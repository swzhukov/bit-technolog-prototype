"""
Импорт техкарт в разных форматах: Excel, PDF, JSON, Word.
Каждый импортёр возвращает dict с полями для insert в БД.
"""
import json
import re
import os
from typing import Optional


def import_from_json(data: dict) -> list[dict]:
    """Импорт из JSON. Ожидаемый формат:
    {
      "details": [
        {
          "id": "d-001",
          "designation": "ЛМША.301314.010",
          "name": "Упор продольный",
          "model": "АЦ-6,0-40",
          "chassis": "КАМАЗ-43118",
          "material": "Сталь 09Г2С",
          "mass_kg": 12.5,
          "surface_treatment": "Грунт ГФ-021, эмаль ПФ-115",
          "parent_id": null,
          "level": "detail",
          "operations": [
            {
              "name": "010 Подготовительная",
              "equipment": "Кедр-300",
              "duration_hours": 0.2,
              "department": "Цех 01",
              "workplace": "Участок 01",
              "profession_code": "19905",
              "profession_grade": 4,
              "materials": [
                {"name": "Проволока Св-08Г2С-О 1,0", "quantity": 0.5, "unit": "кг", "gost": "ГОСТ 2246-70"},
                {"name": "Смесь М21", "quantity": 0.02, "unit": "м3", "gost": "ГОСТ Р ИСО 14175-2010"}
              ]
            }
          ]
        }
      ]
    }
    """
    if isinstance(data, list):
        details = data
    elif "details" in data:
        details = data["details"]
    else:
        details = [data]
    return [_normalize_detail(d) for d in details]


def import_from_excel(filepath: str) -> list[dict]:
    """Импорт из Excel. Ожидаем структуру:
    - Лист 'Детали': колонки id, designation, name, model, chassis, material, mass_kg, ...
    - Лист 'Операции': detail_id, op_index, name, equipment, duration_hours, ...
    Или один лист с разделителем по detail_id.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    details_by_id = {}
    # Если есть лист 'Детали'
    if "Детали" in wb.sheetnames or "details" in wb.sheetnames:
        sn = "Детали" if "Детали" in wb.sheetnames else "details"
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            d = {
                "id": str(row[0]).strip(),
                "designation": str(row[1] or "").strip(),
                "name": str(row[2] or "").strip() if len(row) > 2 else "",
                "model": str(row[3] or "").strip() if len(row) > 3 else "",
                "chassis": str(row[4] or "").strip() if len(row) > 4 else "",
                "material": str(row[5] or "").strip() if len(row) > 5 else "",
                "mass_kg": float(row[6]) if len(row) > 6 and row[6] else 0,
                "operations": []
            }
            details_by_id[d["id"]] = d
    # Если есть лист 'Операции'
    if "Операции" in wb.sheetnames or "operations" in wb.sheetnames:
        sn = "Операции" if "Операции" in wb.sheetnames else "operations"
        ws = wb[sn]
        headers = [str(c.value or "").strip() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            op = dict(zip(headers, row))
            detail_id = str(op.get("detail_id") or "").strip()
            if not detail_id or detail_id not in details_by_id:
                continue
            details_by_id[detail_id]["operations"].append({
                "name": str(op.get("name") or "").strip(),
                "equipment": str(op.get("equipment") or "").strip(),
                "duration_hours": float(op.get("duration_hours") or 0),
                "department": str(op.get("department") or "").strip(),
                "workplace": str(op.get("workplace") or "").strip(),
                "profession_code": str(op.get("profession_code") or "").strip(),
                "profession_grade": int(op.get("profession_grade") or 0) if op.get("profession_grade") else 0,
            })
    return [_normalize_detail(d) for d in details_by_id.values()]


def import_from_pdf(filepath: str) -> list[dict]:
    """Импорт из PDF (ГОСТ 3.1105-2011 Форма 2). Эвристический парсинг:
    - Извлекаем текст
    - Ищем паттерны: 'ЛМША.XXXXXX.XXX' (обозначение)
    - Ищем операции '010 ...', '015 ...', '020 ...'
    - Ищем оборудование 'Кедр-300', 'Станок ...'
    """
    import pdfplumber
    details_by_id = {}
    with pdfplumber.open(filepath) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            try:
                text = page.extract_text() or ""
            except Exception as e:
                # E fix: N5 — пропускаем страницу если extract_text упал
                continue
            if not text:
                continue
            # Ищем обозначение
            m = re.search(r"([А-Я]{2,5}\.\d{6}\.\d{3}(?:/\d{3})?)", text)
            designation = m.group(1) if m else f"pdf-{len(details_by_id)+1:03d}"
            # Ищем название
            name = ""
            nm = re.search(r"Упор[^\n]+|Кронштейн[^\n]+|Рама[^\n]+|Цилиндр[^\n]+|Опора[^\n]+|Вал[^\n]+", text)
            if nm:
                name = nm.group(0)[:80]
            # Ищем операции
            ops = []
            for line in text.split("\n"):
                m = re.match(r"\s*(\d{3})\s+([А-ЯA-Z][^\d]{5,60})", line)
                if m:
                    op_name = f"{m.group(1)} {m.group(2).strip()}"
                    tm = re.search(r"(\d+[.,]\d+)\s*ч", line)
                    duration = float(tm.group(1).replace(",", ".")) if tm else 0
                    eq = ""
                    em = re.search(r"Кедр-\d+|станок\s+\w+|установка\s+\w+", line, re.IGNORECASE)
                    if em:
                        eq = em.group(0)
                    ops.append({
                        "name": op_name,
                        "equipment": eq,
                        "duration_hours": duration
                    })
            if not ops and not designation:
                continue
            details_by_id[designation] = {
                "id": designation.replace(".", "-").replace("/", "-").lower(),
                "designation": designation,
                "name": name or designation,
                "model": "",
                "chassis": "",
                "material": "",
                "mass_kg": 0,
                "operations": ops
            }
    return [_normalize_detail(d) for d in details_by_id.values()]


def import_from_word(filepath: str) -> list[dict]:
    """Импорт из Word (.docx). Эвристический парсинг параграфов."""
    import docx
    doc = docx.Document(filepath)
    text = "\n".join(p.text for p in doc.paragraphs)
    # Используем те же паттерны, что и для PDF
    details = import_from_pdf_text(text)
    return details


def import_from_pdf_text(text: str) -> list[dict]:
    """Парсинг текста техкарты (используется и из PDF, и из Word)."""
    details_by_id = {}
    m = re.search(r"([А-Я]{2,5}\.\d{6}\.\d{3}(?:/\d{3})?)", text)
    designation = m.group(1) if m else f"text-{len(details_by_id)+1:03d}"
    name = ""
    nm = re.search(r"Упор[^\n]+|Кронштейн[^\n]+|Рама[^\n]+|Цилиндр[^\n]+|Опора[^\n]+|Вал[^\n]+", text)
    if nm:
        name = nm.group(0)[:80]
    ops = []
    for line in text.split("\n"):
        m = re.match(r"\s*(\d{3})\s+([А-ЯA-Z][^\d]{5,60})", line)
        if m:
            op_name = f"{m.group(1)} {m.group(2).strip()}"
            tm = re.search(r"(\d+[.,]\d+)\s*ч", line)
            duration = float(tm.group(1).replace(",", ".")) if tm else 0
            eq = ""
            em = re.search(r"Кедр-\d+|станок\s+\w+|установка\s+\w+", line, re.IGNORECASE)
            if em:
                eq = em.group(0)
            ops.append({
                "name": op_name,
                "equipment": eq,
                "duration_hours": duration
            })
    if ops or designation != f"text-{len(details_by_id)+1:03d}":
        details_by_id[designation] = {
            "id": designation.replace(".", "-").replace("/", "-").lower(),
            "designation": designation,
            "name": name or designation,
            "model": "", "chassis": "", "material": "", "mass_kg": 0,
            "operations": ops
        }
    return [_normalize_detail(d) for d in details_by_id.values()]


def _normalize_detail(d: dict) -> dict:
    """Нормализация полей детали (разные источники — разные форматы)"""
    return {
        "id": str(d.get("id") or "").strip(),
        "designation": str(d.get("designation") or d.get("name") or "UNKNOWN").strip(),
        "name": str(d.get("name") or d.get("designation") or "").strip(),
        "model": str(d.get("model") or "").strip(),
        "chassis": str(d.get("chassis") or "").strip(),
        "material": str(d.get("material") or "").strip(),
        "size_mm": str(d.get("size_mm") or "").strip(),
        "mass_kg": float(d.get("mass_kg") or 0),
        "surface_treatment": str(d.get("surface_treatment") or "").strip(),
        "extra_props": json.dumps(d.get("extra_props") or {}, ensure_ascii=False),
        "parent_id": d.get("parent_id"),
        "level": d.get("level", "detail") if d.get("level") in ("detail", "assembly", "product") else "detail",
        "drawing_path": d.get("drawing_path"),
        "drawing_format": d.get("drawing_format"),
        "operations": d.get("operations") or []
    }


def save_imported_details(details: list[dict], default_author: str = "import") -> dict:
    """Сохраняет распарсенные детали в БД. Возвращает статистику.
    C8 fix: дедупликация по designation (если повтор — обновляем, а не дублируем).
    V7-8: валидация обязательных полей (designation, name)."""
    from app import get_conn
    conn = get_conn()
    # V7-8: валидация
    valid_details = []
    validation_errors = []
    for d in details:
        des = (d.get("designation") or "").strip()
        if not des:
            validation_errors.append({"detail": d.get("id", "?"), "error": "designation is empty"})
            continue
        if len(des) > 200:
            validation_errors.append({"detail": d.get("id", "?"), "error": "designation too long (max 200)"})
            continue
        # Нормализуем
        d["designation"] = des
        if d.get("name"):
            d["name"] = str(d["name"]).strip()[:500]
        # mass_kg: должен быть числом или None
        if "mass_kg" in d and d["mass_kg"] is not None:
            try:
                d["mass_kg"] = float(d["mass_kg"])
            except (TypeError, ValueError):
                validation_errors.append({"detail": d.get("id", "?"), "error": f"mass_kg not a number: {d['mass_kg']}"})
                continue
        valid_details.append(d)
    details = valid_details
    # Дедупликация по designation
    seen_designations = set()
    deduped = []
    for d in details:
        des = d.get("designation", "")
        if not des or des in seen_designations:
            continue
        seen_designations.add(des)
        deduped.append(d)
    details = deduped
    created = 0
    updated = 0
    ops_saved = 0
    seeded_ids = []
    validation_count = len(validation_errors)
    for d in details:
        if not d.get("id"):
            continue
        existing = conn.execute("SELECT id FROM details WHERE id=?", (d["id"],)).fetchone()
        if existing:
            conn.execute("""UPDATE details SET
                designation=?, name=?, model=?, chassis=?, material=?, size_mm=?,
                mass_kg=?, surface_treatment=?, extra_props=?, parent_id=?, level=?,
                drawing_path=?, drawing_format=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?""", (
                d["designation"], d["name"], d["model"], d["chassis"], d["material"],
                d["size_mm"], d["mass_kg"], d["surface_treatment"], d["extra_props"],
                d["parent_id"], d["level"], d["drawing_path"], d["drawing_format"], d["id"]
            ))
            updated += 1
        else:
            conn.execute("""INSERT INTO details
                (id, designation, name, model, chassis, material, size_mm, mass_kg,
                 surface_treatment, extra_props, parent_id, level, drawing_path, drawing_format)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""", (
                d["id"], d["designation"], d["name"], d["model"], d["chassis"], d["material"],
                d["size_mm"], d["mass_kg"], d["surface_treatment"], d["extra_props"],
                d["parent_id"], d["level"], d["drawing_path"], d["drawing_format"]
            ))
            created += 1
        # Сохраняем операции как draft
        if d.get("operations"):
            llm_output = {
                "summary": {
                    "total_operations": len(d["operations"]),
                    "total_hours": sum(float(o.get("duration_hours") or 0) for o in d["operations"]),
                    "prep_hours": 0,
                    "complexity": "medium"
                },
                "operations": [
                    {**op, "confidence": op.get("confidence", 85) if op.get("confidence") is not None else 85}
                    for op in d["operations"]
                ],
                "reasoning": {
                    "operations_choice": f"Импортировано из файла ({default_author})",
                    "duration_estimates": "Из источника",
                    "equipment_choice": "Из источника",
                    "risks": "Требует верификации технологом"
                },
                "source": "import"
            }
            existing_draft = conn.execute("SELECT detail_id FROM drafts WHERE detail_id=?", (d["id"],)).fetchone()
            if existing_draft:
                conn.execute("UPDATE drafts SET llm_output=?, updated_at=CURRENT_TIMESTAMP WHERE detail_id=?",
                             (json.dumps(llm_output, ensure_ascii=False), d["id"]))
            else:
                conn.execute("""INSERT INTO drafts (detail_id, llm_output, status, author)
                    VALUES (?, ?, 'draft', ?)""", (d["id"], json.dumps(llm_output, ensure_ascii=False), default_author))
            ops_saved += len(d["operations"])
            # Сохраняем ресурсы в resource_specs
            for i, op in enumerate(d["operations"]):
                if op.get("profession_code"):
                    conn.execute("""INSERT INTO resource_specs
                        (detail_id, op_index, kind, name, quantity, unit, notes)
                        VALUES (?, ?, 'profession', ?, 1, 'чел', ?)""",
                        (d["id"], i, f"{op.get('profession_code')} {op.get('profession_grade', '')}р",
                         f"Ставка: {op.get('hourly_rate', '—')} ₽/ч"))
                for m in op.get("materials") or []:
                    if isinstance(m, dict):
                        conn.execute("""INSERT INTO resource_specs
                            (detail_id, op_index, kind, name, quantity, unit, notes)
                            VALUES (?, ?, 'material', ?, ?, ?, ?)""",
                            (d["id"], i, m.get("name", ""), m.get("quantity", 1), m.get("unit", ""), m.get("gost", "")))
        seeded_ids.append(d["id"])
    conn.commit()
    conn.close()
    # История после commit
    from app import add_history
    for did in seeded_ids:
        try:
            add_history(did, "imported", {"author": default_author})
        except Exception:
            pass
    return {"created": created, "updated": updated, "operations_saved": ops_saved, "total": len(details),
            "validation_count": validation_count, "validation_errors": validation_errors[:10]}


# F-12 fix: проверка magic bytes для предотвращения загрузки .exe переименованных в .pdf
MAGIC_BYTES = {
    # Office / docs
    "xlsx": [b"PK\x03\x04"],  # ZIP-сигнатура (xlsx/docx оба ZIP)
    "docx": [b"PK\x03\x04"],
    "pdf": [b"%PDF"],
    # Images
    "png": [b"\x89PNG\r\n\x1a\n"],
    "jpg": [b"\xff\xd8\xff"],
    "jpeg": [b"\xff\xd8\xff"],
    "svg": [b"<?xml", b"<svg"],
    # CAD (нет надёжной сигнатуры — принимаем любой бинарь)
    "frw": [],  # КОМПАС-3D
    "dwg": [b"AC10"],  # AutoCAD 2010+
    # JSON можно импортировать как text
    "json": [b"{", b"["],
}


def verify_magic_bytes(contents: bytes, suffix: str) -> bool:
    """Проверяет что содержимое файла соответствует расширению.
    Защита от .exe переименованных в .pdf.
    Возвращает True если magic bytes совпадают ИЛИ если для формата нет ожидаемой сигнатуры.
    """
    if not contents:
        return False
    suffix = suffix.lower()
    expected = MAGIC_BYTES.get(suffix, None)
    # Если для формата нет сигнатуры (frw) — пропускаем проверку
    if expected is None or len(expected) == 0:
        return True
    return any(contents.startswith(sig) for sig in expected)
