"""Extract text from all PDFs in /workspace/attachments."""
import os
import pdfplumber

ATTACH = '/workspace/attachments'
OUT = '/workspace/audit/extracted_text'
os.makedirs(OUT, exist_ok=True)

results = []
for f in sorted(os.listdir(ATTACH)):
    if not f.endswith('.pdf'):
        continue
    path = os.path.join(ATTACH, f)
    name = f.split('__')[1] if '__' in f else f
    out_path = os.path.join(OUT, name + '.txt')
    try:
        with pdfplumber.open(path) as pdf:
            text = ''
            for i, page in enumerate(pdf.pages):
                pt = page.extract_text() or ''
                text += f'\n=== Page {i+1} ===\n' + pt
            with open(out_path, 'w') as out:
                out.write(text)
            results.append((name, len(pdf.pages), len(text), out_path))
            print(f"  {name}: {len(pdf.pages)} pages, {len(text)} chars")
    except Exception as e:
        print(f"  {name}: ERROR {e}")
        results.append((name, 0, 0, str(e)))

# Excel
import openpyxl
for f in sorted(os.listdir(ATTACH)):
    if not f.endswith('.xlsx'):
        continue
    path = os.path.join(ATTACH, f)
    name = f.split('__')[1] if '__' in f else f
    out_path = os.path.join(OUT, name + '.xlsx.txt')
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        out = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            out.append(f'\n=== Sheet: {sheet} ===')
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    out.append(' | '.join(str(c) if c is not None else '' for c in row))
        with open(out_path, 'w') as f_out:
            f_out.write('\n'.join(out))
        print(f"  {name}: {len(wb.sheetnames)} sheets")
    except Exception as e:
        print(f"  {name}: ERROR {e}")

print(f"\nTotal: {len(results)} PDFs processed")
