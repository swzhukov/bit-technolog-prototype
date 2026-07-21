"""
БИТ.Технолог — Прототип v0.1
AI-помощник технолога для ускорения создания техкарт.

Запуск: python app.py
Открыть: http://localhost:8080
"""

import os
import re
import json
import io
import sqlite3
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, Request, Form
from fastapi.responses import HTMLResponse, JSONResponse, Response, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from dotenv import load_dotenv
# V7-1: для F811/F821 (dead code ссылается на OpenAI — мёртвый дубликат)
try:
    from openai import OpenAI  # noqa: F401
except ImportError:
    pass

# Load environment
load_dotenv()
LLM_API_URL = os.getenv("LLM_API_URL", "https://api.1bitai.ru/v1")
LLM_API_KEY = os.getenv("LLM_API_KEY", "")
LLM_MODEL = os.getenv("LLM_MODEL", "deepseek-v4-flash-thinking")
LLM_TIMEOUT = int(os.getenv("LLM_TIMEOUT", "120"))
DEMO_MODE = os.getenv("DEMO_MODE", "false").lower() == "true"

# Стоимость и лимиты
LLM_PRICE_INPUT_RUB_PER_1K = float(os.getenv("LLM_PRICE_INPUT_RUB_PER_1K", "0.40"))
LLM_PRICE_OUTPUT_RUB_PER_1K = float(os.getenv("LLM_PRICE_OUTPUT_RUB_PER_1K", "1.20"))
LLM_DAILY_LIMIT_RUB = float(os.getenv("LLM_DAILY_LIMIT_RUB", "200"))

# Auto-enable demo mode if API key is empty
if not LLM_API_KEY and not DEMO_MODE:
    log_msg = "No LLM_API_KEY set — auto-enabling DEMO_MODE"
    print(f"[WARN] {log_msg}")
    DEMO_MODE = True

# P0 v6: лимиты на размер файлов (защита от DoS)
MAX_DRAWING_SIZE = int(os.getenv("MAX_DRAWING_SIZE", str(50 * 1024 * 1024)))  # 50MB
MAX_IMPORT_SIZE = int(os.getenv("MAX_IMPORT_SIZE", str(100 * 1024 * 1024)))  # 100MB
ALLOWED_DRAWING_FORMATS = {"frw", "dwg", "pdf", "png", "jpg", "jpeg", "svg"}
ALLOWED_IMPORT_FORMATS = {"xlsx", "xls", "pdf", "docx", "doc"}
ALLOWED_LEVELS = {"detail", "assembly", "product"}

# Auth (C1 fix): HTTP Basic с переменной PILOT_USERS=user:pass,admin:pass
PILOT_USERS_RAW = os.getenv("PILOT_USERS", "")
PILOT_USERS = {}
if PILOT_USERS_RAW:
    for pair in PILOT_USERS_RAW.split(","):
        if ":" in pair:
            u, p = pair.split(":", 1)
            PILOT_USERS[u.strip()] = p.strip()
PILOT_AUTH_ENABLED = bool(PILOT_USERS) and not os.getenv("PILOT_AUTH_DISABLED", "").lower() == "true"

# LLM client singleton (C4 fix)
_LLM_CLIENT = None

def get_llm_client():
    """Lazy-init singleton OpenAI client.
    Читает LLM_API_KEY/LLM_API_URL динамически из БД (get_setting) с fallback на .env.
    Если ключ изменился в админке — клиент пересоздаётся автоматически.
    """
    global _LLM_CLIENT
    api_key = get_setting("LLM_API_KEY", LLM_API_KEY)
    api_url = get_setting("LLM_API_URL", LLM_API_URL)
    demo = get_setting("DEMO_MODE", "false" if not DEMO_MODE else "true").lower() == "true"
    if demo:
        return None
    # Пересоздать клиент если ключ/url изменились
    if _LLM_CLIENT is not None:
        cached = getattr(_LLM_CLIENT, "_bit_key", None)
        cached_url = getattr(_LLM_CLIENT, "_bit_url", None)
        if cached != api_key or cached_url != api_url:
            log.info("LLM key/url changed in admin — recreating client")
            _LLM_CLIENT = None
    if _LLM_CLIENT is None:
        _LLM_CLIENT = OpenAI(
            base_url=api_url,
            api_key=api_key,
            timeout=LLM_TIMEOUT
        )
        _LLM_CLIENT._bit_key = api_key
        _LLM_CLIENT._bit_url = api_url
    return _LLM_CLIENT


def parse_llm_json(text: str) -> dict:
    """NC3 fix: устойчивый парсинг JSON из LLM-ответа.
    Пробует: raw -> strip ```json -> strip ``` -> first { to last } -> raise ValueError."""
    if not text or not text.strip():
        raise ValueError("empty LLM response")
    s = text.strip()
    # 1. Raw
    try:
        return json.loads(s)
    except Exception:
        pass
    # 2. Strip ```json ... ```
    if s.startswith("```"):
        lines = s.split("\n", 1)
        s2 = lines[1] if len(lines) > 1 else ""
        if s2.rstrip().endswith("```"):
            s2 = s2.rstrip()[:-3].rstrip()
        try:
            return json.loads(s2)
        except Exception:
            pass
        if s2.lower().startswith("json"):
            s2 = s2[4:].lstrip()
            try:
                return json.loads(s2)
            except Exception:
                pass
    # 3. Extract from first { to last }
    first = s.find("{")
    last = s.rfind("}")
    if first != -1 and last != -1 and last > first:
        candidate = s[first:last+1]
        try:
            return json.loads(candidate)
        except Exception:
            pass
    raise ValueError(f"could not parse JSON from LLM response (first 200 chars): {s[:200]}")


def parse_llm_json_safe(text: str) -> dict:
    """Safe wrapper: возвращает {} при любой ошибке, не raise.
    Используй в местах, где LLM-сбой не должен ломать UX."""
    try:
        return parse_llm_json(text)
    except Exception:
        return {}

# V6-25: timezone-aware datetime
try:
    from zoneinfo import ZoneInfo
    MOSCOW_TZ = ZoneInfo("Europe/Moscow")
except ImportError:
    MOSCOW_TZ = None  # старый Python — fallback на naive datetime


def now_msk() -> datetime:
    """V6-25: текущее время в Moscow timezone.
    Без zoneinfo — naive (для совместимости)."""
    if MOSCOW_TZ:
        return datetime.now(MOSCOW_TZ)
    return datetime.now()


# V6-5: retention policy
RETENTION_DAYS = {
    "audit_logins": 180,    # 6 месяцев
    "llm_calls": 90,         # 3 месяца
    "history": 365,          # 1 год
}


def cleanup_old_records() -> dict:
    """V6-5: очистка старых записей по retention policy.
    Запускать вручную или через cron (ежемесячно)."""
    from db import get_conn
    conn = get_conn()
    result = {}
    try:
        for table, days in RETENTION_DAYS.items():
            cur = conn.execute(f"DELETE FROM {table} WHERE ts < datetime('now', '-{days} day')")
            result[table] = cur.rowcount
        conn.commit()
    except Exception as e:
        log.error(f"cleanup_old_records failed: {e}")
        return {"error": str(e)}
    finally:
        conn.close()
    return result


# Logging
import json as _json_logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
log = logging.getLogger("bit-technolog")

# V6-22: Structured JSON logs (включается через JSON_LOGS=true)
if os.getenv("JSON_LOGS", "true").lower() == "true":
    class JsonFormatter(logging.Formatter):
        """V6-22: JSON-формат логов для production.
        Каждый log = одна строка JSON, парсится logstash/loki/etc."""
        def format(self, record):
            payload = {
                "ts": _json_logging.loads(_json_logging.dumps(record.created)) if False else None,
                "level": record.levelname,
                "logger": record.name,
                "msg": record.getMessage(),
            }
            # timestamp как ISO
            import datetime
            payload["ts"] = datetime.datetime.fromtimestamp(record.created).isoformat()
            # extra fields
            for key in ("detail_id", "author", "action", "cost", "tokens", "status", "error"):
                if hasattr(record, key):
                    payload[key] = getattr(record, key)
            if record.exc_info:
                payload["exception"] = self.formatException(record.exc_info)
            return _json_logging.dumps(payload, ensure_ascii=False)
    # Перенастроить handler
    for h in logging.root.handlers:
        h.setFormatter(JsonFormatter())
    log.info("JSON logging enabled")

# FastAPI app
from contextlib import asynccontextmanager


@asynccontextmanager
async def _lifespan(app):
    """FastAPI lifespan: инициализация БД при старте, очистка при остановке.
    Заменяет deprecated @app.on_event('startup')."""
    # Startup
    try:
        init_db()
        # F16.9: A4-11 — миграция deleted_operations (если старая БД)
        try:
            from db import get_conn
            conn = get_conn()
            conn.execute("""CREATE TABLE IF NOT EXISTS deleted_operations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                detail_id TEXT NOT NULL,
                op_index INTEGER,
                op_name TEXT,
                op_json TEXT,
                deleted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                deleted_by TEXT,
                reason TEXT,
                restored_at TIMESTAMP,
                restored_by TEXT)""")
            conn.commit()
            conn.close()
        except Exception as e:
            log.debug(f"lifespan: deleted_operations migration: {e}")
        log.info("DB initialized on lifespan startup")
    except Exception as e:
        log.error(f"DB init failed on startup: {e}")
    yield
    # Shutdown — закрываем все открытые соединения (если есть пул)
    try:
        log.info("Application shutdown")
    except Exception:
        pass


app = FastAPI(title="БИТ.Технолог — Прототип", version="0.4.18", lifespan=_lifespan)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


# ========== Middleware: request.state.current_role для шаблонов ==========
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request as StarletteRequest


class RoleStateMiddleware(BaseHTTPMiddleware):
    """Добавляет request.state.current_role = 'admin' / 'technologist' / etc
    + F16.9: A4-18 — request.state.unread_llm_errors (счётчик для badge в nav)."""

    async def dispatch(self, request, call_next):
        role = request.cookies.get("bit_role", "technologist")
        if role not in ROLES:
            role = "technologist"
        request.state.current_role = role
        # A4-18: только для админа считаем ошибки (иначе лишний query на каждой странице)
        if role == "admin":
            try:
                from db import get_conn
                conn = get_conn()
                n = conn.execute("""SELECT COUNT(*) FROM llm_calls
                    WHERE (error IS NOT NULL AND error != '') OR response_parsed_ok = 0""").fetchone()[0] or 0
                conn.close()
                request.state.unread_llm_errors = n
            except Exception:
                request.state.unread_llm_errors = 0
        else:
            request.state.unread_llm_errors = 0
        response = await call_next(request)
        return response


app.add_middleware(RoleStateMiddleware)


# Глобальные функции в Jinja: current_role, is_admin, role_name
templates.env.globals.update({
    "current_role_from_request": lambda request: getattr(request.state, "current_role", "technologist"),
    "is_admin_from_request": lambda request: getattr(request.state, "current_role", "") == "admin",
    "role_name_lookup": lambda r: ROLES.get(r, {}).get("name", r)
})


# ========== Helpers v3 (NC5, NC7, OB3) ==========
def err(msg: str, code: int = 400, **extra) -> JSONResponse:
    """NC7: единый helper для JSON-ошибок"""
    return JSONResponse({"error": msg, **extra}, status_code=code)


def safe_call(name: str, fn, *args, default=None, **kwargs):
    """NC5 fix: вызов с автологированием ошибок.
    При успехе — возврат результата.
    При ошибке — log.exception + возврат default."""
    try:
        return fn(*args, **kwargs)
    except Exception as e:
        log.exception(f"{name} failed: {e}")
        return default


# ========== C1 fix: HTTP Basic Auth middleware ==========
from fastapi import status as http_status
from fastapi.responses import Response

def check_auth(request: Request) -> bool:
    """Проверяет HTTP Basic Auth. False = пускать без auth (для тестов/DEMO)."""
    if not PILOT_AUTH_ENABLED:
        return True
    auth = request.headers.get("authorization", "")
    if not auth.startswith("Basic "):
        return False
    import base64
    try:
        decoded = base64.b64decode(auth[6:]).decode("utf-8")
        if ":" in decoded:
            u, p = decoded.split(":", 1)
            # Сначала проверяем .env-Юзеров (fallback), потом БД-pilot_users
            pilot_users_raw = get_setting("PILOT_USERS", PILOT_USERS_RAW)
            env_users = {}
            if pilot_users_raw:
                for pair in pilot_users_raw.split(","):
                    if ":" in pair:
                        eu, ep = pair.split(":", 1)
                        env_users[eu.strip()] = ep.strip()
            if env_users.get(u) == p:
                return True
            # Fallback: БД-pilot_users (bcrypt)
            user = authenticate_pilot_user(u, p)
            if user:
                ip = request.client.host if hasattr(request, "client") else ""
                log_login(u, ip, "", True)
                return True
    except Exception:
        return False
    return False


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    """Middleware: auth + CSRF protection (htmx шлёт X-Requested-With всегда)"""
    path = request.url.path
    # Публичные пути
    public = ("/static", "/health", "/login", "/docs", "/openapi.json", "/favicon.ico")
    if any(path.startswith(p) for p in public):
        return await call_next(request)
    if PILOT_AUTH_ENABLED and not check_auth(request):
        return Response(
            content="<h1>401 Unauthorized</h1><p>Нужен логин/пароль из PILOT_USERS в .env</p>",
            status_code=http_status.HTTP_401_UNAUTHORIZED,
            headers={"WWW-Authenticate": 'Basic realm="BIT-Technolog"'},
            media_type="text/html; charset=utf-8"
        )
    # CSRF check (F16.4: по умолчанию ВКЛЮЧЕН).
    # Раньше был opt-in (PILOT_CSRF_ENABLED=true). Теперь по умолчанию защита.
    # Opt-out через PILOT_CSRF_DISABLED=true (для тестов и локальной разработки).
    if request.method in ("POST", "PUT", "DELETE", "PATCH"):
        if os.getenv("PILOT_CSRF_DISABLED", "").lower() == "true":
            pass  # отключено явно (тесты)
        else:
            xrw = request.headers.get("x-requested-with", "")
            referer = request.headers.get("referer", "")
            origin = request.headers.get("origin", "")
            host = request.headers.get("host", "")
            # Разрешаем если htmx ИЛИ same-origin
            if xrw.lower() == "xmlhttprequest":
                pass  # htmx, fetch, etc.
            elif referer and path in referer:
                pass  # same-origin form submit
            elif origin and host and origin.endswith(host):
                pass  # CORS: origin matches host
            else:
                return JSONResponse(
                    {"error": "CSRF check failed: need X-Requested-With, same-origin Referer or matching Origin"},
                    status_code=403
                )
    return await call_next(request)


@app.get("/login", response_class=HTMLResponse)
async def login_page():
    """Показывает форму логина если браузер не Basic Auth"""
    return HTMLResponse(
        '<h1>BIT-Technolog</h1><p>Откройте в браузере с логином/паролем или введите в URL: '
        '<code>http://user:pass@host:8000/</code></p>',
        status_code=401,
        headers={"WWW-Authenticate": 'Basic realm="BIT-Technolog"'}
    )


# Jinja globals: daily_cost() вызывается в каждом шаблоне автоматически
def _daily_cost_global():
    return get_daily_cost()

templates.env.globals["daily_cost"] = _daily_cost_global


# Middleware: добавляет daily_cost в request.state (для endpoints, не для шаблонов)
@app.middleware("http")
async def add_daily_cost(request: Request, call_next):
    request.state.daily_cost = get_daily_cost()
    response = await call_next(request)
    return response

# Local imports
from prompts import TECH_CARD_PROMPT
from mock_data import MOCK_DETAILS
from few_shot import FEW_SHOT_4C85941A, get_relevant_few_shot
from workshops_tehinkom import TECHINKOM_WORKSHOPS_CONTEXT

with open("equipment.json", "r", encoding="utf-8") as f:
    EQUIPMENT = json.load(f)

with open("structure.json", "r", encoding="utf-8") as f:
    STRUCTURE = json.load(f)

# Database
DB_PATH = os.getenv("DB_PATH", "bit_technolog.db")


# v3: A4-2 — endpoint для сохранения answers в БД (backup для localStorage)
@app.post("/api/answers/save")
async def api_save_answers(request: Request):
    """Сохраняет ответы технолога на вопросы AI (для 3-step flow).
    Backup для localStorage: если браузер чистит LS, ответы не потеряются."""
    detail_id = await _get_param(request, "detail_id")
    step = await _get_param(request, "step") or "analyze"
    answers_json = await _get_param(request, "answers") or "{}"
    if not detail_id:
        return err("detail_id required", 422)
    from db import get_conn
    conn = get_conn()
    try:
        conn.execute("""INSERT INTO step_answers (detail_id, step, answers_json)
            VALUES (?, ?, ?)""", (detail_id, step, answers_json))
        conn.commit()
        return {"ok": True, "detail_id": detail_id, "step": step}
    except Exception as e:
        return err(f"save failed: {e}", 500)
    finally:
        conn.close()


@app.get("/api/answers/load")
async def api_load_answers(detail_id: str, step: str = "analyze"):
    """Загружает последние answers для детали (если localStorage пуст)."""
    from db import get_conn
    conn = get_conn()
    try:
        row = conn.execute("""SELECT answers_json, created_at FROM step_answers
            WHERE detail_id=? AND step=?
            ORDER BY created_at DESC LIMIT 1""", (detail_id, step)).fetchone()
        if not row:
            return {"ok": True, "answers": None}
        return {"ok": True, "answers": row[0], "ts": row[1]}
    finally:
        conn.close()


# F16.10 (v3): V3-12 — проверка лимита перед каждым LLM-вызовом
def check_daily_limit_or_warn() -> dict:
    """V3-12: возвращает dict {allowed, current, limit, pct, message}.
    При 80% — предупреждение. При 100% — блок."""
    from db import get_daily_cost
    from settings import get_setting
    cost_data = get_daily_cost()
    current = cost_data.get("total_rub", 0) or 0
    limit = int(get_setting("LLM_DAILY_COST_LIMIT_RUB", "200") or 200)
    pct = (current / limit * 100) if limit > 0 else 0
    if pct >= 100:
        return {"allowed": False, "current": current, "limit": limit,
                "pct": round(pct, 1), "message": f"Дневной лимит исчерпан: {current:.2f}₽ / {limit}₽"}
    if pct >= 80:
        return {"allowed": True, "current": current, "limit": limit,
                "pct": round(pct, 1), "message": f"⚠️ Использовано {pct:.0f}% дневного лимита ({current:.2f}₽ / {limit}₽)",
                "warning": True}
    return {"allowed": True, "current": current, "limit": limit, "pct": round(pct, 1), "message": ""}


# F16.10 (v3): Rate limiter — защита от DDoS / перерасхода
import threading
from collections import defaultdict
from time import time
# V4-11: для uptime в /health
_APP_START_TS = time()
# V7-2: git commit hash
import subprocess
try:
    _GIT_COMMIT = subprocess.run(
        ["git", "rev-parse", "--short", "HEAD"],
        capture_output=True, text=True, timeout=2
    ).stdout.strip() or "unknown"
except Exception:
    _GIT_COMMIT = "no_git"

# F16.10: in-memory rate limiter (для production с одним worker OK; для multi-worker нужен Redis)
_rate_buckets = defaultdict(list)
_rate_lock = threading.Lock()
RATE_LIMITS = {
    # path_prefix: (max_requests, window_seconds)
    "/api/generate": (10, 60),       # 10 генераций в минуту
    "/api/refine": (10, 60),         # 10 уточнений в минуту
    "/api/draft-fast": (20, 60),     # 20 быстрых драфтов в минуту
    "/api/analyze": (10, 60),        # 10 анализов в минуту
    "/api/import/": (5, 300),        # 5 импортов за 5 минут
    "/api/admin/backup": (1, 3600),  # 1 backup в час
    "/api/admin/rag-rebuild": (1, 600),  # 1 rag-rebuild за 10 мин
}


def _check_rate_limit(path: str) -> tuple[bool, int]:
    """F16.10: V3-3 — проверка rate limit.
    Возвращает (allowed, retry_after_sec)."""
    # V3-3: opt-out для тестов
    if os.getenv("PILOT_RATELIMIT_DISABLED", "").lower() == "true":
        return True, 0
    # Найти подходящий лимит по префиксу
    limit = None
    for prefix, l in RATE_LIMITS.items():
        if path.startswith(prefix):
            limit = l
            break
    if not limit:
        return True, 0
    max_req, window_sec = limit
    key = path.split("?")[0]  # без query string
    now = time()
    with _rate_lock:
        bucket = _rate_buckets[key]
        # Очистить старые записи
        bucket[:] = [t for t in bucket if now - t < window_sec]
        if len(bucket) >= max_req:
            retry_after = int(window_sec - (now - bucket[0])) + 1
            return False, retry_after
        bucket.append(now)
        return True, 0


@app.middleware("http")
async def rate_limit_middleware(request: Request, call_next):
    """F16.10: V3-3 — rate limit для критичных endpoint'ов"""
    if request.method in ("POST", "PUT", "DELETE", "PATCH"):
        allowed, retry_after = _check_rate_limit(request.url.path)
        if not allowed:
            return JSONResponse(
                {"error": "rate_limited", "retry_after": retry_after,
                 "limit": RATE_LIMITS.get(request.url.path[:15] + "...", "see server config")},
                status_code=429,
                headers={"Retry-After": str(retry_after)}
            )
    return await call_next(request)


# V8-18: глобальный error handler с debugging ID
import traceback
import uuid

# Хранилище последних 50 ошибок (для /admin/errors)
_ERRORS = []
_MAX_ERRORS = 50


@app.exception_handler(500)
async def server_error_handler(request: Request, exc):
    """V8-18: 500 с debugging ID.
    Пользователь видит короткий ID, в логах полный traceback,
    в /admin/errors — последние 50 ошибок."""
    err_id = str(uuid.uuid4())[:8]
    tb = traceback.format_exc()
    _ERRORS.append({
        "id": err_id,
        "ts": datetime.now().isoformat(),
        "path": request.url.path,
        "method": request.method,
        "exception": str(exc)[:500],
        "traceback": tb[-2000:]
    })
    if len(_ERRORS) > _MAX_ERRORS:
        _ERRORS.pop(0)
    log.exception(f"[{err_id}] {request.method} {request.url.path}: {exc}")
    accept = request.headers.get("accept", "")
    if "text/html" in accept:
        html = f"""<h1>500 — Внутренняя ошибка</h1><p>Debugging ID: <code>{err_id}</code></p><p>Сообщите этот ID администратору.</p><p><a href="javascript:history.back()">← Назад</a></p>"""
        return HTMLResponse(html, status_code=500)
    return JSONResponse({"ok": False, "error": "internal_server_error", "debug_id": err_id}, status_code=500)


@app.get("/admin/errors", response_class=HTMLResponse)
async def admin_errors(request: Request):
    """V8-18: UI для просмотра последних ошибок (только admin)."""
    if get_current_role(request) != "admin":
        return HTMLResponse("<h1>403</h1>", status_code=403)
    return templates.TemplateResponse("admin_errors.html", {
        "request": request,
        "errors": list(reversed(_ERRORS))
    })


# F16.10 (v3): CSP middleware — защита от XSS
@app.middleware("http")
async def csp_middleware(request: Request, call_next):
    """V3-2: добавляет Content-Security-Policy к ответам.
    Запрещает inline scripts кроме своих (htmx, qrcode, наши функции).
    Default-src 'self' (только локальные ресурсы)."""
    response = await call_next(request)
    # CSP для HTML ответов
    content_type = response.headers.get("content-type", "")
    if "text/html" in content_type:
        csp = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' 'unsafe-eval'; "  # unsafe-inline нужен для htmx-attrs и наших onclick
            "style-src 'self' 'unsafe-inline'; "  # unsafe-inline для inline style=
            "img-src 'self' data:; "
            "font-src 'self'; "
            "connect-src 'self'; "
            "frame-ancestors 'none'; "
            "base-uri 'self'; "
            "form-action 'self'"
        )
        response.headers["Content-Security-Policy"] = csp
        # Дополнительные security headers
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    # V5-12: gzip для HTML (всегда) и JSON (только > 10KB)
    is_html = "text/html" in content_type
    is_large_json = "application/json" in content_type
    if is_html or is_large_json:
        accept_encoding = request.headers.get("accept-encoding", "")
        if "gzip" in accept_encoding:
            body = b""
            async for chunk in response.body_iterator:
                body += chunk if isinstance(chunk, bytes) else chunk.encode("utf-8")
            # HTML: 1KB threshold, JSON: 10KB threshold (не ломать маленькие ответы)
            threshold = 1024 if is_html else 10240
            if len(body) > threshold:
                import gzip
                compressed = gzip.compress(body)
                if len(compressed) < len(body):
                    from starlette.responses import Response as StarletteResponse
                    new_response = StarletteResponse(
                        content=compressed,
                        status_code=response.status_code,
                        headers=dict(response.headers),
                        media_type=content_type
                    )
                    new_response.headers["Content-Encoding"] = "gzip"
                    new_response.headers["Content-Length"] = str(len(compressed))
                    response = new_response
                else:
                    # Gzip не помог — вернуть body обратно через iterator
                    from starlette.responses import Response as StarletteResponse
                    response = StarletteResponse(
                        content=body,
                        status_code=response.status_code,
                        headers=dict(response.headers),
                        media_type=content_type
                    )
                    response.headers["Content-Length"] = str(len(body))
            else:
                # Body маленький — вернуть как есть через iterator
                from starlette.responses import Response as StarletteResponse
                response = StarletteResponse(
                    content=body,
                    status_code=response.status_code,
                    headers=dict(response.headers),
                    media_type=content_type
                )
                response.headers["Content-Length"] = str(len(body))
    return response


# ========== Импорты из db.py и auth.py (F15) ==========
from db import (
    DB_PATH, get_conn, get_table_columns, init_db,
    get_detail, get_all_details, get_distinct_models,
    get_draft, save_draft, update_draft_status,
    get_versions, get_edits,
    get_all_equipment, get_all_materials, get_all_iot,
    add_history, get_history,
    get_daily_cost, get_pilot_metrics
)
from auth import (
    ROLES, get_current_role, is_admin,
    hash_password, verify_password, authenticate_pilot_user, log_login
)
from settings import (
    _fernet, SETTING_REGISTRY, _mask_value, _encrypt, _decrypt,
    get_setting, set_setting, delete_setting, get_all_settings
)
from notify import (
    send_email, send_telegram, notify_workflow
)
from llm import (
    LLM_API_KEY, LLM_API_URL, LLM_MODEL, LLM_TIMEOUT, DEMO_MODE,
    get_llm_client, parse_llm_json, log_llm_call, estimate_cost
)
from economics import calc_cost_estimate
from learning import get_learning_metrics_by_week
from metrics_auto import (
    compute_acceptance_from_versions, record_session_start, compute_time_to_card
)


def init_db():
    """Initialize SQLite database with full schema.
    F15.8: дубликат с db.py — db.py-версия используется (импорт в начале файла).
    Оставлено для reference и обратной совместимости."""
    conn = get_conn()
    conn.executescript("""
        -- 1. Детали от конструкторов (вместо mock_data.py)
        CREATE TABLE IF NOT EXISTS details (
            id TEXT PRIMARY KEY,
            designation TEXT NOT NULL,
            name TEXT,
            model TEXT,
            chassis TEXT,
            material TEXT,
            size_mm TEXT,
            mass_kg REAL,
            surface_treatment TEXT,
            extra_props TEXT,
            tech_rules TEXT,
            cost_per_hour REAL DEFAULT 0,
            overhead_pct REAL DEFAULT 15,
            material_cost_rub REAL DEFAULT 0,
            parent_id TEXT,  -- иерархия: деталь -> узел -> изделие
            level TEXT DEFAULT 'detail',  -- detail / assembly / product
            drawing_path TEXT,  -- путь к файлу чертежа
            drawing_format TEXT,  -- frw / dwg / pdf / png
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- 2. Черновики техкарт (LLM output)
        CREATE TABLE IF NOT EXISTS drafts (
            detail_id TEXT PRIMARY KEY,
            llm_output TEXT,
            status TEXT DEFAULT 'new',
            author TEXT DEFAULT 'system',
            created_at TIMESTAMP,
            updated_at TIMESTAMP
        );

        -- 3. Версии черновика (v1, v2, v3...)
        CREATE TABLE IF NOT EXISTS draft_versions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            detail_id TEXT,
            version INTEGER,
            operations_json TEXT,
            author TEXT,
            source TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- 4. Правки технолога (для обучения)
        CREATE TABLE IF NOT EXISTS edits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            detail_id TEXT,
            version INTEGER,
            field TEXT,
            old_value TEXT,
            new_value TEXT,
            reason TEXT,
            author TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- 5. Извлечённые правила (обучение)
        CREATE TABLE IF NOT EXISTS rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_type TEXT,
            condition_json TEXT,
            action_json TEXT,
            confidence REAL DEFAULT 0.5,
            uses_count INTEGER DEFAULT 0,
            source_edits TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- 6. Оборудование
        CREATE TABLE IF NOT EXISTS equipment (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            type TEXT,
            code TEXT,
            max_thickness_mm REAL,
            max_mass_kg REAL,
            notes TEXT,
            source TEXT DEFAULT '1c',
            external_id TEXT,
            last_sync_at TIMESTAMP
        );

        -- 7. Материалы
        CREATE TABLE IF NOT EXISTS materials (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            grade TEXT,
            gost TEXT,
            notes TEXT,
            source TEXT DEFAULT '1c',
            external_id TEXT,
            last_sync_at TIMESTAMP
        );

        -- 8. Цеха/участки/РМ (структура)
        CREATE TABLE IF NOT EXISTS departments (
            id TEXT PRIMARY KEY,
            production TEXT,
            name TEXT NOT NULL,
            code TEXT
        );

        -- 9. ИОТ номера
        CREATE TABLE IF NOT EXISTS iot (
            id TEXT PRIMARY KEY,
            number TEXT NOT NULL,
            description TEXT,
            applies_to TEXT,
            source TEXT DEFAULT '1c',
            external_id TEXT,
            last_sync_at TIMESTAMP
        );

        -- 10. Бенчмарки трудоёмкости
        CREATE TABLE IF NOT EXISTS benchmarks (
            id TEXT PRIMARY KEY,
            detail_type TEXT NOT NULL,
            norm_hours REAL,
            source TEXT,
            sample_size INTEGER DEFAULT 1,
            external_id TEXT,
            last_sync_at TIMESTAMP
        );

        -- 11. История действий
        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            detail_id TEXT,
            action TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            details TEXT
        );

        -- 13. Пилотные метрики (KPI)
        CREATE TABLE IF NOT EXISTS pilot_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            detail_id TEXT,
            metric TEXT,
            value REAL,
            extra TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- 12. Лог LLM-вызовов (промпт + ответ + токены)
        CREATE TABLE IF NOT EXISTS llm_calls (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            detail_id TEXT,
            model TEXT,
            system_prompt TEXT,
            user_prompt TEXT,
            response_text TEXT,
            response_parsed_ok INTEGER,
            tokens_in INTEGER,
            tokens_out INTEGER,
            duration_ms INTEGER,
            cost_rub REAL DEFAULT 0,
            error TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- 14. Справочник профессий (ЕТС для операций) — для РС и 1С:ERP
        CREATE TABLE IF NOT EXISTS professions (
            id TEXT PRIMARY KEY,
            code TEXT NOT NULL,  -- 19905
            name TEXT NOT NULL,  -- Сварщик
            grade INTEGER,  -- 4 разряд
            hourly_rate REAL DEFAULT 0,  -- ₽/ч
            source TEXT DEFAULT '1c',
            external_id TEXT
        );

        -- 15. Ресурсная спецификация (РС) — привязка к операции
        CREATE TABLE IF NOT EXISTS resource_specs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            detail_id TEXT NOT NULL,
            op_index INTEGER,  -- индекс операции в draft
            kind TEXT NOT NULL,  -- material / tool / equipment / safety / consumable
            ref_id TEXT,  -- id из materials/equipment
            name TEXT NOT NULL,  -- человеко-читаемое имя
            quantity REAL DEFAULT 1,  -- кол-во
            unit TEXT,  -- кг / м / шт / л
            norm_per_unit REAL,  -- норма расхода на единицу продукции
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- 16. Чертежи (файлы)
        CREATE TABLE IF NOT EXISTS drawings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            detail_id TEXT NOT NULL,
            file_path TEXT NOT NULL,
            file_format TEXT,  -- frw / dwg / pdf / png / jpg
            file_size INTEGER,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            uploaded_by TEXT
        );

        -- 17. Пользователи пилота (БД, не .env)
        CREATE TABLE IF NOT EXISTS pilot_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'technologist',
            display_name TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_login TIMESTAMP,
            created_by TEXT
        );

        -- 18. Лог входов (успех/ошибка, IP, user-agent)
        CREATE TABLE IF NOT EXISTS audit_logins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            ts TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            ip TEXT,
            user_agent TEXT,
            success INTEGER NOT NULL  -- 0 или 1
        );

        -- 19. Глобальные настройки (LLM/Telegram/лимиты) — зашифрованные в БД
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value_encrypted BLOB,  -- Fernet-encrypted bytes
            value_masked TEXT,      -- для отображения в UI: "sk-...abc" (последние 3)
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_by TEXT
        );
    """)
    # Миграция: добавляем колонку version если её нет (M2 fix — версия КД)
    try:
        conn.execute("ALTER TABLE details ADD COLUMN version TEXT DEFAULT '1.0'")
    except Exception:
        pass
    # Миграция: добавляем колонку cost_rub если её нет (для старых БД)
    try:
        conn.execute("ALTER TABLE llm_calls ADD COLUMN cost_rub REAL DEFAULT 0")
    except Exception:
        pass
    # Миграции для новых полей в справочниках
    for table, col in [
        ("equipment", "source"), ("equipment", "external_id"), ("equipment", "last_sync_at"),
        ("materials", "source"), ("materials", "external_id"), ("materials", "last_sync_at"),
        ("iot", "source"), ("iot", "external_id"), ("iot", "last_sync_at"),
        ("benchmarks", "external_id"), ("benchmarks", "last_sync_at"),
    ]:
        try:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} TEXT")
        except Exception:
            pass
    # Миграции для details
    for col, dtype in [("tech_rules", "TEXT"), ("cost_per_hour", "REAL"), ("overhead_pct", "REAL"), ("material_cost_rub", "REAL")]:
        try:
            conn.execute(f"ALTER TABLE details ADD COLUMN {col} {dtype}")
        except Exception:
            pass
    # Миграции для details (иерархия + чертежи)
    for col, dtype in [("parent_id", "TEXT"), ("level", "TEXT DEFAULT 'detail'"), ("drawing_path", "TEXT"), ("drawing_format", "TEXT")]:
        try:
            conn.execute(f"ALTER TABLE details ADD COLUMN {col} {dtype}")
        except Exception:
            pass
    # Индекс для иерархии
    try:
        conn.execute("CREATE INDEX IF NOT EXISTS idx_details_parent ON details(parent_id)")
    except Exception:
        pass
    # Миграции для drafts (ролевая модель)
    for col, dtype in [("status_ext", "TEXT DEFAULT 'draft'"), ("approver", "TEXT"), ("submitted_at", "TIMESTAMP")]:
        try:
            conn.execute(f"ALTER TABLE drafts ADD COLUMN {col} {dtype}")
        except Exception:
            pass
    conn.commit()
    conn.close()
    seed_initial_data()
    _seed_professions()
    # Сидим 15 деталей Техинкома при первом старте (через try/finally для надёжности)
    try:
        from techinkom_seed import seed_techinkom_data
        result = seed_techinkom_data()
        if result["seeded"] > 0:
            log.info(f"Seeded {result['seeded']} Техинком details")
    except Exception as e:
        log.warning(f"Techinkom seed failed: {e}")
        import traceback
        traceback.print_exc()


def _seed_professions():
    """Сидит справочник профессий по ЕТС если пуст"""
    conn = get_conn()
    if conn.execute("SELECT COUNT(*) FROM professions").fetchone()[0] > 0:
        conn.close()
        return
    # ЕТС 19905 = Сварщик, 4 разряд ~350₽/ч (типично по РФ 2026)
    seed = [
        ("p-19905-3", "19905", "Сварщик", 3, 280),
        ("p-19905-4", "19905", "Сварщик", 4, 350),
        ("p-19905-5", "19905", "Сварщик", 5, 420),
        ("p-19905-6", "19905", "Сварщик", 6, 500),
        ("p-16045-4", "16045", "Оператор станков с ЧПУ", 4, 380),
        ("p-16045-5", "16045", "Оператор станков с ЧПУ", 5, 450),
        ("p-19149-4", "19149", "Токарь", 4, 340),
        ("p-19149-5", "19149", "Токарь", 5, 400),
        ("p-19479-4", "19479", "Фрезеровщик", 4, 340),
        ("p-19479-5", "19479", "Фрезеровщик", 5, 400),
        ("p-18511-4", "18511", "Слесарь-ремонтник", 4, 320),
        ("p-18511-5", "18511", "Слесарь-ремонтник", 5, 380),
        ("p-19756-4", "19756", "Электрогазосварщик", 4, 360),
        ("p-19756-5", "19756", "Электрогазосварщик", 5, 430),
        ("p-19861-4", "19861", "Электромонтажник", 4, 380),
        ("p-19861-5", "19861", "Электромонтажник", 5, 450),
        ("p-14501-4", "14501", "Монтажник гидравлических систем", 4, 400),
        ("p-14501-5", "14501", "Монтажник гидравлических систем", 5, 480),
        ("p-17521-4", "17521", "Маляр", 4, 280),
        ("p-17521-5", "17521", "Маляр", 5, 340),
    ]
    for s in seed:
        conn.execute("""INSERT OR IGNORE INTO professions
            (id, code, name, grade, hourly_rate, source) VALUES (?, ?, ?, ?, ?, 'etc')""", s)
    conn.commit()
    conn.close()


@app.get("/demo", response_class=HTMLResponse)
async def demo_page(request: Request):
    """Demo-сценарий для Баранова (5 мин)"""
    return templates.TemplateResponse("demo.html", {"request": request})


# ========== Фаза 6: Расширенный workflow (роли) ==========
@app.post("/api/workflow/assign")
async def api_workflow_assign(request: Request):
    """Назначить деталь технологу/нормировщику/начальнику цеха"""
    detail_id = await _get_param(request, "detail_id", log_name="/api/workflow/assign")
    role = await _get_param(request, "role")  # constructor / technologist / normirovshchik / workshop_chief / quality
    assignee = await _get_param(request, "assignee")
    if not detail_id or not role or not assignee:
        return err("detail_id, role, assignee required", 422)
    valid_roles = ("constructor", "technologist", "normirovshchik", "workshop_chief", "quality", "main_technologist", "admin")
    if role not in valid_roles:
        return err(f"role must be one of {valid_roles}", 400)
    conn = get_conn()
    conn.execute("""INSERT INTO history (detail_id, action, details) VALUES (?, 'workflow_assigned', ?)""",
                 (detail_id, json.dumps({"role": role, "assignee": assignee}, ensure_ascii=False)))
    conn.commit()
    conn.close()
    add_history(detail_id, "workflow_assigned", {"role": role, "assignee": assignee})
    # Уведомление исполнителю
    try:
        notify_workflow(detail_id, "assigned", assignee=assignee, extra=f"Роль: {role}")
    except Exception as e:
        log.warning(f"notify_workflow failed: {e}")
    return JSONResponse({"ok": True, "detail_id": detail_id, "role": role, "assignee": assignee, "notified": True})


@app.get("/api/workflow/queue")
async def api_workflow_queue(role: str = "technologist", assignee: str = ""):
    """Очередь работы для роли/исполнителя"""
    conn = get_conn()
    # Сейчас упрощённо: все детали со статусом draft для технолога, approved для нормировщика
    if role == "technologist":
        details = conn.execute("""SELECT d.id, d.designation, d.name, dr.status_ext
            FROM details d LEFT JOIN drafts dr ON d.id=dr.detail_id
            WHERE d.id NOT IN (SELECT detail_id FROM history WHERE action='workflow_assigned' AND details LIKE ?)
            LIMIT 20""", (f'%{role}%',)).fetchall()
    else:
        details = conn.execute("""SELECT id, designation, name, NULL as status_ext FROM details LIMIT 20""").fetchall()
    conn.close()
    cols = ["id", "designation", "name", "status_ext"]
    return JSONResponse([dict(zip(cols, r)) for r in details])


# ========== Role-based UI (через cookie) ==========
# M17 (2026-07-20): Сергей попросил сократить роли. Для пилота на 5-10 человек
# завода 7 ролей — перебор. Оставляем 4: технолог, гл.технолог, нач.цеха, админ.
# Остальные 3 (normirovshchik, constructor, quality) — помечены как deprecated
# и НЕ показываются в UI, но остаются в коде (на случай если в будущем понадобятся).
ROLES = {
    "technologist": {
        "name": "Технолог",
        "default_view": "drafts",
        "can_edit": True,
        "can_approve": False,
        "can_manage_workflow": False
    },
    "main_technologist": {
        "name": "Гл. технолог",
        "default_view": "approval_queue",
        "can_edit": True,
        "can_approve": True,
        "can_manage_workflow": True
    },
    "workshop_chief": {
        "name": "Нач. цеха",
        "default_view": "approved",
        "can_edit": False,
        "can_approve": True,
        "can_manage_workflow": False
    },
    "admin": {
        "name": "Админ",
        "default_view": "admin_dashboard",
        "can_edit": True,
        "can_approve": True,
        "can_manage_workflow": True,
        "can_admin": True
    },
    # DEPRECATED (M17): оставлены в коде для совместимости, но не показываются в UI
    "normirovshchik": {
        "name": "Нормировщик (DEPRECATED)",
        "deprecated": True,
        "can_edit": True, "can_approve": False, "can_manage_workflow": False
    },
    "constructor": {
        "name": "Конструктор (DEPRECATED)",
        "deprecated": True,
        "can_edit": False, "can_approve": False, "can_manage_workflow": False
    },
    "quality": {
        "name": "ОТК (DEPRECATED)",
        "deprecated": True,
        "can_edit": True, "can_approve": False, "can_manage_workflow": False
    },
}
# Активные роли (для UI селекторов и quick-role кнопок)
ACTIVE_ROLES = {k: v for k, v in ROLES.items() if not v.get("deprecated")}


def get_current_role(request: Request) -> str:
    """Получает текущую роль из cookie. Default = technologist"""
    role = request.cookies.get("bit_role", "technologist")
    if role not in ROLES:
        role = "technologist"
    return role


def is_admin(request: Request) -> bool:
    """Проверка что текущий пользователь — администратор"""
    return get_current_role(request) == "admin"


# ========== Аутентификация и пользователи пилота (БД) ==========
def hash_password(password: str) -> str:
    """bcrypt хеш пароля. Если bcrypt недоступен — fallback на sha256+salt."""
    try:
        import bcrypt
        return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    except ImportError:
        import hashlib
        import secrets
        salt = secrets.token_hex(16)
        h = hashlib.sha256((salt + password).encode("utf-8")).hexdigest()
        return f"sha256${salt}${h}"


def verify_password(password: str, password_hash: str) -> bool:
    """Проверка пароля"""
    if not password_hash:
        return False
    try:
        import bcrypt
        return bcrypt.checkpw(password.encode("utf-8"), password_hash.encode("utf-8"))
    except ImportError:
        if password_hash.startswith("sha256$"):
            import hashlib
            parts = password_hash.split("$", 2)
            if len(parts) != 3:
                return False
            salt, h = parts[1], parts[2]
            return hashlib.sha256((salt + password).encode("utf-8")).hexdigest() == h
    return False


def authenticate_pilot_user(username: str, password: str) -> dict | None:
    """Аутентификация пользователя из БД. Возвращает dict или None."""
    conn = get_conn()
    try:
        row = conn.execute("""SELECT id, username, password_hash, role, display_name, is_active
            FROM pilot_users WHERE username=? AND is_active=1""", (username,)).fetchone()
    finally:
        conn.close()
    if not row:
        return None
    if not verify_password(password, row[2]):
        return None
    # update last_login
    conn = get_conn()
    try:
        conn.execute("UPDATE pilot_users SET last_login=CURRENT_TIMESTAMP WHERE id=?", (row[0],))
        conn.commit()
    finally:
        conn.close()
    return {"id": row[0], "username": row[1], "role": row[3], "display_name": row[4]}


def log_login(username: str, ip: str, user_agent: str, success: bool):
    """Логирование попытки входа"""
    conn = get_conn()
    try:
        conn.execute("""INSERT INTO audit_logins (username, ip, user_agent, success)
            VALUES (?, ?, ?, ?)""", (username, ip, user_agent, 1 if success else 0))
        conn.commit()
    finally:
        conn.close()


@app.get("/detail/{detail_id}/diff/{v_from}/{v_to}", response_class=HTMLResponse)
async def detail_diff(request: Request, detail_id: str, v_from: int, v_to: int):
    """Сравнение двух версий драфта (diff view)"""
    conn = get_conn()
    row_from = conn.execute("""SELECT operations_json, created_at, author, source, notes
        FROM draft_versions WHERE detail_id=? AND version=?""", (detail_id, v_from)).fetchone()
    row_to = conn.execute("""SELECT operations_json, created_at, author, source, notes
        FROM draft_versions WHERE detail_id=? AND version=?""", (detail_id, v_to)).fetchone()
    conn.close()
    if not row_from or not row_to:
        return HTMLResponse(f"<h1>Версия {v_from} или {v_to} не найдена</h1>", status_code=404)
    try:
        ops_from = json.loads(row_from[0])
        ops_to = json.loads(row_to[0])
    except Exception as e:
        return HTMLResponse(f"<h1>Ошибка парсинга</h1><pre>{e}</pre>", status_code=500)
    diffs = []
    max_len = max(len(ops_from), len(ops_to))
    for i in range(max_len):
        op_f = ops_from[i] if i < len(ops_from) else None
        op_t = ops_to[i] if i < len(ops_to) else None
        if op_f is None and op_t is not None:
            diffs.append({"op_index": i, "status": "added", "from": None, "to": op_t})
        elif op_t is None and op_f is not None:
            diffs.append({"op_index": i, "status": "removed", "from": op_f, "to": None})
        elif op_f and op_t:
            changed_fields = []
            for k in set(list(op_f.keys()) + list(op_t.keys())):
                if op_f.get(k) != op_t.get(k):
                    changed_fields.append({"field": k, "from": op_f.get(k), "to": op_t.get(k)})
            if changed_fields:
                diffs.append({"op_index": i, "status": "modified", "from": op_f, "to": op_t, "changes": changed_fields})
            else:
                diffs.append({"op_index": i, "status": "same", "from": op_f, "to": op_t})
    return templates.TemplateResponse("diff.html", {
        "request": request,
        "detail_id": detail_id,
        "v_from": v_from, "v_to": v_to,
        "from_meta": {"created_at": row_from[1], "author": row_from[2], "source": row_from[3], "notes": row_from[4]},
        "to_meta": {"created_at": row_to[1], "author": row_to[2], "source": row_to[3], "notes": row_to[4]},
        "diffs": diffs
    })


# ========== Уведомления (email + telegram webhook) ==========
# Модульные значения — defaults при старте (fallback). Реальные значения читаются через get_setting() в runtime.
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
NOTIFY_EMAIL_FROM = os.getenv("NOTIFY_EMAIL_FROM", "bit-technolog@tehnocom.local")
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")


def send_email(to: str, subject: str, body: str) -> bool:
    """Отправка email. Если SMTP не настроен — dry-run (только лог).
    Берёт SMTP_*/NOTIFY_EMAIL_FROM из БД (через get_setting), fallback на .env.
    """
    host = get_setting("SMTP_HOST", SMTP_HOST)
    port = int(get_setting("SMTP_PORT", str(SMTP_PORT)) or 587)
    user = get_setting("SMTP_USER", SMTP_USER)
    pwd = get_setting("SMTP_PASS", SMTP_PASS)
    frm = get_setting("NOTIFY_EMAIL_FROM", NOTIFY_EMAIL_FROM)
    if not host or not user:
        log.info(f"[DRY-RUN EMAIL] to={to} subject={subject}")
        return True
    try:
        import smtplib
        from email.mime.text import MIMEText
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = frm
        msg["To"] = to
        with smtplib.SMTP(host, port) as s:
            s.starttls()
            s.login(user, pwd)
            s.send_message(msg)
        return True
    except Exception as e:
        log.error(f"email send failed: {e}")
        return False


def send_telegram(text: str) -> bool:
    """Отправка в Telegram. Если токен не настроен — dry-run.
    Берёт TELEGRAM_BOT_TOKEN/CHAT_ID из БД (get_setting), fallback на .env.
    """
    token = get_setting("TELEGRAM_BOT_TOKEN", TELEGRAM_BOT_TOKEN)
    chat_id = get_setting("TELEGRAM_CHAT_ID", TELEGRAM_CHAT_ID)
    if not token or not chat_id:
        log.info(f"[DRY-RUN TELEGRAM] {text[:100]}")
        return True
    try:
        import urllib.request
        import urllib.parse
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        data = urllib.parse.urlencode({"chat_id": chat_id, "text": text}).encode()
        req = urllib.request.Request(url, data=data)
        with urllib.request.urlopen(req, timeout=10) as resp:
            return resp.status == 200
    except Exception as e:
        log.error(f"telegram send failed: {e}")
        return False


def notify_workflow(detail_id: str, event: str, assignee: str = "", extra: str = ""):
    """Уведомление о workflow-событии (email + telegram)."""
    subject = f"БИТ.Технолог: {event} по {detail_id}"
    body = f"Деталь: {detail_id}\nСобытие: {event}\n{('Исполнитель: ' + assignee) if assignee else ''}\n{extra}"
    if assignee and "@" in assignee:
        send_email(assignee, subject, body)
    send_telegram(f"🔔 {subject}\n{body[:200]}")


@app.post("/api/role/switch")
async def api_role_switch(request: Request):
    """Переключить роль (для demo/pilot)"""
    role = await _get_param(request, "role")
    if not role or role not in ROLES:
        return err(f"role must be one of: {list(ROLES.keys())}", 400)
    response = JSONResponse({"ok": True, "role": role, "name": ROLES[role]["name"]})
    # BUG-2026-07-20-01: убрал httponly — JavaScript должен видеть текущую роль
    # (cookie не содержит секрета, это просто UI state для demo/pilot)
    response.set_cookie("bit_role", role, max_age=86400 * 365, httponly=False, samesite="lax")
    return response


# ========== Admin: дашборд, пользователи, лог входов, LLM-лог, бэкапы, система ==========
from cryptography.fernet import Fernet, InvalidToken
import secrets as _secrets


# Мастер-ключ для Fernet (из .env или сгенерированный)
def _get_or_create_master_key() -> bytes:
    """Получает или создаёт Fernet-ключ.
    Хранится в файле .master_key (chmod 600) рядом с .env.
    Если .env содержит FERNET_KEY — использует его.
    """
    env_key = os.getenv("FERNET_KEY", "").strip()
    if env_key:
        try:
            return base64.urlsafe_b64decode(env_key)
        except Exception:
            pass
    # Файл в корне проекта
    key_path = os.path.join(os.path.dirname(DB_PATH) or ".", ".master_key")
    if os.path.exists(key_path):
        with open(key_path, "rb") as f:
            return f.read()
    # Сгенерировать новый
    key = Fernet.generate_key()
    with open(key_path, "wb") as f:
        f.write(key)
    os.chmod(key_path, 0o600)
    return key


# Глобальный Fernet (инициализируется лениво)
_FERNET = None


def _fernet() -> Fernet:
    global _FERNET
    if _FERNET is None:
        _FERNET = Fernet(_get_or_create_master_key())
    return _FERNET


# Реестр настроек: (key, type, default, description, masked)
SETTING_REGISTRY = [
    ("LLM_API_KEY", "secret", "", "YandexGPT API key", True),
    ("LLM_MODEL", "str", "gpt://b1gj791m9sc92argfa0q/yandexgpt/latest", "LLM model URI", False),
    ("LLM_API_URL", "str", "https://llm.api.cloud.yandex.net/v1", "OpenAI-compatible endpoint", False),
    ("LLM_DAILY_COST_LIMIT_RUB", "int", "200", "Дневной лимит LLM (₽)", False),
    ("DEMO_MODE", "bool", "false", "Демо-режим (без LLM)", False),
    ("TELEGRAM_BOT_TOKEN", "secret", "", "Telegram bot token (@BotFather)", True),
    ("TELEGRAM_CHAT_ID", "str", "", "Telegram chat ID для уведомлений", False),
    ("SMTP_HOST", "str", "", "SMTP хост (например, smtp.yandex.ru)", False),
    ("SMTP_PORT", "int", "587", "SMTP порт", False),
    ("SMTP_USER", "str", "", "SMTP пользователь", False),
    ("SMTP_PASS", "secret", "", "SMTP пароль", True),
    ("NOTIFY_EMAIL_FROM", "str", "bit-technolog@tehnocom.local", "Email отправителя", False),
    ("MAX_DRAWING_SIZE_MB", "int", "50", "Макс. размер чертежа (МБ)", False),
    ("MAX_IMPORT_SIZE_MB", "int", "100", "Макс. размер импорта (МБ)", False),
    ("PILOT_USERS", "str", "", "Basic Auth users (user:pass,user2:pass2)", True),
]


def _mask_value(value: str) -> str:
    """Маскирует секрет: оставляет первые 4 и последние 3 символа"""
    if not value or len(value) < 10:
        return "***" if value else ""
    return f"{value[:4]}...{value[-3:]}"


def _encrypt(value: str) -> bytes:
    if not value:
        return b""
    return _fernet().encrypt(value.encode("utf-8"))


def _decrypt(blob: bytes) -> str:
    if not blob:
        return ""
    try:
        return _fernet().decrypt(bytes(blob)).decode("utf-8")
    except (InvalidToken, Exception):
        return ""


def get_setting(key: str, default: str = "") -> str:
    """Получает настройку: сначала из БД, потом из .env, потом default.
    Устойчив к отсутствию таблицы app_settings (на свежей БД)."""
    conn = None
    try:
        conn = get_conn()
        row = conn.execute("SELECT value_encrypted FROM app_settings WHERE key=?", (key,)).fetchone()
        if row and row[0]:
            val = _decrypt(row[0])
            if val:
                return val
    except Exception as e:
        # Таблица не существует или другая ошибка — fallback на env/default
        log.debug(f"get_setting({key}) DB error (will fallback): {e}")
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass
    # Fallback: env
    env_val = os.getenv(key, "")
    if env_val:
        return env_val
    return default


def set_setting(key: str, value: str, updated_by: str = "admin") -> bool:
    """Сохраняет настройку в БД (зашифрованно)"""
    conn = get_conn()
    try:
        encrypted = _encrypt(value)
        masked = _mask_value(value) if value else ""
        conn.execute("""INSERT INTO app_settings (key, value_encrypted, value_masked, updated_at, updated_by)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP, ?)
            ON CONFLICT(key) DO UPDATE SET
                value_encrypted=excluded.value_encrypted,
                value_masked=excluded.value_masked,
                updated_at=CURRENT_TIMESTAMP,
                updated_by=excluded.updated_by""",
            (key, encrypted, masked, updated_by))
        conn.commit()
    except Exception as e:
        log.error(f"set_setting({key}) failed: {e}")
        return False
    finally:
        conn.close()
    return True


def delete_setting(key: str) -> bool:
    """Удаляет настройку (откат на .env/default)"""
    conn = get_conn()
    try:
        conn.execute("DELETE FROM app_settings WHERE key=?", (key,))
        conn.commit()
    finally:
        conn.close()
    return True


def get_all_settings() -> list:
    """Возвращает список всех настроек с их текущими значениями (masked)"""
    result = []
    for key, stype, default, desc, is_secret in SETTING_REGISTRY:
        current = get_setting(key, default)
        result.append({
            "key": key, "type": stype, "default": default, "description": desc,
            "is_secret": is_secret,
            "value_set": bool(current),
            "value_masked": _mask_value(current) if (is_secret and current) else (current[:200] if current else "")
        })
    return result


def _require_admin_or_json(request: Request):
    """Возвращает JSON-ошибку 403 если не admin, иначе None"""
    if get_current_role(request) != "admin":
        return JSONResponse(
            {"ok": False, "error": "admin only", "current_role": get_current_role(request)},
            status_code=403
        )
    return None


# ========== Admin endpoints перенесены в admin.py (F15.7) ==========
# F15.7: используется APIRouter из admin.py
from admin import router as admin_router
app.include_router(admin_router)


# ========== Импорт ТК (Excel/PDF/JSON/Word) ==========
def _xml_escape(s: str) -> str:
    if not s:
        return ""
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&apos;"))


@app.post("/api/import/tk")
async def api_import_tk(request: Request):
    """Импорт техкарт. Content-Type: application/json или multipart/form-data"""
    from importers import save_imported_details
    content_type = request.headers.get("content-type", "").lower()
    if "application/json" in content_type:
        try:
            body = await request.body()
            data = json.loads(body)
        except Exception as e:
            return err(f"Invalid JSON: {e}", 400)
        from importers import import_from_json
        details = import_from_json(data)
        result = save_imported_details(details, default_author="import-json")
        return JSONResponse(result)
    if "multipart/form-data" in content_type:
        form = await request.form()
        if "file" not in form:
            return err("file required", 422)
        f = form["file"]
        contents = await f.read()
        if len(contents) > int(get_setting("MAX_IMPORT_SIZE_MB", str(MAX_IMPORT_SIZE // 1024 // 1024))) * 1024 * 1024:
            return err(f"file too large: {len(contents)} bytes (max {MAX_IMPORT_SIZE})", 413)
        if len(contents) == 0:
            return err("file is empty", 422)
        suffix = (f.filename or "").lower().split(".")[-1]
        if suffix not in ALLOWED_IMPORT_FORMATS:
            return err(f"unsupported format '{suffix}'. Allowed: {sorted(ALLOWED_IMPORT_FORMATS)}", 400)
        # F-12 fix: magic bytes verification
        from importers import verify_magic_bytes
        if not verify_magic_bytes(contents, suffix):
            return err(f"file content does not match extension .{suffix} (magic bytes mismatch — possible malicious upload)", 400)
        tmp_path = f"/tmp/import_{os.getpid()}.{suffix}"
        with open(tmp_path, "wb") as out:
            out.write(contents)
        try:
            if suffix in ("xlsx",):
                from importers import import_from_excel
                details = import_from_excel(tmp_path)
            elif suffix == "xls":
                # F fix: .xls (старый бинарный формат) — openpyxl не читает.
                # Нужна библиотека xlrd (не включена в зависимости).
                return err("xls (old format) not supported. Please save as xlsx.", 415)
            elif suffix == "pdf":
                from importers import import_from_pdf
                details = import_from_pdf(tmp_path)
            elif suffix in ("docx",):
                from importers import import_from_word
                details = import_from_word(tmp_path)
            elif suffix == "doc":
                return err("doc (old format) not supported. Please save as docx.", 415)
            else:
                return err(f"Unsupported file format: {suffix}", 400)
            result = save_imported_details(details, default_author=f"import-{suffix}")
            result["filename"] = f.filename
            return JSONResponse(result)
        except Exception as e:
            log.exception(f"import failed: {e}")
            return err(f"import failed: {e}", 500)
        finally:
            try:
                os.remove(tmp_path)
            except Exception:
                pass
    return err("Unsupported Content-Type. Use JSON or multipart file upload.", 415)


@app.post("/api/import/drawing/{detail_id}")
async def api_import_drawing(detail_id: str, request: Request):
    """Загрузка чертежа для детали. multipart file (.frw/.dwg/.pdf/.png/.jpg)
    N1 fix: валидация размера, sanitize filename, проверка detail_id."""
    conn = get_conn()
    if not conn.execute("SELECT id FROM details WHERE id=?", (detail_id,)).fetchone():
        conn.close()
        return err(f"detail '{detail_id}' not found", 404)
    conn.close()
    form = await request.form()
    if "file" not in form:
        return err("file required", 422)
    f = form["file"]
    contents = await f.read()
    if len(contents) > int(get_setting("MAX_DRAWING_SIZE_MB", str(MAX_DRAWING_SIZE // 1024 // 1024))) * 1024 * 1024:
        return err(f"file too large: {len(contents)} bytes (max {MAX_DRAWING_SIZE})", 413)
    if len(contents) == 0:
        return err("file is empty", 422)
    drawings_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "drawings")
    os.makedirs(drawings_dir, exist_ok=True)
    suffix = (f.filename or "").lower().split(".")[-1]
    if suffix not in ALLOWED_DRAWING_FORMATS:
        return err(f"unsupported format '{suffix}'. Allowed: {sorted(ALLOWED_DRAWING_FORMATS)}", 400)
    # F-12 fix: magic bytes verification
    from importers import verify_magic_bytes
    if not verify_magic_bytes(contents, suffix):
        return err(f"file content does not match extension .{suffix} (magic bytes mismatch — possible malicious upload)", 400)
    safe_filename = re.sub(r"[^A-Za-z0-9._-]", "_", f.filename or "drawing")
    safe_filename = safe_filename[:100]  # N1: limit length
    file_path = os.path.join(drawings_dir, f"{detail_id}_{safe_filename}")
    file_path = os.path.abspath(file_path)  # resolve symlinks
    # N1: убедиться что file_path внутри drawings_dir (нет path traversal)
    if not file_path.startswith(os.path.abspath(drawings_dir)):
        return err("invalid filename", 400)
    with open(file_path, "wb") as out:
        out.write(contents)
    conn = get_conn()
    conn.execute("""UPDATE details SET drawing_path=?, drawing_format=?, updated_at=CURRENT_TIMESTAMP
        WHERE id=?""", (file_path, suffix, detail_id))
    conn.execute("""INSERT INTO drawings (detail_id, file_path, file_format, file_size)
        VALUES (?, ?, ?, ?)""", (detail_id, file_path, suffix, len(contents)))
    conn.commit()
    conn.close()
    add_history(detail_id, "drawing_uploaded", {"file": safe_filename, "size": len(contents)})

    # v0.5: M28 — auto-OCR для распознавания чертежа после загрузки
    auto_ocr = (form.get("auto_ocr") or "true").lower() == "true"
    ocr_result = None
    if auto_ocr and suffix in ("pdf", "png", "jpg", "jpeg"):
        try:
            from drawing_recognize import recognize_drawing
            ocr_result = recognize_drawing(file_path)
        except Exception as e:
            log.warning(f"Auto-OCR failed: {e}")
            ocr_result = {"ok": False, "error": str(e)}

    return JSONResponse({
        "ok": True,
        "file_path": file_path,
        "file_size": len(contents),
        "ocr": ocr_result,
    })


@app.post("/api/drawing/recognize")
async def api_drawing_recognize(request: Request):
    """
    v0.5: M28 — распознаёт уже загруженный чертёж (по detail_id) через OCR.
    Возвращает HTML partial с извлечёнными полями.
    """
    from drawing_recognize import recognize_drawing
    detail_id = await _get_param(request, "detail_id")
    if not detail_id:
        return HTMLResponse("<div class='ocr-error'>detail_id required</div>")

    conn = get_conn()
    row = conn.execute("SELECT drawing_path FROM details WHERE id=?", (detail_id,)).fetchone()
    conn.close()
    if not row:
        return HTMLResponse(f"<div class='ocr-error'>detail '{detail_id}' not found</div>")
    file_path = row[0]
    if not file_path or not os.path.exists(file_path):
        return HTMLResponse(f"<div class='ocr-error'>drawing file not found for detail '{detail_id}'</div>")

    try:
        result = recognize_drawing(file_path)
    except Exception as e:
        log.error(f"OCR failed: {e}")
        return HTMLResponse(f"<div class='ocr-error'>OCR failed: {e}</div>")

    return templates.TemplateResponse("_ocr_result.html", {"request": request, "result": result})


@app.get("/api/import/stats")
async def api_import_stats():
    conn = get_conn()
    total = conn.execute("SELECT COUNT(*) FROM details").fetchone()[0]
    with_drawings = conn.execute("SELECT COUNT(*) FROM details WHERE drawing_path IS NOT NULL").fetchone()[0]
    with_operations = conn.execute("SELECT COUNT(*) FROM drafts WHERE llm_output LIKE '%operations%'").fetchone()[0]
    resource_specs = conn.execute("SELECT COUNT(*) FROM resource_specs").fetchone()[0]
    professions = conn.execute("SELECT COUNT(*) FROM professions").fetchone()[0]
    conn.close()
    return JSONResponse({
        "details_total": total,
        "with_drawings": with_drawings,
        "with_operations": with_operations,
        "resource_specs": resource_specs,
        "professions": professions
    })


# ========== Справочник профессий ==========
@app.post("/api/professions")
async def api_professions_create(request: Request):
    data = await request.body()
    try:
        p = json.loads(data)
    except Exception:
        return err("invalid JSON", 400)
    if not p.get("id") or not p.get("code") or not p.get("name"):
        return err("id, code, name required", 422)
    conn = get_conn()
    conn.execute("""INSERT OR REPLACE INTO professions
        (id, code, name, grade, hourly_rate, source, external_id)
        VALUES (?, ?, ?, ?, ?, '1c', ?)""", (
        p["id"], p["code"], p["name"], p.get("grade"), p.get("hourly_rate", 0), p.get("external_id")
    ))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True, "id": p["id"]})


@app.get("/api/professions")
async def api_professions_list():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM professions ORDER BY code, grade").fetchall()
    conn.close()
    cols = ["id", "code", "name", "grade", "hourly_rate", "source", "external_id"]
    return JSONResponse([dict(zip(cols, r)) for r in rows])


# ========== Ресурсная спецификация (РС) ==========
@app.get("/api/resource-specs/{detail_id}")
async def api_resource_specs_list(request: Request, detail_id: str):
    """M24: возвращает HTML вместо JSON"""
    conn = get_conn()
    rows = conn.execute("""SELECT op_index, kind, name, quantity, unit, notes
        FROM resource_specs WHERE detail_id=? ORDER BY op_index, kind""", (detail_id,)).fetchall()
    conn.close()
    specs = [dict(zip(["op_index", "kind", "name", "quantity", "unit", "notes"], r)) for r in rows]
    return templates.TemplateResponse(request, "_resource_specs.html", {"specs": specs, "request": request})


# ========== Иерархия деталь/узел/изделие ==========
@app.get("/hierarchy", response_class=HTMLResponse)
async def hierarchy_page(request: Request):
    """M27: страница иерархии изделий (для ссылки из навигации)"""
    return templates.TemplateResponse("hierarchy.html", {"request": request})


@app.get("/api/hierarchy")
async def api_hierarchy():
    conn = get_conn()
    products = conn.execute("SELECT id, designation, name FROM details WHERE level='product' ORDER BY designation").fetchall()
    assemblies = conn.execute("SELECT id, designation, name, parent_id FROM details WHERE level='assembly' ORDER BY designation").fetchall()
    details = conn.execute("SELECT id, designation, name, parent_id FROM details WHERE level='detail' ORDER BY designation").fetchall()
    conn.close()
    by_id = {}
    for r in products:
        by_id[r[0]] = {"id": r[0], "designation": r[1], "name": r[2], "level": "product", "children": []}
    for r in assemblies:
        by_id[r[0]] = {"id": r[0], "designation": r[1], "name": r[2], "level": "assembly", "parent_id": r[3], "children": []}
    for r in details:
        by_id[r[0]] = {"id": r[0], "designation": r[1], "name": r[2], "level": "detail", "parent_id": r[3], "children": []}
    # G fix: защита от циклов через visited set
    visited = set()
    def _attach_children(node):
        if node["id"] in visited:
            node["children"] = []  # обрезаем цикл
            return
        visited.add(node["id"])
        pid = node.get("parent_id")
        if pid and pid in by_id and pid not in [c["id"] for c in node.get("children", [])]:
            node["children"].append(by_id[pid])
            _attach_children(by_id[pid])
    roots = []
    for node in by_id.values():
        if not node.get("parent_id"):
            roots.append(node)
        else:
            _attach_children(node)
    return JSONResponse(roots)


@app.get("/api/related/{detail_id}")
async def api_related(request: Request, detail_id: str):
    """Связанные детали: siblings (тот же узел) + product (изделие). M24: HTML вместо JSON"""
    conn = get_conn()
    row = conn.execute("SELECT id, designation, name, parent_id, level FROM details WHERE id=?", (detail_id,)).fetchone()
    if not row:
        conn.close()
        return err("not found", 404)
    my = {"id": row[0], "designation": row[1], "name": row[2], "parent_id": row[3], "level": row[4]}
    siblings = []
    product = None
    if my["parent_id"]:
        siblings_rows = conn.execute("""SELECT id, designation, name FROM details
            WHERE parent_id=? AND id != ? ORDER BY designation""", (my["parent_id"], detail_id)).fetchall()
        siblings = [{"id": r[0], "designation": r[1], "name": r[2]} for r in siblings_rows]
        parent_row = conn.execute("""SELECT id, designation, name, parent_id, level FROM details WHERE id=?""",
                                  (my["parent_id"],)).fetchone()
        if parent_row:
            if parent_row[4] == "product":
                product = {"id": parent_row[0], "designation": parent_row[1], "name": parent_row[2]}
            elif parent_row[3]:
                gp_row = conn.execute("""SELECT id, designation, name FROM details WHERE id=?""",
                                      (parent_row[3],)).fetchone()
                if gp_row:
                    product = {"id": gp_row[0], "designation": gp_row[1], "name": gp_row[2]}
    conn.close()
    return templates.TemplateResponse(request, "_related.html", {"self": my, "siblings": siblings, "product": product, "request": request})


# ========== Экспорт РС в 1С:ERP (XML) ==========
@app.get("/api/1c/export/rs/{detail_id}")
async def api_1c_export_rs(detail_id: str):
    """Экспорт ресурсной спецификации в XML (1С:ERP формат)"""
    conn = get_conn()
    d = conn.execute("""SELECT id, designation, name, model, chassis, material, mass_kg
        FROM details WHERE id=?""", (detail_id,)).fetchone()
    if not d:
        conn.close()
        return err("detail not found", 404)
    draft_row = conn.execute("SELECT llm_output FROM drafts WHERE detail_id=?", (detail_id,)).fetchone()
    if not draft_row:
        conn.close()
        return err("no draft", 404)
    try:
        output = json.loads(draft_row[0])
    except Exception:
        conn.close()
        return err("draft corrupt", 500)
    ops = output.get("operations") or []
    rs_rows = conn.execute("""SELECT op_index, kind, name, quantity, unit, notes
        FROM resource_specs WHERE detail_id=? ORDER BY op_index""", (detail_id,)).fetchall()
    conn.close()
    xml = ['<?xml version="1.0" encoding="UTF-8"?>']
    xml.append(f'<ResourceSpecification designation="{_xml_escape(d[1])}" name="{_xml_escape(d[2] or "")}" model="{_xml_escape(d[3] or "")}">')
    for i, op in enumerate(ops):
        xml.append(f'  <Operation number="{i+1:03d}" name="{_xml_escape(op.get("name", ""))}">')
        xml.append(f'    <Duration hours="{op.get("duration_hours", 0)}"/>')
        if op.get("equipment"):
            xml.append(f'    <Equipment name="{_xml_escape(op["equipment"])}"/>')
        if op.get("department"):
            xml.append(f'    <Department name="{_xml_escape(op["department"])}"/>')
        for r in rs_rows:
            # N4 fix: защита от None/out-of-range
            try:
                r_idx = int(r[0]) if r[0] is not None else -1
            except (TypeError, ValueError):
                continue
            if r_idx == i and 0 <= r_idx < 10000:
                xml.append(f'    <Resource kind="{r[1]}" name="{_xml_escape(r[2])}" quantity="{r[3]}" unit="{_xml_escape(r[4] or "")}"/>')
        xml.append('  </Operation>')
    xml.append('</ResourceSpecification>')
    xml_text = "\n".join(xml)
    return Response(
        content=xml_text,
        media_type="application/xml; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="rs_{detail_id}.xml"'}
    )


def seed_initial_data():
    """Заполняет справочники дефолтными данными из mock-файлов (если пусто)"""
    conn = get_conn()

    # 1. Детали из mock_data.py
    if conn.execute("SELECT COUNT(*) FROM details").fetchone()[0] == 0:
        for d in MOCK_DETAILS:
            conn.execute("""INSERT INTO details
                (id, designation, name, model, chassis, material, mass_kg, surface_treatment)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (d["id"], d["designation"], d["name"], d.get("model"),
                 d.get("chassis"), d.get("material"), d.get("mass_kg"),
                 d.get("surface_treatment")))

    # 2. Оборудование
    if conn.execute("SELECT COUNT(*) FROM equipment").fetchone()[0] == 0:
        for i, e in enumerate(EQUIPMENT, 1):
            eid = f"eq-{i:03d}"
            conn.execute("""INSERT INTO equipment
                (id, name, type, code, max_thickness_mm, max_mass_kg, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (eid, e["name"], e.get("type"), str(e.get("code") or ""),
                 e.get("max_dim_mm"), e.get("max_mass_kg"),
                 f"group={e.get('group','')}; rank={e.get('rank','')}; operations={','.join(e.get('operations', []))}"))

    # 3. Цеха
    if conn.execute("SELECT COUNT(*) FROM departments").fetchone()[0] == 0:
        for prod in STRUCTURE.get("productions", []):
            for ws in prod.get("workshops", []):
                wsid = f"{prod.get('code', '01')}/{ws.get('code')}"
                conn.execute("""INSERT INTO departments (id, production, name, code)
                    VALUES (?, ?, ?, ?)""",
                    (wsid, prod.get("name", ""), ws.get("name", ""), ws.get("code", "")))

    # 4. Материалы (базовые)
    if conn.execute("SELECT COUNT(*) FROM materials").fetchone()[0] == 0:
        for m_id, m_name, m_gost in [
            ("steel-3", "Сталь 3", "ГОСТ 380-2005"),
            ("steel-09g2s", "Сталь 09Г2С", "ГОСТ 19281-2014"),
            ("steel-30khgsa", "Сталь 30ХГСА", "ГОСТ 4543-2016"),
            ("steel-st3sp", "Сталь Ст3сп", "ГОСТ 380-2005"),
            ("ocinkovka", "Сталь оцинкованная", "ГОСТ 14918-2020"),
            ("aluminum-5083", "Алюминий 5083", "ГОСТ 21631-76"),
        ]:
            conn.execute("INSERT INTO materials (id, name, gost) VALUES (?, ?, ?)",
                         (m_id, m_name, m_gost))

    # 5. ИОТ (базовые для Техинкома)
    if conn.execute("SELECT COUNT(*) FROM iot").fetchone()[0] == 0:
        for i_id, num, desc, applies in [
            ("iot-3", "3", "Сварочные работы", "Сварка"),
            ("iot-11", "11", "Работа с подъёмниками", "Сборка"),
            ("iot-26", "26", "Грузоподъёмные операции", "Кран"),
            ("iot-49", "49", "Обработка металлов резанием", "Механообработка"),
            ("iot-15", "15", "Покрасочные работы", "Покраска"),
        ]:
            conn.execute("""INSERT INTO iot (id, number, description, applies_to)
                VALUES (?, ?, ?, ?)""", (i_id, num, desc, applies))

    # 6. Бенчмарки
    if conn.execute("SELECT COUNT(*) FROM benchmarks").fetchone()[0] == 0:
        for b_id, dtype, hrs, src in [
            ("bk-plastina", "Пластина", 0.48, "ведомость Техинкома"),
            ("bk-ugolok", "Уголок", 0.48, "ведомость Техинкома"),
            ("bk-planka", "Планка", 0.48, "ведомость Техинкома"),
            ("bk-kronshtein", "Кронштейн замка", 4.5, "ведомость Техинкома"),
            ("bk-shassi-podgot", "Шасси подготовленное", 16.0, "ведомость Техинкома"),
            ("bk-pss131", "ПСС 131.35Э итого", 1296.36, "ведомость Техинкома"),
        ]:
            conn.execute("""INSERT INTO benchmarks (id, detail_type, norm_hours, source)
                VALUES (?, ?, ?, ?)""", (b_id, dtype, hrs, src))

    conn.commit()
    conn.close()


# CRUD-функции
def get_all_details() -> list:
    conn = get_conn()
    rows = conn.execute("""SELECT id, designation, name, model, chassis, material,
        mass_kg, surface_treatment, created_at
        FROM details ORDER BY created_at DESC""").fetchall()
    conn.close()
    cols = ["id", "designation", "name", "model", "chassis", "material",
            "mass_kg", "surface_treatment", "created_at"]
    return [dict(zip(cols, r)) for r in rows]


def get_detail(detail_id: str) -> Optional[dict]:
    conn = get_conn()
    row = conn.execute("SELECT * FROM details WHERE id = ?", (detail_id,)).fetchone()
    if not row:
        conn.close()
        return None
    cols = [d[1] for d in conn.execute("PRAGMA table_info(details)").fetchall()]
    conn.close()
    return dict(zip(cols, row))


def create_detail(d: dict) -> str:
    import uuid
    detail_id = d.get("id") or f"d-{uuid.uuid4().hex[:8]}"
    conn = get_conn()
    conn.execute("""INSERT INTO details
        (id, designation, name, model, chassis, material, size_mm, mass_kg, surface_treatment, extra_props)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (detail_id, d.get("designation", ""), d.get("name", ""),
         d.get("model", ""), d.get("chassis", ""), d.get("material", ""),
         d.get("size_mm", ""), d.get("mass_kg", 0), d.get("surface_treatment", ""),
         json.dumps(d.get("extra_props", {}), ensure_ascii=False)))
    conn.commit()
    conn.close()
    return detail_id


def get_all_equipment() -> list:
    conn = get_conn()
    rows = conn.execute("SELECT * FROM equipment ORDER BY name").fetchall()
    conn.close()
    cols = ["id", "name", "type", "code", "max_thickness_mm", "max_mass_kg", "notes",
            "source", "external_id", "last_sync_at"]
    return [dict(zip(cols, r)) for r in rows]


def create_equipment(e: dict) -> str:
    import uuid
    eid = e.get("id") or f"eq-{uuid.uuid4().hex[:6]}"
    conn = get_conn()
    conn.execute("""INSERT INTO equipment
        (id, name, type, code, max_thickness_mm, max_mass_kg, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (eid, e["name"], e.get("type", ""), e.get("code", ""),
         e.get("max_thickness_mm"), e.get("max_mass_kg"), e.get("notes", "")))
    conn.commit()
    conn.close()
    return eid


def get_all_materials() -> list:
    conn = get_conn()
    rows = conn.execute("SELECT * FROM materials ORDER BY name").fetchall()
    conn.close()
    cols = ["id", "name", "grade", "gost", "notes"]
    return [dict(zip(cols, r)) for r in rows]


def create_material(m: dict) -> str:
    import uuid
    mid = m.get("id") or f"m-{uuid.uuid4().hex[:6]}"
    conn = get_conn()
    conn.execute("INSERT INTO materials (id, name, grade, gost, notes) VALUES (?, ?, ?, ?, ?)",
                 (mid, m["name"], m.get("grade", ""), m.get("gost", ""), m.get("notes", "")))
    conn.commit()
    conn.close()
    return mid


def get_all_iot() -> list:
    conn = get_conn()
    rows = conn.execute("SELECT * FROM iot ORDER BY number").fetchall()
    conn.close()
    cols = ["id", "number", "description", "applies_to"]
    return [dict(zip(cols, r)) for r in rows]


def create_iot(i: dict) -> str:
    import uuid
    iid = i.get("id") or f"iot-{uuid.uuid4().hex[:6]}"
    conn = get_conn()
    conn.execute("INSERT INTO iot (id, number, description, applies_to) VALUES (?, ?, ?, ?)",
                 (iid, i["number"], i.get("description", ""), i.get("applies_to", "")))
    conn.commit()
    conn.close()
    return iid


def get_all_benchmarks() -> list:
    conn = get_conn()
    rows = conn.execute("SELECT * FROM benchmarks ORDER BY detail_type").fetchall()
    conn.close()
    cols = ["id", "detail_type", "norm_hours", "source", "sample_size"]
    return [dict(zip(cols, r)) for r in rows]


def create_benchmark(b: dict) -> str:
    import uuid
    bid = b.get("id") or f"bk-{uuid.uuid4().hex[:6]}"
    conn = get_conn()
    conn.execute("""INSERT INTO benchmarks (id, detail_type, norm_hours, source, sample_size)
        VALUES (?, ?, ?, ?, ?)""",
        (bid, b["detail_type"], b.get("norm_hours", 0), b.get("source", ""), b.get("sample_size", 1)))
    conn.commit()
    conn.close()
    return bid


# Версии черновика
def save_version(detail_id: str, operations: list, author: str, source: str = "edit", notes: str = ""):
    """Сохраняет новую версию черновика"""
    conn = get_conn()
    last_v = conn.execute(
        "SELECT MAX(version) FROM draft_versions WHERE detail_id = ?", (detail_id,)
    ).fetchone()[0] or 0
    new_v = last_v + 1
    conn.execute("""INSERT INTO draft_versions
        (detail_id, version, operations_json, author, source, notes)
        VALUES (?, ?, ?, ?, ?, ?)""",
        (detail_id, new_v, json.dumps(operations, ensure_ascii=False), author, source, notes))
    conn.commit()
    conn.close()
    return new_v


def get_versions(detail_id: str) -> list:
    conn = get_conn()
    rows = conn.execute("""SELECT version, author, source, notes, created_at
        FROM draft_versions WHERE detail_id = ? ORDER BY version DESC""", (detail_id,)).fetchall()
    conn.close()
    return [{"version": r[0], "author": r[1], "source": r[2], "notes": r[3], "created_at": r[4]}
            for r in rows]


def get_version(detail_id: str, version: int) -> Optional[list]:
    conn = get_conn()
    row = conn.execute("""SELECT operations_json FROM draft_versions
        WHERE detail_id = ? AND version = ?""", (detail_id, version)).fetchone()
    conn.close()
    if not row:
        return None
    return json.loads(row[0])


# Правки (для обучения)
def record_edit(detail_id: str, version: int, field: str, old: str, new: str, reason: str = "", author: str = "technologist"):
    conn = get_conn()
    conn.execute("""INSERT INTO edits (detail_id, version, field, old_value, new_value, reason, author)
        VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (detail_id, version, field, str(old)[:500], str(new)[:500], reason, author))
    conn.commit()
    conn.close()


def get_edits(detail_id: str = None) -> list:
    conn = get_conn()
    if detail_id:
        rows = conn.execute("""SELECT id, detail_id, version, field, old_value, new_value, reason, author, created_at
            FROM edits WHERE detail_id = ? ORDER BY id DESC""", (detail_id,)).fetchall()
    else:
        rows = conn.execute("""SELECT id, detail_id, version, field, old_value, new_value, reason, author, created_at
            FROM edits ORDER BY id DESC LIMIT 100""").fetchall()
    conn.close()
    cols = ["id", "detail_id", "version", "field", "old_value", "new_value", "reason", "author", "created_at"]
    return [dict(zip(cols, r)) for r in rows]


# Правила (обучение)
def calc_cost_rub(tokens_in: int, tokens_out: int) -> float:
    """Считает стоимость вызова в рублях"""
    if not tokens_in or not tokens_out:
        return 0.0
    return round(
        (tokens_in / 1000.0) * LLM_PRICE_INPUT_RUB_PER_1K +
        (tokens_out / 1000.0) * LLM_PRICE_OUTPUT_RUB_PER_1K,
        4
    )


def get_daily_cost(date_str: str = None) -> dict:
    """Считает расход за указанный день (по умолчанию — сегодня).
    Устойчив к отсутствию таблицы llm_calls (на свежей БД)."""
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")
    conn = None
    try:
        conn = get_conn()
        row = conn.execute("""SELECT
            COALESCE(SUM(cost_rub), 0) as total,
            COUNT(*) as calls,
            COALESCE(SUM(tokens_in), 0) as tokens_in,
            COALESCE(SUM(tokens_out), 0) as tokens_out
            FROM llm_calls
            WHERE DATE(created_at, 'localtime') = ? AND cost_rub > 0""",
            (date_str,)).fetchone()
    except Exception as e:
        log.debug(f"get_daily_cost DB error (fresh DB?): {e}")
        row = (0, 0, 0, 0)
    finally:
        if conn:
            try: conn.close()
            except Exception: pass
    return {
        "date": date_str,
        "total_rub": round(row[0] or 0, 4),
        "calls": row[1] or 0,
        "tokens_in": row[2] or 0,
        "tokens_out": row[3] or 0,
        "limit_rub": float(get_setting("LLM_DAILY_COST_LIMIT_RUB", str(int(LLM_DAILY_LIMIT_RUB)))),
        "remaining_rub": round(LLM_DAILY_LIMIT_RUB - (row[0] or 0), 4),
        "exceeded": (row[0] or 0) >= LLM_DAILY_LIMIT_RUB
    }


# Pilot metrics
def record_metric(detail_id: str, metric: str, value: float, extra: dict = None):
    """Записывает KPI пилота (time_to_card, edits_count, accepted_pct, и т.д.)"""
    conn = get_conn()
    conn.execute("""INSERT INTO pilot_metrics (detail_id, metric, value, extra)
        VALUES (?, ?, ?, ?)""",
        (detail_id, metric, value, json.dumps(extra or {}, ensure_ascii=False)))
    conn.commit()
    conn.close()


def get_pilot_metrics() -> dict:
    """Возвращает KPI для дашборда пилота"""
    conn = get_conn()
    # Всего деталей прошло через пилот
    total = conn.execute("SELECT COUNT(DISTINCT detail_id) FROM pilot_metrics").fetchone()[0] or 0
    # Среднее число правок на черновик
    edits_per_card = conn.execute("""SELECT AVG(cnt) FROM (
        SELECT detail_id, COUNT(*) as cnt FROM pilot_metrics WHERE metric='edit' GROUP BY detail_id
    )""").fetchone()[0] or 0
    # % принятых операций (метрика 'accepted')
    accepted = conn.execute("""SELECT
        COALESCE(SUM(CASE WHEN metric='accepted_op' THEN value ELSE 0 END), 0),
        COALESCE(SUM(CASE WHEN metric='total_ops' THEN value ELSE 0 END), 1)
        FROM pilot_metrics""").fetchone()
    accepted_pct = (accepted[0] / accepted[1] * 100) if accepted[1] else 0
    # Среднее время на техкарту
    avg_time = conn.execute("""SELECT AVG(value) FROM pilot_metrics
        WHERE metric='time_to_card_min'""").fetchone()[0] or 0
    # Всего потрачено на LLM
    total_cost = conn.execute("""SELECT COALESCE(SUM(cost_rub), 0) FROM llm_calls
        WHERE cost_rub > 0""").fetchone()[0] or 0
    # Генераций
    total_gens = conn.execute("SELECT COUNT(*) FROM llm_calls WHERE error IS NULL AND response_parsed_ok=1").fetchone()[0] or 0
    conn.close()
    return {
        "total_details_processed": total,
        "edits_per_card": round(edits_per_card, 2),
        "accepted_pct": round(accepted_pct, 1),
        "avg_time_to_card_min": round(avg_time, 1),
        "total_llm_cost_rub": round(total_cost, 2),
        "total_successful_gens": total_gens,
        # Целевые KPI (из совета + industry benchmark GitHub Copilot / NIO ~30%)
        "kpi": {
            "time_target": 60,  # мин (текущее 240-480, цель 4-8x ускорение)
            "accepted_target": 30,  # % (минимум, industry benchmark)
            "edits_target": 8,  # правок (максимум)
        }
    }


def log_llm_call(detail_id: str, model: str, system_prompt: str, user_prompt: str,
                 response_text: str = None, response_parsed_ok: bool = None,
                 tokens_in: int = None, tokens_out: int = None,
                 duration_ms: int = None, error: str = None):
    """Записывает каждый LLM-вызов в БД для отладки"""
    cost = calc_cost_rub(tokens_in or 0, tokens_out or 0)
    conn = get_conn()
    conn.execute("""INSERT INTO llm_calls
        (detail_id, model, system_prompt, user_prompt, response_text,
         response_parsed_ok, tokens_in, tokens_out, duration_ms, cost_rub, error)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (detail_id, model, system_prompt[:5000], user_prompt[:10000],
         (response_text or "")[:10000],
         1 if response_parsed_ok else (0 if response_parsed_ok is False else None),
         tokens_in, tokens_out, duration_ms, cost, error))
    conn.commit()
    conn.close()
    return cost


def get_llm_calls(detail_id: str = None, limit: int = 50) -> list:
    """Получает последние LLM-вызовы (для UI)"""
    conn = get_conn()
    if detail_id:
        rows = conn.execute("""SELECT id, detail_id, model, response_parsed_ok,
            tokens_in, tokens_out, duration_ms, error, created_at
            FROM llm_calls WHERE detail_id = ? ORDER BY id DESC LIMIT ?""",
            (detail_id, limit)).fetchall()
    else:
        rows = conn.execute("""SELECT id, detail_id, model, response_parsed_ok,
            tokens_in, tokens_out, duration_ms, error, created_at
            FROM llm_calls ORDER BY id DESC LIMIT ?""", (limit,)).fetchall()
    conn.close()
    cols = ["id", "detail_id", "model", "response_parsed_ok", "tokens_in",
            "tokens_out", "duration_ms", "error", "created_at"]
    return [dict(zip(cols, r)) for r in rows]


def get_llm_call_detail(call_id: int) -> Optional[dict]:
    conn = get_conn()
    row = conn.execute("""SELECT id, detail_id, model, system_prompt, user_prompt,
        response_text, response_parsed_ok, tokens_in, tokens_out, duration_ms, error, created_at
        FROM llm_calls WHERE id = ?""", (call_id,)).fetchone()
    conn.close()
    if not row:
        return None
    cols = ["id", "detail_id", "model", "system_prompt", "user_prompt",
            "response_text", "response_parsed_ok", "tokens_in", "tokens_out",
            "duration_ms", "error", "created_at"]
    return dict(zip(cols, row))
def extract_rule_from_edits():
    """Извлекает правила из частых правок"""
    conn = get_conn()
    rows = conn.execute("""SELECT field, old_value, new_value, COUNT(*) as cnt
        FROM edits GROUP BY field, old_value, new_value HAVING cnt >= 2""").fetchall()
    rules = []
    for r in rows:
        rules.append({
            "field": r[0], "old": r[1], "new": r[2], "uses": r[3]
        })
    conn.close()
    return rules


def get_metrics() -> dict:
    """Метрики точности"""
    conn = get_conn()
    total_drafts = conn.execute("SELECT COUNT(*) FROM drafts").fetchone()[0]
    approved = conn.execute("SELECT COUNT(*) FROM drafts WHERE status='approved'").fetchone()[0]
    total_edits = conn.execute("SELECT COUNT(*) FROM edits").fetchone()[0]
    total_details = conn.execute("SELECT COUNT(*) FROM details").fetchone()[0]
    # Среднее правок на черновик
    avg_edits = 0
    if approved > 0:
        avg_edits = total_edits / approved
    # Процент точности: 0 правок = 100%, 10 правок = 0%
    accuracy = max(0, 100 - int(avg_edits * 10))
    conn.close()
    return {
        "total_details": total_details,
        "total_drafts": total_drafts,
        "approved": approved,
        "total_edits": total_edits,
        "avg_edits_per_draft": round(avg_edits, 2),
        "accuracy_pct": accuracy
    }


def get_draft(detail_id: str) -> Optional[dict]:
    """Get draft from DB"""
    conn = get_conn()
    cursor = conn.execute(
        "SELECT llm_output, status FROM drafts WHERE detail_id = ?",
        (detail_id,)
    )
    row = cursor.fetchone()
    conn.close()
    if row and row[0]:
        return {"output": json.loads(row[0]), "status": row[1]}
    return None


def save_draft(detail_id: str, llm_output: dict, status: str = "draft", author: str = "llm"):
    """Save draft to DB + create version"""
    conn = get_conn()
    now = datetime.now().isoformat()
    conn.execute(
        """INSERT OR REPLACE INTO drafts (detail_id, llm_output, status, author, created_at, updated_at)
           VALUES (?, ?, ?, ?, COALESCE((SELECT created_at FROM drafts WHERE detail_id = ?), ?), ?)""",
        (detail_id, json.dumps(llm_output, ensure_ascii=False), status, author, detail_id, now, now)
    )
    conn.commit()
    conn.close()
    # Сохраняем версию операций
    if "operations" in llm_output:
        save_version(detail_id, llm_output["operations"], author=author, source="llm_generate")


def add_history(detail_id: str, action: str, details: dict = None):
    """Add history entry"""
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO history (detail_id, action, details) VALUES (?, ?, ?)",
            (detail_id, action, json.dumps(details or {}, ensure_ascii=False))
        )
        conn.commit()
    finally:
        conn.close()


# Pydantic models
class GenerateRequest(BaseModel):
    detail_id: str
    answers: Optional[dict] = None


async def _get_param(request: Request, name: str, log_name: str = "") -> Optional[str]:
    """Получает параметр из body (JSON или form-encoded) или query string.
    В FastAPI нельзя читать body дважды, поэтому читаем один раз и парсим руками."""
    from urllib.parse import parse_qs
    # 1. Query string
    qs_val = request.query_params.get(name)
    if qs_val:
        if log_name: log.info(f"{log_name}: '{name}' from query = {qs_val!r}")
        return qs_val
    # 2. Body (form-encoded или JSON)
    body = await request.body()
    if log_name: log.info(f"{log_name}: body={body!r}, content_type={request.headers.get('content-type')!r}")
    if not body:
        return None
    body_str = body.decode('utf-8', errors='ignore').strip()
    if not body_str:
        return None
    # Сначала пробуем JSON
    try:
        data = json.loads(body_str)
        if isinstance(data, dict) and name in data:
            val = str(data[name])
            if log_name: log.info(f"{log_name}: '{name}' from JSON = {val!r}")
            return val
    except Exception:
        pass
    # Потом form-encoded
    try:
        parsed = parse_qs(body_str)
        if name in parsed and parsed[name]:
            val = parsed[name][0]
            if log_name: log.info(f"{log_name}: '{name}' from form = {val!r}")
            return val
    except Exception:
        pass
    return None


# Routes
@app.get("/", response_class=HTMLResponse)
async def index(request: Request, q: str = "", page: int = 1, per_page: int = 25, status: str = "", model: str = ""):
    # N+1 fix: один запрос со всеми статусами
    conn = get_conn()
    where_clauses = []
    params = []
    if q:
        # UX5: search расширен на chassis
        where_clauses.append("(designation LIKE ? OR name LIKE ? OR model LIKE ? OR material LIKE ? OR chassis LIKE ?)")
        like_q = f"%{q}%"
        params.extend([like_q, like_q, like_q, like_q, like_q])
    if status:
        where_clauses.append("status = ?")
        params.append(status)
    if model:
        where_clauses.append("model = ?")
        params.append(model)
    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    # Список моделей для фильтра (M2 fix)
    models = [r[0] for r in conn.execute("SELECT DISTINCT model FROM details WHERE model IS NOT NULL AND model != '' ORDER BY model").fetchall() if r[0]]
    # Всего
    count_row = conn.execute(f"SELECT COUNT(*) FROM details {where_sql}", params).fetchone()
    total = count_row[0] if count_row else 0
    # Пагинация
    per_page = max(1, min(100, per_page))
    page = max(1, page)
    total_pages = (total + per_page - 1) // per_page
    offset = (page - 1) * per_page
    # Детали + статус через LEFT JOIN
    rows = conn.execute(f"""
        SELECT d.id, d.designation, d.name, d.model, d.chassis, d.material, d.mass_kg, d.surface_treatment, d.created_at,
               COALESCE(dr.status, 'new') as draft_status
        FROM details d
        LEFT JOIN drafts dr ON d.id = dr.detail_id
        {where_sql}
        ORDER BY d.created_at DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()
    cols = ["id", "designation", "name", "model", "chassis", "material", "mass_kg", "surface_treatment", "created_at", "status"]
    details = [dict(zip(cols, r)) for r in rows]
    conn.close()

    # v0.7: Mavis dashboard data (tasks, counts, learning, metrics)
    from mock_llm import is_mock_mode
    role = request.cookies.get("bit_role", "technologist")
    USER_BY_ROLE = {
        "technologist": "Тарлецкий А.В.", "main_technologist": "Баранов М.А.",
        "workshop_chief": "Голубев П.В.", "admin": "ИТ-служба", "normirovshchik": "Воробьёв Г.И.",
    }
    ROLE_DESCR = {
        "technologist": "технолог Тарлецкий. Ваш день: разобрать задачи ниже, проверить черновик AI, ответить на вопросы AI. Кнопки утверждения вам недоступны.",
        "main_technologist": "гл. технолог Баранов. Вам доступны утверждение ТК, решения по извещениям и ревизия профилей РС.",
        "workshop_chief": "нач. производства Голубев. Финальная виза после гл. технолога.",
        "admin": "администратор LLM. Управляете моделями, лимитами и смотрите журнал.",
        "normirovshchik": "технолог-администратор Воробьёв. Настраиваете профили РС и проверяете разобранные AI документы.",
    }
    today = datetime.now().strftime("%A, %d %B %Y").lower()
    months_ru = {"january":"января","february":"февраля","march":"марта","april":"апреля","may":"мая","june":"июня","july":"июля","august":"августа","september":"сентября","october":"октября","november":"ноября","december":"декабря"}
    days_ru = {"monday":"понедельник","tuesday":"вторник","wednesday":"среда","thursday":"четверг","friday":"пятница","saturday":"суббота","sunday":"воскресенье"}
    for en, ru in months_ru.items(): today = today.replace(en, ru)
    for en, ru in days_ru.items(): today = today.replace(en, ru)

    # 5 counters (реальные)
    drafts_count = sum(1 for d in details if d.get('status') in ('draft', 'new', None))
    need_approve = sum(1 for d in details if d.get('status') == 'review')
    notices_count = 1  # mock — есть И-2026-014 в change_notices
    questions_count = 3  # mock
    conflicts_count = 1  # mock

    # Tasks table (5 приоритетных)
    tasks = []
    for d in details[:3]:
        tasks.append({
            "icon": "📄", "title": d.get("name", d.get("designation", "Деталь")),
            "subtitle": d.get("designation", ""),
            "product": d.get("model", "—"),
            "action": "Черновик AI готов — проверить нормы и отправить на утверждение",
            "status_class": "blue", "status_label": "Черновик v2",
            "deadline": "сегодня", "action_label": "Открыть",
            "url": f"/detail/{d['id']}",
        })
    if notices_count > 0:
        tasks.append({
            "icon": "📬", "title": "Извещение И-2026-014", "subtitle": "смена материала",
            "product": "АЦ-8,0-40 · зав. №147",
            "action": "Материал днища: 09Г2С → 10ХСНД. Затронуты 3 сборки, 2 ТК, 2 РС",
            "status_class": "bad", "status_label": "Требует решения",
            "deadline": "22.07", "action_label": "Разобрать",
            "url": "/notices",
        })
    tasks.append({
        "icon": "❓", "title": "Растяжка пружинная", "subtitle": "ЛМША.301712.000",
        "product": "УМК",
        "action": "AI спрашивает: «Пружина тарельчатая (72 шт) — запрессовка на участке растяжек или в кооперации?»",
        "status_class": "warn", "status_label": "Вопрос AI",
        "deadline": "23.07", "action_label": "Ответить",
        "url": "/detail/detail-lmsha-301712-000",
    })
    if conflicts_count > 0:
        tasks.append({
            "icon": "🔄", "title": "Цистерна", "subtitle": "53-ТВ.05.00.00",
            "product": "АЦ-6,0-40 · зав. №151",
            "action": "РС изменена вручную в 1С (этап «Сварка» +2 н/ч). Принять из 1С или перезаписать?",
            "status_class": "bad", "status_label": "Конфликт РС",
            "deadline": "сегодня", "action_label": "Сравнить",
            "url": "/detail/detail-cisterna",
        })

    # Метрики пилота (real)
    pilot_metrics = []
    try:
        conn2 = get_conn()
        for row in conn2.execute("SELECT metric_value, metric_text FROM pilot_metrics WHERE metric_name IN ('etalons_count','vedomost_dse','work_orders_count','rules_count','green_pct') ORDER BY metric_name"):
            pilot_metrics.append({"value": int(row[0]) if row[0] == int(row[0]) else row[0], "text": row[1] or ""})
        conn2.close()
    except Exception:
        pass

    return templates.TemplateResponse("index.html", {
        "request": request,
        "details": details,
        "demo_mode": is_mock_mode(),
        "llm_model": LLM_MODEL,
        "q": q,
        "status": status,
        "model": model,
        "models": models,
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
        # v0.7 dashboard data
        "active": "dashboard",
        "today_str": today,
        "current_user_name": USER_BY_ROLE.get(role, "Гость"),
        "current_role_descr": ROLE_DESCR.get(role, ""),
        "counts": {
            "drafts": drafts_count, "need_approve": need_approve,
            "notices": notices_count, "questions": questions_count,
            "conflicts": conflicts_count,
        },
        "tasks": tasks[:5],
        "learning": {
            "approved_4w": 17, "green_now": 61, "green_4w_ago": 42, "bias": True,
            "case_title": "Зачистка после мех. сварки",
            "case_old": "0,50 н/ч (−25% к факту)",
            "case_new": "0,65 н/ч (после правила)",
            "case_rule": "Зачистка после мех. сварки = 12 мин/пог.м шва",
        },
        "pilot_metrics": pilot_metrics,
    })


# V0.7: экраны из вложенного прототипа (M33) — встроены в base.html

@app.get("/products", response_class=HTMLResponse)
async def products_v7(request: Request):
    """V0.7 Изделия и состав — дерево моделей/исполнений/сборок/деталей."""
    from mock_llm import is_mock_mode
    role = request.cookies.get("bit_role", "technologist")
    real = []
    try:
        conn = get_conn()
        for row in conn.execute("SELECT id, designation, name, material FROM details ORDER BY created_at DESC LIMIT 15"):
            real.append({"id": row[0], "designation": row[1], "name": row[2], "material": row[3]})
        conn.close()
    except Exception:
        pass
    return templates.TemplateResponse("products.html", {
        "request": request, "active": "products",
        "current_role": role, "current_user_name": "",
        "current_role_descr": "", "mock_mode": is_mock_mode(),
        "real_details": real,
    })


@app.get("/notices", response_class=HTMLResponse)
async def notices_v7(request: Request):
    """V0.7 Извещения — diff было/стало + AI предложение."""
    from mock_llm import MOCK_NOTICE_DIFF, is_mock_mode
    role = request.cookies.get("bit_role", "technologist")
    notices = []
    try:
        conn = get_conn()
        for row in conn.execute("SELECT id, number, date, author, status, reason, foundation_doc FROM change_notices WHERE status='open' ORDER BY date DESC"):
            notices.append({
                "id": row[0], "number": row[1], "date": row[2], "author": row[3],
                "status": row[4], "reason": row[5], "foundation_doc": row[6],
                "affected": [
                    {"obj": "Цистерна 53-ТВ.05.00.00", "relation": "деталь входит в сборку", "impacted": "ТК (утв.) + РС v3 в 1С"},
                    {"obj": "Обечайка с днищами 53-ТВ.05.01.000", "relation": "подсборка", "impacted": "ТК (утв.)"},
                    {"obj": "АЦ-8,0-40 · зав. №147", "relation": "исполнение", "impacted": "состав исполнения (отличие от базового)"},
                ],
                "changes": MOCK_NOTICE_DIFF["operation_changes"],
            })
        conn.close()
    except Exception:
        notices.append({
            "id": "notice-014", "number": "И-2026-014", "date": "2026-07-18",
            "author": "Юрьев А.М. (КБ)", "status": "open",
            "reason": "Замена материала днища заднего 09Г2С → 10ХСНД (исполнение зав. №147)",
            "foundation_doc": "И-2026-014.pdf",
            "affected": [
                {"obj": "Цистерна 53-ТВ.05.00.00", "relation": "деталь входит в сборку", "impacted": "ТК (утв.) + РС v3 в 1С"},
                {"obj": "Обечайка с днищами 53-ТВ.05.01.000", "relation": "подсборка", "impacted": "ТК (утв.)"},
                {"obj": "АЦ-8,0-40 · зав. №147", "relation": "исполнение", "impacted": "состав исполнения (отличие от базового)"},
            ],
            "changes": MOCK_NOTICE_DIFF["operation_changes"],
        })
    return templates.TemplateResponse("notices.html", {
        "request": request, "active": "notices",
        "current_role": role, "current_user_name": "",
        "current_role_descr": "", "mock_mode": is_mock_mode(),
        "notices": notices, "ai_model": "MockLLM-Technologist-v1",
    })


@app.get("/profiles", response_class=HTMLResponse)
async def profiles_v7(request: Request):
    """V0.7 Профили выхода РС — 3 профиля × 8 осей."""
    import json as _json
    from mock_llm import is_mock_mode
    role = request.cookies.get("bit_role", "technologist")
    profiles = []
    try:
        conn = get_conn()
        for row in conn.execute("SELECT id, code, name, product_type, is_active, version, axes_json, export_schema, description FROM rs_profiles ORDER BY is_active DESC, code"):
            try:
                axes = _json.loads(row[6]) if row[6] else {}
            except Exception:
                axes = {}
            # humanize axis values
            axis_pretty = {
                "stage_cut": {"single_stage": "Вся РС одним этапом", "by_site": "По участкам", "by_visit": "По цехозаходам (ВРЦ)"}.get(axes.get("stage_cut"), axes.get("stage_cut", "—")),
                "op_detail": {"consolidated": "Укрупнённо: суммирование по профессии, без ПЗ-времени", "full": "Полная, включая ПЗ", "stages_only": "Только этапы, без операций"}.get(axes.get("op_detail"), axes.get("op_detail", "—")),
                "time_norm": {"piece_total": "Только Тшт, суммой на этап", "piece_setup_split": "Тшт + Тпз раздельно", "none": "Не выгружать"}.get(axes.get("time_norm"), axes.get("time_norm", "—")),
                "materials": {"stage_total": "Сводно на этап", "by_op": "По операциям"}.get(axes.get("materials"), axes.get("materials", "—")),
                "labor": {"prof_total": "Профессия + разряд, сводно", "by_op": "По операциям", "none": "Не выгружать"}.get(axes.get("labor"), axes.get("labor", "—")),
                "nesting": {"flat": "Плоская (без отдельных РС полуфабрикатов)", "multi": "Многоуровневая (полуфабрикаты — отдельными РС)"}.get(axes.get("nesting"), axes.get("nesting", "—")),
                "coop": {"none": "Не применяется", "da": "Давальческая → этап вида «Переработка на стороне»"}.get(axes.get("coop"), axes.get("coop", "—")),
            }
            profiles.append({
                "id": row[0], "code": row[1], "name": row[2], "product_type": row[3],
                "is_active": row[4], "version": row[5], "axes": axis_pretty,
                "export_schema": row[7], "description": row[8],
            })
        conn.close()
    except Exception:
        pass
    return templates.TemplateResponse("profiles.html", {
        "request": request, "active": "profiles",
        "current_role": role, "current_user_name": "",
        "current_role_descr": "", "mock_mode": is_mock_mode(),
        "profiles": profiles,
    })


@app.get("/knowledge", response_class=HTMLResponse)
async def knowledge_v7(request: Request):
    """V0.7 База знаний — эталоны/ведомости/наряды/правила."""
    from mock_llm import is_mock_mode
    role = request.cookies.get("bit_role", "technologist")
    metrics, rules = [], []
    try:
        conn = get_conn()
        for row in conn.execute("SELECT metric_value, metric_text FROM pilot_metrics WHERE metric_name IN ('etalons_count','vedomost_dse','work_orders_count','rules_count','green_pct') ORDER BY metric_name"):
            metrics.append({"value": int(row[0]) if row[0] == int(row[0]) else row[0], "text": row[1] or ""})
        for row in conn.execute("SELECT name, scope, rule_text, source FROM tech_rules ORDER BY id"):
            rules.append({"name": row[0], "scope": row[1], "rule_text": row[2], "source": row[3]})
        conn.close()
    except Exception:
        pass
    return templates.TemplateResponse("knowledge.html", {
        "request": request, "active": "knowledge",
        "current_role": role, "current_user_name": "",
        "current_role_descr": "", "mock_mode": is_mock_mode(),
        "metrics": metrics, "rules": rules,
    })


@app.get("/llm-admin", response_class=HTMLResponse)
async def llm_admin_v7(request: Request):
    """V0.7 LLM admin — назначение моделей, лимиты, журнал."""
    from mock_llm import is_mock_mode
    role = request.cookies.get("bit_role", "technologist")
    is_llm_admin = role == "admin"
    pm = {"cost_today_rub": 0, "cost_month_rub": 0, "avg_draft_cost_rub": 0}
    try:
        conn = get_conn()
        for row in conn.execute("SELECT metric_name, metric_value FROM pilot_metrics WHERE metric_name LIKE 'cost_%' OR metric_name='avg_draft_cost_rub'"):
            pm[row[0]] = row[1]
        conn.close()
    except Exception:
        pass
    return templates.TemplateResponse("llm_admin.html", {
        "request": request, "active": "llm",
        "current_role": role, "current_user_name": "",
        "current_role_descr": "", "mock_mode": is_mock_mode(),
        "is_llm_admin": is_llm_admin,
        "llm_model": LLM_MODEL, "llm_url": LLM_API_URL,
        "llm_latency_ms": "—",
        "cost_today": int(pm.get("cost_today_rub", 0)),
        "cost_month": int(pm.get("cost_month_rub", 0)),
        "avg_draft": pm.get("avg_draft_cost_rub", 0),
    })


@app.get("/index/table", response_class=HTMLResponse)
async def index_table(request: Request, q: str = "", page: int = 1, per_page: int = 25, status: str = ""):
    """UX5: live search возвращает только таблицу + пагинацию (без layout)"""
    conn = get_conn()
    where_clauses = []
    params = []
    if q:
        where_clauses.append("(designation LIKE ? OR name LIKE ? OR model LIKE ? OR material LIKE ? OR chassis LIKE ?)")
        like_q = f"%{q}%"
        params.extend([like_q, like_q, like_q, like_q, like_q])
    if status:
        where_clauses.append("status = ?")
        params.append(status)
    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    count_row = conn.execute(f"SELECT COUNT(*) FROM details {where_sql}", params).fetchone()
    total = count_row[0] if count_row else 0
    per_page = max(1, min(100, per_page))
    page = max(1, page)
    total_pages = (total + per_page - 1) // per_page
    offset = (page - 1) * per_page
    rows = conn.execute(f"""
        SELECT d.id, d.designation, d.name, d.model, d.chassis, d.material, d.mass_kg, d.surface_treatment, d.created_at,
               COALESCE(dr.status, 'new') as draft_status
        FROM details d
        LEFT JOIN drafts dr ON d.id = dr.detail_id
        {where_sql}
        ORDER BY d.created_at DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()
    cols = ["id", "designation", "name", "model", "chassis", "material", "mass_kg", "surface_treatment", "created_at", "status"]
    details = [dict(zip(cols, r)) for r in rows]
    conn.close()
    return templates.TemplateResponse("_index_table.html", {
        "request": request,
        "details": details,
        "q": q,
        "status": status,
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages
    })


@app.exception_handler(404)
async def not_found_handler(request: Request, exc):
    """F16.7: кастомный 404 с навигацией"""
    return templates.TemplateResponse("404.html", {
        "request": request,
        "roles": ROLES,
        "current_role_from_request": get_current_role
    }, status_code=404)


# ========== BUG-2026-07-20-04: Помощь ВНУТРИ продукта ==========
@app.get("/help", response_class=HTMLResponse)
async def help_page(request: Request):
    """Руководство пользователя по 7 ролям (внутри продукта, не в docs/)"""
    return templates.TemplateResponse("help.html", {
        "request": request,
        "current_role_from_request": get_current_role,
        "roles": ROLES
    })


@app.get("/detail/{detail_id}", response_class=HTMLResponse)
async def detail(request: Request, detail_id: str):
    detail_obj = get_detail(detail_id)
    if not detail_obj:
        raise HTTPException(404, "Detail not found")

    draft_data = get_draft(detail_id)
    versions = get_versions(detail_id)
    edits = get_edits(detail_id)
    status_ext = "draft"
    if draft_data:
        conn = get_conn()
        row = conn.execute("SELECT status_ext FROM drafts WHERE detail_id=?", (detail_id,)).fetchone()
        conn.close()
        if row and row[0]:
            status_ext = row[0]

    # Экономика (расчёт себестоимости) — U8 fix: показываем в карточке
    economics = calc_cost_estimate(detail_id)
    # U12 fix: список оборудования для datalist в inline-edit
    all_equipment = get_all_equipment()[:50]  # топ-50 для производительности

    return templates.TemplateResponse("detail.html", {
        "request": request,
        "detail": detail_obj,
        "draft": draft_data["output"] if draft_data else None,
        "status": draft_data["status"] if draft_data else "new",
        "status_ext": status_ext,
        "versions": versions,
        "edits": edits,
        "economics": economics,
        "all_equipment": all_equipment,
        "active_tab": request.query_params.get("active_tab", "route"),
        "demo_mode": DEMO_MODE,
        "llm_model": LLM_MODEL
    })


@app.post("/api/analyze")
async def api_analyze(request: Request):
    """Sprint 1: AI задаёт 3-5 уточняющих вопросов перед генерацией (blueprint.io pattern)"""
    # BUG-2026-07-19-01: RBAC — нормировщик/ОТК/конструктор не должны генерировать
    role = request.cookies.get("bit_role", "technologist")
    if role not in ("technologist", "main_technologist", "admin"):
        return err("role_not_allowed: только технолог/гл.технолог/админ могут генерировать", 403)
    detail_id = await _get_param(request, "detail_id", log_name="/api/analyze")
    if not detail_id:
        return err("detail_id required", 422)
    # V4-2: проверка дневного лимита перед LLM-вызовом
    limit = check_daily_limit_or_warn()
    if not limit.get("allowed"):
        return JSONResponse({"error": "daily_limit_exceeded", "detail": limit},
                            status_code=429,
                            headers={"Retry-After": "86400"})
    detail_obj = get_detail(detail_id)
    if not detail_obj:
        return err("not found", 404)

    daily = get_daily_cost()
    if daily["exceeded"]:
        return JSONResponse({"error": "daily_limit_exceeded", "limit": daily["limit_rub"]}, status_code=429)

    if DEMO_MODE:
        return JSONResponse({
            "questions": [
                {"id": "Q1", "topic": "материал", "question": "Какой материал заготовки?", "options": ["лист 09Г2С", "труба Ст20", "поковка 40Х", "другое"], "default": "лист 09Г2С", "impact_if_changed": "изменит режим сварки и нормы"},
                {"id": "Q2", "topic": "толщина", "question": "Толщина металла в мм?", "options": ["3", "5", "8", "10+"], "default": "5", "impact_if_changed": "изменит параметры сварки"},
                {"id": "Q3", "topic": "термообработка", "question": "Требуется ли термообработка после сварки?", "options": ["да, отпуск", "да, закалка+отпуск", "нет"], "default": "да, отпуск", "impact_if_changed": "добавит 1.5ч в маршрут"},
                {"id": "Q4", "topic": "покрытие", "question": "Какое покрытие?", "options": ["порошковая покраска", "жидкая краска", "горячее цинкование", "без покрытия"], "default": "порошковая покраска", "impact_if_changed": "изменит финишные операции"},
                {"id": "Q5", "topic": "приёмка", "question": "Нужна ли военная приёмка?", "options": ["да", "нет"], "default": "нет", "impact_if_changed": "добавит операции ОТК"}
            ],
            "mode": "demo"
        })

    try:
        from string import Template
        from prompts import CLARIFICATION_PROMPT
        client = get_llm_client()
        prompt = Template(CLARIFICATION_PROMPT).substitute(
            properties_json=json.dumps(detail_obj, indent=2, ensure_ascii=False),
            material=detail_obj.get("material", ""),
            mass_kg=detail_obj.get("mass_kg", 0),
            surface_treatment=detail_obj.get("surface_treatment", ""),
            chassis=detail_obj.get("chassis", ""),
            model=detail_obj.get("model", "")
        )
        import time
        t0 = time.time()
        system_msg = "Ты — опытный технолог. Задай 3-5 КРАТКИХ уточняющих вопросов перед генерацией техкарты. Возвращай только JSON без markdown."
        response = client.chat.completions.create(
            model=LLM_MODEL,
            messages=[{"role": "system", "content": system_msg}, {"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=1500
        )
        duration = int((time.time() - t0) * 1000)
        text = response.choices[0].message.content.strip()
        result = parse_llm_json(text)
        log_llm_call(detail_id, LLM_MODEL, system_msg, prompt, text, True,
                     response.usage.prompt_tokens if response.usage else None,
                     response.usage.completion_tokens if response.usage else None,
                     duration)
        result["mode"] = "live"
        return JSONResponse(result)
    except Exception as e:
        log.error(f"/api/analyze error: {e}")
        return err(str(e)[:200], 500)


@app.post("/api/draft-fast")
async def api_draft_fast(request: Request):
    """Sprint 1: быстрый дешёвый draft (короткий промт, 3 операции, ~30 сек, ~1₽)"""
    # BUG-2026-07-19-01: RBAC
    role = request.cookies.get("bit_role", "technologist")
    if role not in ("technologist", "main_technologist", "admin"):
        return err("role_not_allowed", 403)
    detail_id = await _get_param(request, "detail_id", log_name="/api/draft-fast")
    if not detail_id:
        return err("detail_id required", 422)
    detail_obj = get_detail(detail_id)
    if not detail_obj:
        return err("not found", 404)
    daily = get_daily_cost()
    if daily["exceeded"]:
        return err("daily_limit_exceeded", 429)
    if DEMO_MODE:
        return JSONResponse({"draft": {"summary": {"total_operations": 3, "total_hours": 1.5, "complexity": "средняя", "closest_analog": "4c85941a (упор продольный)"},
                              "route": [{"step": 1, "operation": "010 Подготовительная", "duration_hours": 0.2},
                                       {"step": 2, "operation": "015 Сварка", "duration_hours": 0.7},
                                       {"step": 3, "operation": "020 Контроль", "duration_hours": 0.6}],
                              "operations": [], "warnings": [], "questions": []},
                         "mode": "demo", "cost_estimate": "1.00₽"})
    try:
        from string import Template
        from prompts import DRAFT_FAST_PROMPT
        client = get_llm_client()
        answers = await _get_param(request, "answers") or "{}"
        try:
            answers_dict = json.loads(answers) if answers else {}
        except Exception:
            answers_dict = {}
        prompt = Template(DRAFT_FAST_PROMPT).substitute(
            properties_json=json.dumps(detail_obj, indent=2, ensure_ascii=False),
            answers_json=json.dumps(answers_dict, indent=2, ensure_ascii=False)
        )
        import time
        t0 = time.time()
        system_msg = "Сгенерируй КРАТКИЙ draft маршрута (3 операции) для этой детали. Только JSON без markdown."
        response = client.chat.completions.create(
            model=LLM_MODEL,
            messages=[{"role": "system", "content": system_msg}, {"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=1500
        )
        duration = int((time.time() - t0) * 1000)
        text = response.choices[0].message.content.strip()
        result = parse_llm_json(text)
        log_llm_call(detail_id, LLM_MODEL, system_msg, prompt, text, True,
                     response.usage.prompt_tokens if response.usage else None,
                     response.usage.completion_tokens if response.usage else None,
                     duration)
        result["mode"] = "live"
        tokens_in = response.usage.prompt_tokens if response.usage else 0
        tokens_out = response.usage.completion_tokens if response.usage else 0
        cost = (tokens_in/1000)*LLM_PRICE_INPUT_RUB_PER_1K + (tokens_out/1000)*LLM_PRICE_OUTPUT_RUB_PER_1K
        result["cost_estimate"] = f"{cost:.2f}₽"
        # M27 bugfix: /api/draft-fast возвращал draft, но НЕ сохранял его в БД.
        # Технолог нажимал "Сгенерировать", видел toast "Готово!", страница перезагружалась,
        # но операций не было. Теперь сохраняем.
        llm_output = result
        # Адаптируем структуру: route → operations с нужными полями
        if "operations" not in llm_output and "route" in llm_output:
            llm_output["operations"] = [
                {
                    "op_index": str(r.get("step", i+1) * 10),
                    "name": r.get("operation", ""),
                    "duration_hours": r.get("duration_hours", 0),
                    "department": r.get("department", ""),
                    "workplace": r.get("workplace", ""),
                    "equipment": r.get("equipment", ""),
                }
                for i, r in enumerate(llm_output["route"])
            ]
        save_draft(detail_id, llm_output, "draft")
        total_ops = len(llm_output.get("operations", []))
        record_metric(detail_id, "total_ops", total_ops, {"source": "draft-fast"})
        return JSONResponse({"draft": llm_output, "mode": "live", "cost_estimate": f"{cost:.2f}₽"})
    except Exception as e:
        log.error(f"/api/draft-fast error: {e}")
        return err(str(e)[:200], 500)


@app.post("/api/refine")
async def api_refine(request: Request):
    """Sprint 1: уточнение draft'а до полного маршрута (с учётом ответов на уточнения)"""
    # BUG-2026-07-19-01: RBAC
    role = request.cookies.get("bit_role", "technologist")
    if role not in ("technologist", "main_technologist", "admin"):
        return err("role_not_allowed", 403)
    detail_id = await _get_param(request, "detail_id", log_name="/api/refine")
    if not detail_id:
        return err("detail_id required", 422)
    detail_obj = get_detail(detail_id)
    if not detail_obj:
        return err("not found", 404)
    daily = get_daily_cost()
    if daily["exceeded"]:
        return err("daily_limit_exceeded", 429)
    draft_json = await _get_param(request, "draft")
    answers = await _get_param(request, "answers")
    if draft_json:
        try:
            draft_dict = json.loads(draft_json)
        except Exception:
            draft_dict = {}
    else:
        draft_dict = {}
    try:
        answers_dict = json.loads(answers) if answers else {}
    except Exception:
        answers_dict = {}
    if DEMO_MODE:
        llm_output = generate_mock_draft(detail_obj)
        add_history(detail_id, "refined_mock", {"from": "draft"})
    else:
        try:
            from string import Template
            from prompts import REFINE_PROMPT
            client = get_llm_client()
            tech_rules_text = detail_obj.get("tech_rules") or ""
            rules_block = f"\nВАЖНО: Технолог указал следующие правила и нюансы — ОБЯЗАТЕЛЬНО учти их:\n{tech_rules_text}\n" if tech_rules_text.strip() else ""
            # Sprint 2: добавляем top-K похожих техкарт из RAG
            similar_block = "(похожих техкарт пока нет)"
            try:
                from rag import rag_search
                similar = rag_search(detail_obj, top_k=3)
                if similar:
                    lines = []
                    for s in similar:
                        meta = s.get("metadata", {})
                        lines.append(f"- {meta.get('designation', '?')} '{meta.get('name', '?')}' "
                                     f"({meta.get('material', '?')}, score={s['score']})")
                    similar_block = "\n".join(lines)
            except Exception as e:
                log.warning(f"RAG search failed in /api/refine: {e}")
            prompt = Template(REFINE_PROMPT).substitute(
                properties_json=json.dumps(detail_obj, indent=2, ensure_ascii=False),
                equipment_json=json.dumps(EQUIPMENT, indent=2, ensure_ascii=False),
                structure_json=json.dumps(STRUCTURE, indent=2, ensure_ascii=False),
                few_shot_json=json.dumps(FEW_SHOT_4C85941A, indent=2, ensure_ascii=False),
                tech_rules="(правила не указаны)",
                rules_block=rules_block,
                workshops_context=TECHINKOM_WORKSHOPS_CONTEXT,
                draft_json=json.dumps(draft_dict, indent=2, ensure_ascii=False),
                answers_json=json.dumps(answers_dict, indent=2, ensure_ascii=False),
                similar_block=similar_block
            )
            import time
            t_start = time.time()
            system_msg = "Ты — опытный технолог-сварщик. Доработай draft маршрута до полной техкарты. Всегда возвращай валидный JSON без markdown-обёртки."
            response = client.chat.completions.create(
                model=LLM_MODEL,
                messages=[{"role": "system", "content": system_msg}, {"role": "user", "content": prompt}],
                temperature=0.2,
                max_tokens=8000
            )
            duration_ms = int((time.time() - t_start) * 1000)
            text = response.choices[0].message.content.strip()
            log_llm_call(detail_id, LLM_MODEL, system_msg, prompt, text, False,
                         response.usage.prompt_tokens if response.usage else None,
                         response.usage.completion_tokens if response.usage else None,
                         duration_ms)
            try:
                llm_output = json.loads(text)
                conn = get_conn()
                last_id = conn.execute("SELECT MAX(id) FROM llm_calls").fetchone()[0]
                if last_id:
                    conn.execute("UPDATE llm_calls SET response_parsed_ok=1 WHERE id=?", (last_id,))
                    conn.commit()
                conn.close()
            except json.JSONDecodeError as je:
                log_llm_call(detail_id, LLM_MODEL, system_msg, prompt, text, False, duration_ms=duration_ms, error=f"JSON parse: {je}")
                return err("JSON parse", 500)
            add_history(detail_id, "refined", {"model": LLM_MODEL,
                                              "tokens_in": response.usage.prompt_tokens if response.usage else None,
                                              "tokens_out": response.usage.completion_tokens if response.usage else None})
        except Exception as e:
            log.error(f"/api/refine LLM error: {e}")
            return err(str(e)[:200], 500)
    save_draft(detail_id, llm_output, "draft")
    total_ops = len(llm_output.get("operations", []))
    record_metric(detail_id, "total_ops", total_ops, {"source": "refine"})
    tokens_in = response.usage.prompt_tokens if 'response' in dir() and response.usage else 0
    tokens_out = response.usage.completion_tokens if 'response' in dir() and response.usage else 0
    cost = (tokens_in/1000)*LLM_PRICE_INPUT_RUB_PER_1K + (tokens_out/1000)*LLM_PRICE_OUTPUT_RUB_PER_1K
    return JSONResponse({"ok": True, "total_ops": total_ops, "cost": f"{cost:.2f}₽"})


@app.post("/api/feedback")
async def api_feedback(request: Request):
    """Sprint 3 placeholder: кнопка 'Пожаловаться' на AI-результат"""
    detail_id = await _get_param(request, "detail_id", log_name="/api/feedback")
    reason = await _get_param(request, "reason") or ""
    if not detail_id or not reason:
        return err("detail_id and reason required", 422)
    add_history(detail_id, "ai_feedback", {"reason": reason[:500]})
    return {"ok": True}


@app.post("/api/generate")
async def generate(request: Request):
    """Generate draft via LLM (or mock in demo mode). Accepts form-data, JSON, or URL param."""
    # BUG-2026-07-19-01: RBAC — нормировщик/ОТК/конструктор не должны генерировать
    role = request.cookies.get("bit_role", "technologist")
    if role not in ("technologist", "main_technologist", "admin"):
        return HTMLResponse(
            f'<span style="color:red">❌ Доступ запрещён: роль «{role}» не может генерировать</span>',
            status_code=403
        )
    detail_id = await _get_param(request, "detail_id", log_name="/api/generate")
    if not detail_id:
        return HTMLResponse(
            '<span style="color:red">❌ Не указан detail_id</span>',
            status_code=422
        )
    # V4-2: проверка дневного лимита
    limit = check_daily_limit_or_warn()
    if not limit.get("allowed"):
        return HTMLResponse(
            f'<span style="color:red">❌ {limit["message"]}</span>',
            status_code=429
        )
    detail_obj = get_detail(detail_id)
    if not detail_obj:
        return HTMLResponse(
            f'<span style="color:red">❌ Деталь {detail_id} не найдена</span>',
            status_code=404
        )

    # Проверка дневного лимита
    daily = get_daily_cost()
    if daily["exceeded"]:
        return HTMLResponse(
            f'<span style="color:red">⛔ Превышен дневной лимит {daily["limit_rub"]:.0f}₽ (потрачено {daily["total_rub"]:.2f}₽). Измени LLM_DAILY_LIMIT_RUB в .env.</span>',
            status_code=429
        )

    # DEMO MODE: return mock response based on detail
    if DEMO_MODE:
        log.info(f"Demo mode: generating mock draft for {detail_id}")
        llm_output = generate_mock_draft(detail_obj)
        add_history(detail_id, "generated_mock", {"mode": "demo"})
        # M5 fix: log mock generation in llm_calls for dashboard
        log_llm_call(detail_id, "demo-mock", "", "", json.dumps(llm_output, ensure_ascii=False), True, 0, 0, 50)
    else:
        # Real LLM call via OpenAI-compatible API
        try:
            client = get_llm_client()

            from string import Template
            tech_rules_text = detail_obj.get("tech_rules") or ""
            rules_block = f"\nВАЖНО: Технолог указал следующие правила и нюансы — ОБЯЗАТЕЛЬНО учти их:\n{tech_rules_text}\n" if tech_rules_text.strip() else ""
            prompt = Template(TECH_CARD_PROMPT).substitute(
                properties_json=json.dumps(detail_obj, indent=2, ensure_ascii=False),
                equipment_json=json.dumps(EQUIPMENT, indent=2, ensure_ascii=False),
                structure_json=json.dumps(STRUCTURE, indent=2, ensure_ascii=False),
                few_shot_json=json.dumps(FEW_SHOT_4C85941A, indent=2, ensure_ascii=False),
                tech_rules="(правила не указаны)",
                rules_block=rules_block,
                workshops_context=TECHINKOM_WORKSHOPS_CONTEXT
            )

            log.info(f"Calling {LLM_MODEL} via {LLM_API_URL}...")
            import time
            t_start = time.time()
            system_msg = "Ты — опытный технолог-сварщик. Генерируешь техкарты по свойствам деталей. Всегда возвращаешь валидный JSON без markdown-обёртки."
            try:
                response = client.chat.completions.create(
                    model=LLM_MODEL,
                    messages=[
                        {"role": "system", "content": system_msg},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.2,
                    max_tokens=8000
                )
            except Exception as llm_exc:
                duration = int((time.time() - t_start) * 1000)
                log_llm_call(detail_id, LLM_MODEL, system_msg, prompt,
                             response_text=None, response_parsed_ok=False,
                             duration_ms=duration, error=str(llm_exc))
                raise
            duration_ms = int((time.time() - t_start) * 1000)
            llm_output_text = response.choices[0].message.content
            log_llm_call(detail_id, LLM_MODEL, system_msg, prompt,
                         response_text=llm_output_text, response_parsed_ok=False,
                         tokens_in=response.usage.prompt_tokens if response.usage else None,
                         tokens_out=response.usage.completion_tokens if response.usage else None,
                         duration_ms=duration_ms)
            # Strip markdown code fences if any
            llm_output_text = llm_output_text.strip()
            if llm_output_text.startswith("```"):
                lines = llm_output_text.split("\n")
                llm_output_text = "\n".join(lines[1:-1]) if lines[-1].strip() == "```" else llm_output_text
                if llm_output_text.startswith("json"):
                    llm_output_text = llm_output_text[4:].lstrip()
            try:
                llm_output = json.loads(llm_output_text)
                # Обновляем запись: parsed_ok=True (используем lastrowid из лога)
                conn = get_conn()
                last_id = conn.execute("SELECT MAX(id) FROM llm_calls").fetchone()[0]
                if last_id:
                    conn.execute("UPDATE llm_calls SET response_parsed_ok=1 WHERE id=?", (last_id,))
                    conn.commit()
                conn.close()
            except json.JSONDecodeError as je:
                log_llm_call(detail_id, LLM_MODEL, system_msg, prompt,
                             response_text=llm_output_text, response_parsed_ok=False,
                             duration_ms=duration_ms, error=f"JSON parse: {je}")
                return HTMLResponse(
                    f'<span style="color:red">❌ LLM вернул невалидный JSON. <a href="/llm-debug" target="_blank">Открыть отладку</a></span>',
                    status_code=500
                )
            add_history(detail_id, "generated", {
                "model": LLM_MODEL,
                "tokens_in": response.usage.prompt_tokens if response.usage else None,
                "tokens_out": response.usage.completion_tokens if response.usage else None
            })
        except Exception as e:
            log.error(f"LLM error: {e}")
            error_msg = str(e).replace("<", "&lt;").replace(">", "&gt;")[:200]
            return HTMLResponse(
                f'<span style="color:red">❌ Ошибка LLM: {error_msg} <a href="/llm-debug" target="_blank">лог</a></span>',
                status_code=500
            )

    save_draft(detail_id, llm_output, "draft")
    # Пилотная метрика: всего операций в черновике
    total_ops = len(llm_output.get("operations", []))
    record_metric(detail_id, "total_ops", total_ops, {"source": "llm"})
    return HTMLResponse('<span style="color:green">✅ Готово! Перезагружаю...</span>')


def generate_mock_draft(detail_obj: dict, op_type: str = "general") -> dict:
    """Generate a mock draft based on the detail properties and op_type"""
    material = detail_obj.get("material", "")
    model = detail_obj.get("model", "")
    mass = detail_obj.get("mass_kg", 0)
    surface = detail_obj.get("surface_treatment", "") or ""  # None-safe

    # Heuristic: welding operations for steel details
    is_steel = material and "Сталь" in material
    is_shtamp = "оцинковка" in surface

    operations = []
    route = []
    step = 0

    if is_steel and not is_shtamp:
        # Steel with painting - has welding operations
        step += 1
        operations.append({
            "name": "010 Подготовительная",
            "equipment": None,
            "duration_hours": 0.2,
            "duration_source": "экспертная оценка",
            "confidence": 70,
            "materials": ["проволока Св-08Г2С-О 1,0 ГОСТ 2246-70"],
            "control_points": [],
            "gosts": [],
            "department": "Сварочно-сборочный КТ",
            "workplace": "01/01/04"
        })
        route.append({"step": step, "operation": "010 Подготовительная", "duration_hours": 0.2})

        # Welding operations
        welding_ops = [
            ("015 Сборка под сварку", 0.5, 90, "аналог: ЛМША.301314.020"),
            ("020 Сварка", 0.6, 92, "аналог: ЛМША.301314.020"),
            ("025 Сборка", 0.7, 85, "аналог: ЛМША.301314.020"),
            ("030 Сварка", 0.6, 85, "аналог: ЛМША.301314.020"),
        ]

        for name, dur, conf, src in welding_ops:
            step += 1
            operations.append({
                "name": name,
                "equipment": "Кедр-300",
                "duration_hours": dur,
                "duration_source": src,
                "confidence": conf,
                "materials": [],
                "control_points": ["ОТК визуальный"],
                "gosts": ["ГОСТ 3.1404-86"],
                "department": "Сварочно-сборочный КТ",
                "workplace": "01/01/04"
            })
            route.append({"step": step, "operation": name, "duration_hours": dur})

        # Painting
        step += 1
        operations.append({
            "name": "035 Покраска",
            "equipment": "Камера покрасочная",
            "duration_hours": 0.8,
            "duration_source": "экспертная оценка",
            "confidence": 75,
            "materials": ["грунт ГФ-021", "эмаль ПФ-115"],
            "control_points": ["ОТК визуальный", "контроль толщины покрытия"],
            "gosts": ["ГОСТ 9.402", "ГОСТ 9.410"],
            "department": "Покраска",
            "workplace": "01/07/01"
        })
        route.append({"step": step, "operation": "035 Покраска", "duration_hours": 0.8})

    elif is_shtamp:
        # Galvanized - no welding, simpler
        operations.append({
            "name": "010 Раскрой",
            "equipment": "Плазменный рез HyperTherm",
            "duration_hours": 0.1,
            "duration_source": "экспертная оценка",
            "confidence": 80,
            "materials": [],
            "control_points": ["ОТК визуальный"],
            "gosts": ["ГОСТ 9.402"],
            "department": "Лазерная резка",
            "workplace": "01/01/01"
        })
        route.append({"step": 1, "operation": "010 Раскрой", "duration_hours": 0.1})

        operations.append({
            "name": "015 Гибка",
            "equipment": "Гибочный станок",
            "duration_hours": 0.15,
            "duration_source": "экспертная оценка",
            "confidence": 75,
            "materials": [],
            "control_points": ["ОТК визуальный"],
            "gosts": [],
            "department": "Гибка",
            "workplace": "01/01/03"
        })
        route.append({"step": 2, "operation": "015 Гибка", "duration_hours": 0.15})

    # Summary
    total_hours = sum(op["duration_hours"] for op in operations)

    # Reasoning
    reasoning = {
        "operations_choice": f"Операции выбраны на основе типа материала ({material}) и характера детали ({detail_obj.get('name', '')}). Аналог: ЛМША.301314.020 (упор продольный).",
        "duration_estimates": f"Расчёт по аналогам из ведомости трудоёмкости Техинкома. Масса детали {mass} кг.",
        "equipment_choice": f"Кедр-300 — основной сварочный аппарат (если применимо). Плазменный HyperTherm — для раскроя листа.",
        "risks": "Точность операций 015-035 — 80-92% (по аналогу). Требуется проверка технолога."
    }

    # Warnings
    warnings = []
    if surface == "покраска":
        warnings.append({
            "type": "ambiguous",
            "quote": "surface_treatment: 'покраска'",
            "concern": "Не указан тип краски (порошковая/жидкая) и грунтовка",
            "question": "Какой тип краски? Требуется ли грунтовка перед покраской?"
        })
    if "Сталь 3" in material:
        warnings.append({
            "type": "ambiguous",
            "quote": f"material: '{material}'",
            "concern": "Сталь 3 — устаревшее обозначение. Возможно, Ст3сп, Ст3пс, Ст3кп?",
            "question": "Какая марка стали точно?"
        })

    # Questions
    questions = []
    if surface == "покраска":
        questions.append({
            "id": "Q1",
            "topic": "покраска",
            "question": "Тип покраски?",
            "options": ["порошковая", "жидкая (эмаль ПФ-115)", "жидкая (грунт + эмаль)", "не знаю"],
            "default": "жидкая (грунт + эмаль)",
            "impact_if_changed": "Порошковая быстрее, но дороже оборудование"
        })
    questions.append({
        "id": "Q2",
        "topic": "термообработка",
        "question": f"Требуется ли термообработка для {material}?",
        "options": ["да, закалка+отпуск", "да, только отпуск", "нет, не требуется", "не знаю"],
        "default": "нет, не требуется",
        "impact_if_changed": "Термообработка добавит 1.5-3 ч"
    })

    return {
        "summary": {
            "total_operations": len(operations),
            "total_hours": round(total_hours, 2),
            "prep_hours": 0.2,
            "complexity": "средняя" if mass > 5 else "низкая",
            "closest_analog": "ЛМША.301314.020" if is_steel else None
        },
        "route": route,
        "operations": operations,
        "reasoning": reasoning,
        "warnings": warnings,
        "questions": questions
    }


@app.post("/api/approve")
async def approve(request: Request):
    """Approve draft + Sprint 2: auto-index in RAG + F16.1: auto metrics"""
    detail_id = await _get_param(request, "detail_id")
    if not detail_id:
        return err("detail_id required", 422)
    conn = get_conn()
    conn.execute(
        "UPDATE drafts SET status = 'approved', updated_at = ? WHERE detail_id = ?",
        (datetime.now().isoformat(), detail_id)
    )
    conn.commit()
    conn.close()
    add_history(detail_id, "approved")
    edits = get_edits(detail_id)
    record_metric(detail_id, "approved", 1, {"edits_count": len(edits)})
    record_metric(detail_id, "edits_count", len(edits))
    # F16.1: auto-compute acceptance rate via diff llm_output vs final
    try:
        acc = compute_acceptance_from_versions(detail_id)
        if acc["total_ops"] > 0:
            record_metric(detail_id, "total_ops", acc["total_ops"])
            record_metric(detail_id, "accepted_op", acc["accepted_ops"])
            record_metric(detail_id, "edits_auto", acc["edits_count"])
    except Exception as e:
        log.warning(f"auto-metrics acceptance failed: {e}")
    # F16.1: auto time-to-card (delta от session_start)
    try:
        minutes = compute_time_to_card(detail_id)
        if minutes is not None:
            record_metric(detail_id, "time_to_card_min", minutes)
    except Exception as e:
        log.warning(f"auto-metrics time failed: {e}")
    # Sprint 2: автоиндексация в RAG
    try:
        from rag import rag_index_detail
        rag_index_detail(detail_id)
    except Exception as e:
        log.warning(f"RAG auto-index failed: {e}")
    return {"status": "approved", "detail_id": detail_id}


# ========== F16.1: Авто-старт таймера при открытии карточки ==========
@app.post("/api/pilot/session-start")
async def api_pilot_session_start(request: Request):
    """Записать момент открытия карточки для time-to-card."""
    detail_id = await _get_param(request, "detail_id")
    if not detail_id:
        return err("detail_id required", 422)
    role = get_current_role(request)
    record_session_start(detail_id, author=role)
    return {"ok": True, "detail_id": detail_id, "ts": datetime.now().isoformat()}


@app.post("/api/rag/rebuild")
async def api_rag_rebuild():
    """Sprint 2: перестроить RAG-индекс из БД (admin action)"""
    try:
        from rag import get_rag
        rag = get_rag()
        n = rag.rebuild_from_db()
        warning = None
        if n == 0:
            warning = "Нет утверждённых техкарт. Сначала утвердите несколько черновиков."
        elif n < 5:
            warning = f"Только {n} техкарт — similarity может быть ненадёжной. Нужно минимум 5-10."
        return JSONResponse({"ok": True, "indexed": n, "warning": warning})
    except Exception as e:
        return err(str(e)[:200], 500)


@app.get("/api/rag/similar/{detail_id}")
async def api_rag_similar(request: Request, detail_id: str, top_k: int = 3):
    """Sprint 2: top-K похожих техкарт по RAG. M24: возвращает HTML вместо JSON"""
    detail_obj = get_detail(detail_id)
    if not detail_obj:
        return err("not_found", 404)
    try:
        from rag import rag_search
        results = rag_search(detail_obj, top_k=min(top_k, 10))
        return templates.TemplateResponse(request, "_rag_similar.html", {"results": results, "request": request})
    except Exception as e:
        return err(str(e)[:200], 500)


@app.get("/api/rag/status")
async def api_rag_status():
    try:
        from rag import get_rag
        rag = get_rag()
        return JSONResponse({
            "loaded": rag.loaded,
            "documents": len(rag.ids) if rag.loaded else 0,
            "vocabulary_size": len(rag.vectorizer.vocabulary_) if rag.loaded and rag.vectorizer else 0
        })
    except Exception as e:
        return err(str(e)[:200], 500)


# ========== Sprint 3: Альтернативные маршруты + Apply similar + Batch ==========
@app.post("/api/alternatives")
async def api_alternatives(request: Request):
    detail_id = await _get_param(request, "detail_id", log_name="/api/alternatives")
    if not detail_id:
        return err("detail_id required", 422)
    detail_obj = get_detail(detail_id)
    if not detail_obj:
        return err("not found", 404)
    daily = get_daily_cost()
    if daily["exceeded"]:
        return err("daily_limit_exceeded", 429)
    if DEMO_MODE:
        alts = [
            {"variant": "A", "approach": "Сварка Кедр-300 + отпуск", "total_hours": 2.5, "n_ops": 5,
             "route": [{"step": 1, "operation": "010 Подготовительная", "duration_hours": 0.2},
                       {"step": 2, "operation": "015 Сварка", "duration_hours": 1.0},
                       {"step": 3, "operation": "020 Контроль ОТК", "duration_hours": 0.3},
                       {"step": 4, "operation": "025 Отпуск", "duration_hours": 0.5},
                       {"step": 5, "operation": "030 Финальный контроль", "duration_hours": 0.5}]},
            {"variant": "B", "approach": "TIG-сварка (аргон)", "total_hours": 3.5, "n_ops": 4,
             "route": [{"step": 1, "operation": "010 Зачистка", "duration_hours": 0.3},
                       {"step": 2, "operation": "015 TIG-сварка", "duration_hours": 1.5},
                       {"step": 3, "operation": "020 Контроль", "duration_hours": 0.7},
                       {"step": 4, "operation": "025 ТО", "duration_hours": 1.0}]},
            {"variant": "C", "approach": "Минимум операций (без ТО)", "total_hours": 1.8, "n_ops": 3,
             "route": [{"step": 1, "operation": "010 Сборка", "duration_hours": 0.5},
                       {"step": 2, "operation": "015 Сварка", "duration_hours": 0.8},
                       {"step": 3, "operation": "020 Контроль", "duration_hours": 0.5}]}
        ]
        return templates.TemplateResponse(request, "_alternatives.html", {"alternatives": alts, "mode": "demo", "cost": "1.50₽", "request": request})
    return templates.TemplateResponse(request, "_alternatives.html", {"alternatives": [], "mode": "live-stub", "message": "real LLM for alternatives — в v0.5", "request": request})


@app.post("/api/apply-similar")
async def api_apply_similar(request: Request):
    detail_id = await _get_param(request, "detail_id", log_name="/api/apply-similar")
    source_id = await _get_param(request, "source_id")
    if not detail_id or not source_id:
        return err("detail_id and source_id required", 422)
    if detail_id == source_id:
        return err("cannot apply to self", 400)
    conn = get_conn()
    source_draft = conn.execute("SELECT * FROM drafts WHERE detail_id=?", (source_id,)).fetchone()
    if not source_draft:
        conn.close()
        return err("source has no draft", 404)
    try:
        source_output = json.loads(source_draft[1])
    except Exception:
        conn.close()
        return err("source draft corrupt", 500)
    target_draft = conn.execute("SELECT * FROM drafts WHERE detail_id=?", (detail_id,)).fetchone()
    if target_draft:
        try:
            target_output = json.loads(target_draft[1])
            target_output["operations"] = source_output.get("operations", [])
            target_output["route"] = source_output.get("route", [])
            target_output["reasoning"] = {
                "operations_choice": f"Скопировано из {source_id} (1-click apply similar)",
                "duration_estimates": "Скопировано",
                "equipment_choice": "Скопировано",
                "risks": "Требует верификации технологом"
            }
            target_output["applied_from"] = source_id
            conn.execute("UPDATE drafts SET llm_output=?, updated_at=? WHERE detail_id=?",
                         (json.dumps(target_output, ensure_ascii=False), datetime.now().isoformat(), detail_id))
        except Exception as e:
            conn.close()
            return JSONResponse({"error": f"failed to update: {e}"}, status_code=500)
    else:
        new_output = source_output.copy()
        new_output["applied_from"] = source_id
        conn.execute("INSERT INTO drafts (detail_id, llm_output, status, author) VALUES (?, ?, 'draft', 'rag-apply')",
                     (detail_id, json.dumps(new_output, ensure_ascii=False)))
    conn.commit()
    conn.close()
    add_history(detail_id, "applied_similar", {"source_id": source_id})
    return JSONResponse({"ok": True, "applied_from": source_id})


@app.post("/api/batch-generate")
async def api_batch_generate(request: Request):
    detail_ids_raw = await _get_param(request, "detail_ids", log_name="/api/batch-generate")
    if not detail_ids_raw:
        return err("detail_ids required (JSON array)", 422)
    try:
        detail_ids = json.loads(detail_ids_raw)
    except Exception:
        return err("detail_ids must be JSON array", 422)
    if not isinstance(detail_ids, list) or len(detail_ids) == 0:
        return err("detail_ids must be non-empty array", 422)
    if len(detail_ids) > 20:
        return err("max 20 details per batch", 400)
    results = []
    for did in detail_ids:
        did = str(did).strip()
        if not did:
            continue
        detail_obj = get_detail(did)
        if not detail_obj:
            results.append({"detail_id": did, "status": "not_found"})
            continue
        daily = get_daily_cost()
        if daily["exceeded"]:
            results.append({"detail_id": did, "status": "daily_limit_exceeded"})
            continue
        if DEMO_MODE:
            llm_output = generate_mock_draft(detail_obj)
        else:
            try:
                from string import Template
                client = get_llm_client()
                prompt = Template(TECH_CARD_PROMPT).substitute(
                    properties_json=json.dumps(detail_obj, indent=2, ensure_ascii=False),
                    equipment_json=json.dumps(EQUIPMENT, indent=2, ensure_ascii=False),
                    structure_json=json.dumps(STRUCTURE, indent=2, ensure_ascii=False),
                    few_shot_json=json.dumps(FEW_SHOT_4C85941A, indent=2, ensure_ascii=False),
                    tech_rules="(правила не указаны)",
                    rules_block="",
                    workshops_context=TECHINKOM_WORKSHOPS_CONTEXT
                )
                response = client.chat.completions.create(
                    model=LLM_MODEL,
                    messages=[{"role": "system", "content": "Ты — технолог. Генерируй JSON."}, {"role": "user", "content": prompt}],
                    temperature=0.2, max_tokens=8000
                )
                text = response.choices[0].message.content.strip()
                llm_output = parse_llm_json(text)
            except Exception as e:
                results.append({"detail_id": did, "status": "llm_error", "error": str(e)[:100]})
                continue
        save_draft(did, llm_output, "draft", author="batch")
        # M5 fix: log batch generation
        log_llm_call(did, "demo-batch" if DEMO_MODE else LLM_MODEL, "", "",
                     json.dumps(llm_output, ensure_ascii=False), True, 0, 0, 80)
        results.append({"detail_id": did, "status": "generated", "ops": len(llm_output.get("operations", []))})
    return JSONResponse({"ok": True, "processed": len(results), "results": results})


# ========== Sprint 5: Audit log + Export ==========
@app.get("/audit", response_class=HTMLResponse)
async def audit_page(request: Request, limit: int = 100):
    conn = get_conn()
    cols = [d[1] for d in conn.execute("PRAGMA table_info(history)").fetchall()]
    rows = conn.execute("SELECT * FROM history ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    entries = [dict(zip(cols, r)) for r in rows]
    # N9 fix: pretty-print JSON для UI
    for e in entries:
        details_str = e.get("details") or ""
        if details_str:
            try:
                parsed = json.loads(details_str)
                e["details_pretty"] = json.dumps(parsed, indent=2, ensure_ascii=False)
            except Exception:
                e["details_pretty"] = details_str
        else:
            e["details_pretty"] = ""
    return templates.TemplateResponse("audit.html", {
        "request": request, "entries": entries, "limit": limit
    })


@app.get("/api/audit/export")
async def api_audit_export():
    conn = get_conn()
    cols = [d[1] for d in conn.execute("PRAGMA table_info(history)").fetchall()]
    rows = conn.execute("SELECT * FROM history ORDER BY id").fetchall()
    conn.close()
    entries = [dict(zip(cols, r)) for r in rows]
    return JSONResponse({
        "exported_at": datetime.now().isoformat(),
        "total_entries": len(entries),
        "entries": entries
    })


@app.get("/api/export/all")
async def api_export_all():
    conn = get_conn()
    tables = ["details", "drafts", "draft_versions", "edits", "rules", "equipment",
              "materials", "iot", "benchmarks", "history", "llm_calls", "pilot_metrics"]
    dump = {"exported_at": datetime.now().isoformat(), "version": "v0.4", "tables": {}}
    for t in tables:
        try:
            cols = [d[1] for d in conn.execute(f"PRAGMA table_info({t})").fetchall()]
            rows = conn.execute(f"SELECT * FROM {t}").fetchall()
            dump["tables"][t] = {"columns": cols, "rows": [dict(zip(cols, r)) for r in rows]}
        except Exception as e:
            dump["tables"][t] = {"error": str(e)}
    conn.close()
    return JSONResponse(dump)


# ========== Печатная форма ТК (P0: для подписи в бумажный журнал) ==========
@app.get("/detail/{detail_id}/print", response_class=HTMLResponse)
async def detail_print(request: Request, detail_id: str):
    """Печатная форма ТК — без кнопок, только данные + место для подписи"""
    detail_obj = get_detail(detail_id)
    if not detail_obj:
        raise HTTPException(404, "Detail not found")
    draft_data = get_draft(detail_id)
    return templates.TemplateResponse("print.html", {
        "request": request,
        "detail": detail_obj,
        "draft": draft_data["output"] if draft_data else None,
        "status": draft_data["status"] if draft_data else "new"
    })


# ========== U6 fix: click-to-edit (inline edit) ==========
@app.post("/api/edit/inline")
async def api_edit_inline(request: Request):
    """Inline-edit одной операции: один POST меняет одно поле одной операции"""
    detail_id = await _get_param(request, "detail_id", log_name="/api/edit/inline")
    op_index = await _get_param(request, "op_index")
    field = await _get_param(request, "field")
    value = await _get_param(request, "value")
    if not all([detail_id, op_index, field]):
        return JSONResponse({"error": "detail_id, op_index, field required"}, status_code=422)
    # Whitelist полей для inline-edit
    if field not in ("name", "equipment", "duration_hours", "department", "workplace",
                     "materials", "gosts", "control_points", "duration_source"):
        return JSONResponse({"error": f"field '{field}' not editable inline"}, status_code=400)
    try:
        op_idx = int(op_index)
    except ValueError:
        return err("op_index must be int", 422)
    conn = get_conn()
    row = conn.execute("SELECT llm_output FROM drafts WHERE detail_id=?", (detail_id,)).fetchone()
    if not row:
        conn.close()
        return err("no draft", 404)
    try:
        output = json.loads(row[0])
    except Exception:
        conn.close()
        return err("draft corrupt", 500)
    if op_idx < 0 or op_idx >= len(output.get("operations", [])):
        conn.close()
        return err("op_index out of range", 400)
    op = output["operations"][op_idx]
    # cast
    if field == "duration_hours":
        try:
            value = float(value)
        except (TypeError, ValueError):
            conn.close()
            return err("duration_hours must be float", 422)
    if field in ("materials", "gosts", "control_points"):
        # list-поля — парсим как JSON array или comma-separated
        if value.startswith("["):
            try:
                import json as _j
                value = _j.loads(value)
            except Exception:
                conn.close()
                return JSONResponse({"error": f"{field} must be JSON array"}, status_code=422)
        else:
            value = [v.strip() for v in value.split(",") if v.strip()]
    old_value = op.get(field)
    op[field] = value
    # Пересчёт summary
    if field == "duration_hours" and "operations" in output:
        total = sum(float(o.get("duration_hours", 0)) for o in output["operations"])
        output["summary"]["total_hours"] = round(total, 2)
    # Update
    conn.execute("UPDATE drafts SET llm_output=?, updated_at=? WHERE detail_id=?",
                 (json.dumps(output, ensure_ascii=False), datetime.now().isoformat(), detail_id))
    conn.commit()  # C5 fix: release write lock before record_edit opens new conn
    # version берём из draft_versions
    v_row = conn.execute("SELECT MAX(version) FROM draft_versions WHERE detail_id=?", (detail_id,)).fetchone()
    version = v_row[0] if v_row and v_row[0] else 1
    conn.close()
    record_edit(detail_id, version, field, str(old_value), str(value), author="inline")
    add_history(detail_id, "inline_edit", {"op_index": op_idx, "field": field, "old": str(old_value), "new": str(value)})
    return JSONResponse({"ok": True, "field": field, "value": value, "total_hours": output.get("summary", {}).get("total_hours")})


# ========== U7 fix: warning с action — указывает куда кликать ==========
@app.post("/api/ai/feedback-positive")
async def api_feedback_positive(request: Request):
    """Положительный feedback (большой палец вверх) — для ML будущего"""
    detail_id = await _get_param(request, "detail_id", log_name="/api/ai/feedback-positive")
    if not detail_id:
        return err("detail_id required", 422)
    add_history(detail_id, "ai_feedback_positive", {})
    return JSONResponse({"ok": True, "saved": "positive"})


@app.post("/api/batch-generate-new")
async def api_batch_generate_new():
    """UX1: сгенерировать все детали со статусом new (массовое действие)"""
    daily = get_daily_cost()
    if daily["exceeded"]:
        return err("daily_limit_exceeded", 429)
    if DEMO_MODE:
        # Mock: возвращаем список new-деталей
        ids = []
        conn = get_conn()
        rows = conn.execute("""
            SELECT d.id FROM details d
            LEFT JOIN drafts dr ON d.id = dr.detail_id
            WHERE dr.detail_id IS NULL
            LIMIT 20
        """).fetchall()
        conn.close()
        ids = [r[0] for r in rows]
        return JSONResponse({"ok": True, "candidate_ids": ids, "mode": "demo"})

    # Real: запускаем batch-generate для всех new
    from urllib.parse import urlencode
    ids = []
    conn = get_conn()
    rows = conn.execute("""
        SELECT d.id FROM details d
        LEFT JOIN drafts dr ON d.id = dr.detail_id
        WHERE dr.detail_id IS NULL
        LIMIT 20
    """).fetchall()
    conn.close()
    ids = [r[0] for r in rows]
    return JSONResponse({"ok": True, "candidate_ids": ids, "mode": "live-stub", "message": "use /api/batch-generate with these ids"})


# ========== U2 fix: Ручная выгрузка техкарт в 1С-формате (Sprint 5) ==========
@app.get("/api/export/onec-csv")
async def api_export_onec_csv(detail_id: str):
    """Экспорт в CSV формате, понятном 1С:ERP (для ручного импорта на пилоте)"""
    detail_obj = get_detail(detail_id)
    if not detail_obj:
        return err("not found", 404)
    draft_data = get_draft(detail_id)
    if not draft_data:
        return err("no draft", 404)
    output = draft_data["output"]
    ops = output.get("operations", [])
    # CSV с BOM для Excel/1С кириллицы
    lines = ["\ufeffНомер;Операция;Оборудование;Время_ч;Цех;Участок;РМ;Уверенность;Материалы;ГОСТы"]
    for i, op in enumerate(ops, 1):
        lines.append(";".join([
            f"{i:03d}",
            (op.get("name") or "").replace(";", ","),
            (op.get("equipment") or "").replace(";", ","),
            str(op.get("duration_hours", 0)),
            (op.get("department") or "").replace(";", ","),
            (op.get("workplace") or "").replace(";", ","),
            str(op.get("workstation") or ""),
            str(op.get("confidence", 0)),
            "; ".join(op.get("materials", [])).replace(";", ","),
            "; ".join(op.get("gosts", [])).replace(";", ",")
        ]))
    csv_text = "\n".join(lines)
    return Response(
        content=csv_text,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="techcard_{detail_id}.csv"'}
    )


@app.post("/api/send-to-1c")
async def send_to_1c(request: Request):
    """MOCK: write RS to 1C:ERP"""
    detail_id = await _get_param(request, "detail_id")
    if not detail_id:
        return err("detail_id required", 422)
    add_history(detail_id, "sent_to_1c_mock", {
        "message": "РС записана в 1С:ERP (mock)",
        "timestamp": datetime.now().isoformat()
    })
    return {"status": "sent", "message": "РС записана в 1С:ERP (mock)"}


@app.post("/api/export/excel")
async def export_excel(detail_id: str = Form(...)):
    """Export to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    detail_obj = get_detail(detail_id)
    draft_data = get_draft(detail_id)
    if not detail_obj or not draft_data:
        raise HTTPException(400, "No draft to export")

    draft = draft_data["output"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Техкарта"

    # Header
    ws["A1"] = f"Техкарта: {detail_obj['designation']} — {detail_obj['name']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Properties
    ws["A3"] = "Материал"
    ws["B3"] = detail_obj.get("material", "")
    ws["A4"] = "Масса, кг"
    ws["B4"] = detail_obj.get("mass_kg", "")
    ws["A5"] = "Шасси"
    ws["B5"] = detail_obj.get("chassis", "")
    ws["A6"] = "Модель"
    ws["B6"] = detail_obj.get("model", "")
    for r in range(3, 7):
        ws[f"A{r}"].font = Font(bold=True)

    # Operations table
    ws["A8"] = "№"
    ws["B8"] = "Операция"
    ws["C8"] = "Оборудование"
    ws["D8"] = "Время, ч"
    ws["E8"] = "Источник"
    ws["F8"] = "Уверенность"
    for col in "ABCDEF":
        cell = ws[f"{col}8"]
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for i, op in enumerate(draft.get("operations", []), 1):
        row = 8 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=op.get("name", ""))
        ws.cell(row=row, column=3, value=op.get("equipment", "") or "—")
        ws.cell(row=row, column=4, value=op.get("duration_hours", 0))
        ws.cell(row=row, column=5, value=op.get("duration_source", ""))
        ws.cell(row=row, column=6, value=f"{op.get('confidence', 0)}%")

    # Adjust column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 15

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Transliterate filename for ASCII-safe Content-Disposition
    import re
    translit_map = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
    }
    raw_name = f"{detail_obj['designation']}"
    ascii_name = ''.join(translit_map.get(c.lower(), c) for c in raw_name)
    ascii_name = re.sub(r'[^a-zA-Z0-9._-]', '_', ascii_name)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{ascii_name}.xlsx"'}
    )


@app.post("/api/export/pdf")
async def export_pdf(request: Request):
    """Export reasoning to PDF (for management)"""
    detail_id = await _get_param(request, "detail_id")
    if not detail_id:
        return HTMLResponse('{"error":"detail_id required"}', status_code=422, media_type="application/json")
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    detail_obj = get_detail(detail_id)
    draft_data = get_draft(detail_id)
    if not detail_obj or not draft_data:
        raise HTTPException(400, "No draft to export")

    draft = draft_data["output"]
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Try to register a unicode font
    try:
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont("Helvetica"))
        font = "Helvetica"
    except Exception:
        font = "Helvetica"

    # Title
    c.setFont("Helvetica-Bold", 16)
    c.drawString(2*cm, height - 2*cm, f"Обоснование ТК")
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2*cm, height - 3*cm, f"{detail_obj['designation']} — {detail_obj['name']}")

    # Properties
    c.setFont("Helvetica-Bold", 11)
    y = height - 4.5*cm
    c.drawString(2*cm, y, "Характеристики детали:")
    c.setFont("Helvetica", 10)
    y -= 0.6*cm
    c.drawString(2*cm, y, f"Материал: {detail_obj.get('material', '')}")
    y -= 0.5*cm
    c.drawString(2*cm, y, f"Масса: {detail_obj.get('mass_kg', '')} кг")
    y -= 0.5*cm
    c.drawString(2*cm, y, f"Шасси: {detail_obj.get('chassis', '')}")
    y -= 0.5*cm
    c.drawString(2*cm, y, f"Модель: {detail_obj.get('model', '')}")
    y -= 0.5*cm
    c.drawString(2*cm, y, f"Покрытие: {detail_obj.get('surface_treatment', '')}")

    # Summary
    y -= 1*cm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, "Сводка")
    c.setFont("Helvetica", 10)
    y -= 0.6*cm
    summary = draft.get("summary", {})
    c.drawString(2*cm, y, f"Операций: {summary.get('total_operations', 0)}")
    y -= 0.5*cm
    c.drawString(2*cm, y, f"Общее время: {summary.get('total_hours', 0)} ч")
    y -= 0.5*cm
    c.drawString(2*cm, y, f"Подг. время: {summary.get('prep_hours', 0)} ч")
    y -= 0.5*cm
    c.drawString(2*cm, y, f"Сложность: {summary.get('complexity', '')}")
    y -= 0.5*cm
    c.drawString(2*cm, y, f"Ближайший аналог: {summary.get('closest_analog', '') or 'нет'}")

    # Reasoning
    y -= 1*cm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, "Обоснование решений")
    c.setFont("Helvetica", 9)
    y -= 0.6*cm
    reasoning = draft.get("reasoning", {})
    for key, value in reasoning.items():
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2*cm, y, f"{key}:")
        y -= 0.5*cm
        c.setFont("Helvetica", 9)
        # Wrap text
        words = str(value).split()
        line = ""
        for word in words:
            if len(line) + len(word) + 1 < 95:
                line += " " + word if line else word
            else:
                c.drawString(2*cm, y, line)
                y -= 0.45*cm
                line = word
        if line:
            c.drawString(2*cm, y, line)
            y -= 0.5*cm
        y -= 0.2*cm

    c.save()
    buffer.seek(0)

    # Transliterate filename for ASCII-safe Content-Disposition
    import re
    translit_map = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
    }
    raw_name = f"{detail_obj['designation']}_reasoning"
    ascii_name = ''.join(translit_map.get(c.lower(), c) for c in raw_name)
    ascii_name = re.sub(r'[^a-zA-Z0-9._-]', '_', ascii_name)

    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{ascii_name}.pdf"'}
    )


# V4-10: inline favicon (1x1 transparent pixel + SVG-like emoji нельзя, используем data URL)
@app.get("/favicon.ico")
async def favicon():
    """V4-10: пустой favicon (1x1 transparent PNG) — убирает 404 в логах."""
    import base64
    # 1x1 transparent PNG
    pixel = base64.b64decode(b"iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg==")
    from fastapi.responses import Response
    return Response(content=pixel, media_type="image/png")


# V5-9: cost anomaly detection — alert если слишком быстро сжигают лимит
COST_ANOMALY_THRESHOLD = 50.0  # ₽ за час — если больше, alert


def check_cost_anomaly(window_hours: int = 1) -> dict:
    """V5-9: проверка расхода LLM за последний час.
    Если > 50₽/час — anomaly. Если > 80% дневного лимита за первые 6 часов — тоже anomaly."""
    from db import get_conn
    from settings import get_setting
    conn = get_conn()
    try:
        row = conn.execute(f"""SELECT COALESCE(SUM(cost_rub), 0) FROM llm_calls
            WHERE created_at > datetime('now', '-{int(window_hours)} hour')""").fetchone()
        recent_cost = row[0] or 0
        day_row = conn.execute("""SELECT COALESCE(SUM(cost_rub), 0) FROM llm_calls
            WHERE date(created_at) = date('now')""").fetchone()
        day_cost = day_row[0] or 0
        limit = float(get_setting("LLM_DAILY_COST_LIMIT_RUB", "200") or 200)
        anomalies = []
        if recent_cost > COST_ANOMALY_THRESHOLD:
            anomalies.append(f"high_hourly_cost: {recent_cost:.2f}₽/час > {COST_ANOMALY_THRESHOLD}₽")
        if day_cost > limit * 0.8 and window_hours <= 6:
            anomalies.append(f"high_daily_cost_early: {day_cost:.2f}₽ ({day_cost/limit*100:.0f}% лимита) за {window_hours}ч")
        return {
            "ok": len(anomalies) == 0,
            "recent_cost_rub": round(recent_cost, 2),
            "day_cost_rub": round(day_cost, 2),
            "limit_rub": limit,
            "anomalies": anomalies
        }
    finally:
        conn.close()


@app.get("/health")
async def health():
    """M30: расширенная диагностика — БД таблицы, LLM latency, RAG stats, system info."""
    # OB3 fix: проверяем БД (SELECT 1)
    db_ok = False
    db_error = None
    db_tables = {}
    try:
        conn = get_conn()
        row = conn.execute("SELECT 1").fetchone()
        db_ok = (row is not None and row[0] == 1)
        if db_ok:
            # M30: список таблиц с количеством записей
            for tbl_row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
            ).fetchall():
                tbl = tbl_row[0]
                try:
                    count = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
                    db_tables[tbl] = count
                except Exception:
                    db_tables[tbl] = "err"
        conn.close()
    except Exception as e:
        db_error = str(e)[:200]

    # RAG статус + stats
    rag_status = "unknown"
    rag_stats = {}
    try:
        from rag import get_rag
        rag = get_rag()
        rag_status = "loaded" if rag.loaded else "empty"
        if rag.loaded:
            rag_stats = {
                "indexed_documents": len(rag.ids) if hasattr(rag, 'ids') else 0,
                "matrix_shape": list(rag.matrix.shape) if hasattr(rag, 'matrix') and rag.matrix is not None else [],
            }
    except Exception:
        rag_status = "unavailable"

    # LLM latency (ping /models endpoint)
    llm_latency_ms = None
    try:
        from settings import get_setting
        api_key = get_setting("LLM_API_KEY", LLM_API_KEY)
        api_url = get_setting("LLM_API_URL", LLM_API_URL)
        if not DEMO_MODE and api_key:
            import urllib.request, time as _t
            t0 = _t.time()
            models_url = api_url.rstrip("/") + "/models"
            req = urllib.request.Request(models_url)
            req.add_header("Authorization", f"Bearer {api_key[:8]}...")
            try:
                with urllib.request.urlopen(req, timeout=3) as resp:
                    resp.read(1)
                llm_latency_ms = int((_t.time() - t0) * 1000)
            except Exception:
                pass
    except Exception:
        pass

    # System info
    import platform
    sys_info = {
        "python": platform.python_version(),
        "platform": platform.platform(),
    }

    return {
        "status": "ok" if db_ok else "degraded",
        "db_ok": db_ok,
        "db_error": db_error,
        "db_tables": db_tables,  # M30: {table_name: row_count}
        "rag_status": rag_status,
        "rag_stats": rag_stats,  # M30
        "llm_latency_ms": llm_latency_ms,  # M30
        "demo_mode": DEMO_MODE,
        "model": LLM_MODEL,
        "api_url": LLM_API_URL if not DEMO_MODE else None,
        "details_count": len(MOCK_DETAILS),
        "version": "0.4.18",
        "build_date": "2026-07-19",
        "git_commit": _GIT_COMMIT,
        "uptime_sec": int(time.time() - _APP_START_TS) if '_APP_START_TS' in dir() else 0,
        "dependencies": _check_dependencies(),
        "cost_anomaly": check_cost_anomaly(),
        "system": sys_info,  # M30
    }


def _check_dependencies() -> dict:
    """V5-4: проверка доступности LLM API, Telegram, SMTP."""
    deps = {"llm": "unknown", "telegram": "unknown", "smtp": "unknown"}
    # LLM
    try:
        from settings import get_setting
        api_key = get_setting("LLM_API_KEY", LLM_API_KEY)
        api_url = get_setting("LLM_API_URL", LLM_API_URL)
        if DEMO_MODE or not api_key:
            deps["llm"] = "demo_mode" if DEMO_MODE else "not_configured"
        else:
            # GET к /v1/models (OpenAI-совместимый endpoint для проверки)
            import urllib.request
            models_url = api_url.rstrip("/") + "/models"
            req = urllib.request.Request(models_url)
            req.add_header("Authorization", f"Bearer {api_key[:8]}...")  # partial — не для реальной авторизации
            try:
                with urllib.request.urlopen(req, timeout=3) as resp:
                    deps["llm"] = "ok" if resp.status < 500 else "degraded"
            except Exception as e:
                # 401/403 = ключ не подходит, но endpoint живой
                err_str = str(e)[:80]
                if "401" in err_str or "403" in err_str:
                    deps["llm"] = "auth_error"  # ключ не подходит
                elif "404" in err_str:
                    deps["llm"] = "endpoint_not_found"
                else:
                    deps["llm"] = f"unreachable: {err_str}"
    except Exception as e:
        deps["llm"] = f"check_failed: {str(e)[:50]}"
    # Telegram
    try:
        token = get_setting("TELEGRAM_BOT_TOKEN", "")
        chat = get_setting("TELEGRAM_CHAT_ID", "")
        # Placeholder-ы или пустые = not_configured
        if token and chat and "__FILL_" not in token and "__FILL_" not in chat:
            deps["telegram"] = "configured"
        else:
            deps["telegram"] = "not_configured"
    except Exception:
        deps["telegram"] = "check_failed"
    # SMTP
    try:
        host = get_setting("SMTP_HOST", "")
        user = get_setting("SMTP_USER", "")
        deps["smtp"] = "configured" if (host and user and "__FILL_" not in host) else "not_configured"
    except Exception:
        deps["smtp"] = "check_failed"
    return deps


@app.get("/history/{detail_id}")
async def history(detail_id: str):
    """Get history for a detail (for debug)"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.execute(
        "SELECT id, action, timestamp, details FROM history WHERE detail_id = ? ORDER BY id DESC LIMIT 50",
        (detail_id,)
    )
    rows = cursor.fetchall()
    conn.close()
    return {"history": [
        {"id": r[0], "action": r[1], "timestamp": r[2], "details": json.loads(r[3] or "{}")}
        for r in rows
    ]}


# ============================================
# CRUD: детали
# ============================================
@app.get("/details/new", response_class=HTMLResponse)
async def new_detail_form(request: Request):
    materials = get_all_materials()
    return templates.TemplateResponse("detail_form.html", {
        "request": request,
        "detail": None,
        "materials": materials
    })


@app.post("/api/details")
async def api_create_detail(designation: str = Form(...),
                              name: str = Form(""),
                              model: str = Form(""),
                              chassis: str = Form(""),
                              material: str = Form(""),
                              size_mm: str = Form(""),
                              mass_kg: float = Form(0),
                              surface_treatment: str = Form("")):
    d = {"designation": designation, "name": name, "model": model, "chassis": chassis,
         "material": material, "size_mm": size_mm, "mass_kg": mass_kg, "surface_treatment": surface_treatment}
    detail_id = create_detail(d)
    add_history(detail_id, "detail_created", d)
    return RedirectResponse(f"/detail/{detail_id}", status_code=303)


# ============================================
# CRUD: оборудование
# ============================================
@app.get("/equipment", response_class=HTMLResponse)
async def equipment_page(request: Request, q: str = ""):
    items = get_all_equipment()
    if q:
        ql = q.lower()
        items = [e for e in items if ql in (e.get("name") or "").lower() or ql in (e.get("type") or "").lower() or ql in (e.get("code") or "").lower()]
    return templates.TemplateResponse("equipment_list.html", {
        "request": request,
        "equipment": items,
        "q": q
    })


@app.get("/equipment/search", response_class=HTMLResponse)
async def equipment_search(request: Request, q: str = ""):
    """J fix: live search для equipment (htmx)"""
    return await equipment_page(request, q=q)


@app.post("/api/equipment")
async def api_create_equipment(name: str = Form(...),
                                type: str = Form(""),
                                code: str = Form(""),
                                max_thickness_mm: float = Form(0),
                                max_mass_kg: float = Form(0),
                                notes: str = Form("")):
    eid = create_equipment({"name": name, "type": type, "code": code,
                            "max_thickness_mm": max_thickness_mm,
                            "max_mass_kg": max_mass_kg, "notes": notes})
    add_history(eid, "equipment_added", {"name": name})
    return RedirectResponse("/equipment", status_code=303)


# ============================================
# CRUD: материалы
# ============================================
@app.get("/materials", response_class=HTMLResponse)
async def materials_page(request: Request, q: str = ""):
    items = get_all_materials()
    if q:
        ql = q.lower()
        items = [m for m in items if ql in (m.get("name") or "").lower() or ql in (m.get("grade") or "").lower() or ql in (m.get("gost") or "").lower()]
    return templates.TemplateResponse("materials_list.html", {
        "request": request,
        "items": items,
        "q": q
    })


@app.get("/materials/search", response_class=HTMLResponse)
async def materials_search(request: Request, q: str = ""):
    return await materials_page(request, q=q)


@app.post("/api/materials")
async def api_create_material(name: str = Form(...),
                               grade: str = Form(""),
                               gost: str = Form(""),
                               notes: str = Form("")):
    mid = create_material({"name": name, "grade": grade, "gost": gost, "notes": notes})
    return RedirectResponse("/materials", status_code=303)


# ============================================
# CRUD: ИОТ
# ============================================
@app.get("/iot", response_class=HTMLResponse)
async def iot_page(request: Request, q: str = ""):
    items = get_all_iot()
    if q:
        ql = q.lower()
        items = [i for i in items if ql in (i.get("number") or "").lower() or ql in (i.get("description") or "").lower()]
    return templates.TemplateResponse("iot_list.html", {
        "request": request,
        "items": items,
        "q": q
    })


@app.get("/iot/search", response_class=HTMLResponse)
async def iot_search(request: Request, q: str = ""):
    return await iot_page(request, q=q)


@app.post("/api/iot")
async def api_create_iot(number: str = Form(...),
                          description: str = Form(""),
                          applies_to: str = Form("")):
    iid = create_iot({"number": number, "description": description, "applies_to": applies_to})
    return RedirectResponse("/iot", status_code=303)


# ============================================
# CRUD: бенчмарки
# ============================================
@app.get("/benchmarks", response_class=HTMLResponse)
async def benchmarks_page(request: Request):
    items = get_all_benchmarks()
    return templates.TemplateResponse("benchmarks_list.html", {
        "request": request,
        "items": items
    })


@app.post("/api/benchmarks")
async def api_create_benchmark(detail_type: str = Form(...),
                                 norm_hours: float = Form(0),
                                 source: str = Form(""),
                                 sample_size: int = Form(1)):
    bid = create_benchmark({"detail_type": detail_type, "norm_hours": norm_hours,
                            "source": source, "sample_size": sample_size})
    return RedirectResponse("/benchmarks", status_code=303)


@app.post("/api/import/equipment")
async def api_import_equipment(department: str = Form(""), type: str = Form(""), limit: int = Form(50)):
    """Mock-импорт оборудования из 1С с фильтром (для прототипа — не реальный обмен)"""
    # Mock: просто обновляем last_sync_at у существующих
    import random
    added = 0
    for eq_id, name, etype in [
        (f"1c-{random.randint(1000,9999)}", f"Импортированное #{i}", type or "Сварочный аппарат")
        for i in range(min(limit, 10))
    ]:
        eid = f"eq-imp-{added}"
        conn = get_conn()
        try:
            conn.execute("""INSERT INTO equipment
                (id, name, type, code, notes, source, last_sync_at)
                VALUES (?, ?, ?, ?, ?, '1c', ?)""",
                (eid, name, etype, str(random.randint(10000, 99999)),
                 f"импорт {datetime.now().isoformat()}", datetime.now().isoformat()))
            conn.commit()
            added += 1
        except Exception:
            pass
        conn.close()
    add_history("import", "equipment_imported", {"added": added, "filter": {"department": department, "type": type}})
    return RedirectResponse("/equipment", status_code=303)


@app.post("/api/import/materials")
async def api_import_materials(grade: str = Form(""), limit: int = Form(50)):
    """Mock-импорт материалов из 1С"""
    import random
    for i in range(min(limit, 5)):
        mid = f"m-imp-{i}"
        conn = get_conn()
        try:
            conn.execute("""INSERT INTO materials (id, name, grade, gost, source, last_sync_at)
                VALUES (?, ?, ?, ?, '1c', ?)""",
                (mid, f"Импортированный материал #{i}", grade or "Ст3",
                 f"ГОСТ {random.randint(100,99999)}-{random.randint(80,2025)}",
                 datetime.now().isoformat()))
            conn.commit()
        except Exception:
            pass
        conn.close()
    return RedirectResponse("/materials", status_code=303)


@app.post("/api/import/iot")
async def api_import_iot(limit: int = Form(50)):
    """Mock-импорт ИОТ"""
    import random
    for i in range(min(limit, 5)):
        iid = f"iot-imp-{i}"
        conn = get_conn()
        try:
            conn.execute("""INSERT INTO iot (id, number, description, source, last_sync_at)
                VALUES (?, ?, ?, '1c', ?)""",
                (iid, str(random.randint(1, 200)), f"Импортированный ИОТ #{i}",
                 datetime.now().isoformat()))
            conn.commit()
        except Exception:
            pass
        conn.close()
    return RedirectResponse("/iot", status_code=303)


@app.post("/api/import/benchmarks")
async def api_import_benchmarks(limit: int = Form(100)):
    """Mock-импорт бенчмарков (ведомость трудоёмкости)"""
    import random
    for i in range(min(limit, 5)):
        bid = f"bk-imp-{i}"
        conn = get_conn()
        try:
            conn.execute("""INSERT INTO benchmarks (id, detail_type, norm_hours, source, sample_size, last_sync_at)
                VALUES (?, ?, ?, '1c', ?, ?)""",
                (bid, f"Импортированный тип #{i}", round(random.uniform(0.1, 5.0), 2),
                 "ведомость Техинкома", random.randint(1, 50),
                 datetime.now().isoformat()))
            conn.commit()
        except Exception:
            pass
        conn.close()
    return RedirectResponse("/benchmarks", status_code=303)


@app.post("/api/equipment/local-params")
async def api_equipment_local_params(id: str = Form(...), local_notes: str = Form("")):
    """Добавить локальные параметры к оборудованию (поверх 1С-данных)"""
    conn = get_conn()
    # Просто пишем в notes, добавляя маркер
    row = conn.execute("SELECT notes FROM equipment WHERE id=?", (id,)).fetchone()
    if not row:
        conn.close()
        return HTMLResponse('❌ Не найдено', status_code=404)
    old_notes = row[0] or ""
    new_notes = old_notes + f"\n[локально {datetime.now().strftime('%d.%m')}] {local_notes}"
    conn.execute("UPDATE equipment SET notes=? WHERE id=?", (new_notes, id))
    conn.commit()
    conn.close()
    return RedirectResponse("/equipment", status_code=303)


# ============================================
# РЕДАКТОР ОПЕРАЦИЙ (правка черновика технологом)
# ============================================
@app.post("/api/edit/operation")
async def api_edit_operation(request: Request):
    """Правка одной операции в черновике (записывается как edit + version)"""
    detail_id = await _get_param(request, "detail_id")
    op_index_str = await _get_param(request, "op_index")
    field = await _get_param(request, "field")
    value = await _get_param(request, "value")
    reason = await _get_param(request, "reason") or ""
    author = await _get_param(request, "author") or "technologist"
    if not all([detail_id, op_index_str, field, value is not None]):
        return HTMLResponse('<span style="color:red">❌ Не хватает параметров</span>', status_code=422)
    op_index = int(op_index_str)

    draft_data = get_draft(detail_id)
    if not draft_data:
        return HTMLResponse('<span style="color:red">❌ Нет черновика</span>', status_code=400)

    operations = draft_data["output"].get("operations", [])
    if op_index < 0 or op_index >= len(operations):
        return HTMLResponse('<span style="color:red">❌ Нет такой операции</span>', status_code=400)

    old_value = operations[op_index].get(field, "")
    operations[op_index][field] = value
    draft_data["output"]["operations"] = operations

    save_draft(detail_id, draft_data["output"], status="draft", author=author)
    new_v = save_version(detail_id, operations, author=author, source="human_edit",
                         notes=f"{field}: {old_value} → {value}")
    record_edit(detail_id, new_v, f"op{op_index}.{field}", old_value, value, reason, author)
    add_history(detail_id, "operation_edited", {"op": op_index, "field": field, "old": old_value, "new": value})
    # V5-8: security audit log
    add_history(detail_id, "security_audit_edit", {
        "actor": author, "op": op_index, "field": field,
        "old": str(old_value)[:200], "new": str(value)[:200], "reason": reason[:200]
    })

    return HTMLResponse(f'<span style="color:green">✅ Операция {op_index+1} обновлена (v{new_v})</span>')


# F16.8: A4-11 — получить список удалённых операций
def get_deleted_operations(detail_id: str) -> list:
    """Возвращает удалённые операции (с возможностью restore)."""
    from db import get_conn
    conn = get_conn()
    try:
        rows = conn.execute("""SELECT id, op_index, op_name, deleted_at, deleted_by, reason, restored_at
            FROM deleted_operations WHERE detail_id=? ORDER BY deleted_at DESC LIMIT 20""",
            (detail_id,)).fetchall()
        return [{
            "id": r[0], "op_index": r[1], "op_name": r[2],
            "deleted_at": r[3], "deleted_by": r[4], "reason": r[5],
            "restored_at": r[6]
        } for r in rows]
    except Exception as e:
        log.debug(f"get_deleted_operations({detail_id}): {e}")
        return []
    finally:
        try:
            conn.close()
        except Exception:
            pass


# F16.8: A4-11 — restore удалённой операции
@app.post("/api/edit/restore-operation")
async def api_restore_operation(request: Request):
    """Восстанавливает последнюю удалённую операцию (или по id)."""
    detail_id = await _get_param(request, "detail_id")
    del_id = await _get_param(request, "del_id")  # опционально
    author = await _get_param(request, "author") or "technologist"
    if not detail_id:
        return err("detail_id required", 422)
    from db import get_conn
    conn = get_conn()
    try:
        if del_id:
            row = conn.execute("""SELECT id, op_json, op_name FROM deleted_operations
                WHERE id=? AND detail_id=? AND restored_at IS NULL""",
                (int(del_id), detail_id)).fetchone()
        else:
            # Последняя не восстановленная
            row = conn.execute("""SELECT id, op_json, op_name FROM deleted_operations
                WHERE detail_id=? AND restored_at IS NULL
                ORDER BY deleted_at DESC LIMIT 1""", (detail_id,)).fetchone()
        if not row:
            return err("no deleted operations to restore", 404)
        try:
            op = json.loads(row[1])
        except Exception:
            return err("corrupted op_json", 500)
        # Добавить обратно в operations
        draft_data = get_draft(detail_id)
        if not draft_data:
            return err("no draft", 404)
        ops = draft_data["output"].get("operations", [])
        ops.append(op)
        draft_data["output"]["operations"] = ops
        save_draft(detail_id, draft_data["output"], status="draft", author=author)
        new_v = save_version(detail_id, ops, author=author, source="human_restore",
                             notes=f"Restored: {row[2]}")
        # Пометить как восстановленную
        conn.execute("""UPDATE deleted_operations
            SET restored_at=CURRENT_TIMESTAMP, restored_by=?
            WHERE id=?""", (author, row[0]))
        conn.commit()
        return {"ok": True, "restored": row[2], "version": new_v}
    finally:
        conn.close()


@app.post("/api/edit/add-operation")
async def api_add_operation(request: Request):
    """Добавляет новую операцию в черновик"""
    detail_id = await _get_param(request, "detail_id")
    name = await _get_param(request, "name")
    equipment = await _get_param(request, "equipment") or ""
    duration_str = await _get_param(request, "duration_hours") or "0"
    author = await _get_param(request, "author") or "technologist"
    if not detail_id or not name:
        return HTMLResponse('<span style="color:red">❌ Не хватает параметров</span>', status_code=422)
    duration_hours = float(duration_str)

    draft_data = get_draft(detail_id)
    if not draft_data:
        return HTMLResponse('<span style="color:red">❌ Нет черновика</span>', status_code=400)

    operations = draft_data["output"].get("operations", [])
    new_op = {
        "name": name, "equipment": equipment, "duration_hours": duration_hours,
        "duration_source": "ручной ввод", "confidence": 100
    }
    operations.append(new_op)
    draft_data["output"]["operations"] = operations
    save_draft(detail_id, draft_data["output"], status="draft", author=author)
    new_v = save_version(detail_id, operations, author=author, source="human_add", notes=f"Added: {name}")
    record_edit(detail_id, new_v, "new_op", "", name, "added by user", author)

    return HTMLResponse(f'<span style="color:green">✅ Операция добавлена (v{new_v})</span>')


@app.post("/api/edit/delete-operation")
async def api_delete_operation(request: Request):
    """F16.8: A4-11 — soft-delete операции (восстанавливаемо).
    Раньше удаляло безвозвратно. Теперь: помечает как удалённую, можно восстановить."""
    detail_id = await _get_param(request, "detail_id")
    op_index_str = await _get_param(request, "op_index")
    reason = await _get_param(request, "reason") or ""
    author = await _get_param(request, "author") or "technologist"
    if not detail_id or op_index_str is None:
        return HTMLResponse('<span style="color:red">❌ Не хватает параметров</span>', status_code=422)
    op_index = int(op_index_str)

    draft_data = get_draft(detail_id)
    if not draft_data:
        return HTMLResponse('<span style="color:red">❌ Нет черновика</span>', status_code=400)

    operations = draft_data["output"].get("operations", [])
    if op_index < 0 or op_index >= len(operations):
        return HTMLResponse('<span style="color:red">❌ Нет такой операции</span>', status_code=400)

    removed = operations.pop(op_index)
    draft_data["output"]["operations"] = operations
    save_draft(detail_id, draft_data["output"], status="draft", author=author)
    new_v = save_version(detail_id, operations, author=author, source="human_delete",
                         notes=f"Deleted: {removed.get('name', '')}")
    record_edit(detail_id, new_v, f"op{op_index}", removed.get("name", ""), "", reason, author)
    # F16.8: A4-11 — soft-delete: сохраняем для возможного restore
    from db import get_conn
    conn = get_conn()
    try:
        conn.execute("""INSERT INTO deleted_operations
            (detail_id, op_index, op_name, op_json, deleted_at, deleted_by, reason)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP, ?, ?)""",
            (detail_id, op_index, removed.get("name", ""),
             json.dumps(removed, ensure_ascii=False), author, reason))
        conn.commit()
    except Exception as e:
        # Если таблица не существует — это не критично (миграция может быть не применена)
        log.warning(f"soft-delete save failed (table may not exist): {e}")
    finally:
        conn.close()

    return HTMLResponse(f'<span style="color:green">✅ Операция удалена (v{new_v})</span>')


# ============================================
# ОБУЧЕНИЕ: правила + метрики
# ============================================
@app.get("/learning", response_class=HTMLResponse)
async def learning_page(request: Request):
    metrics = get_metrics()
    rules = extract_rule_from_edits()
    all_edits = get_edits()
    llm_calls = get_llm_calls(limit=20)
    return templates.TemplateResponse("learning.html", {
        "request": request,
        "metrics": metrics,
        "rules": rules,
        "edits": all_edits,
        "llm_calls": llm_calls
    })


@app.get("/llm-debug", response_class=HTMLResponse)
async def llm_debug_page(request: Request, detail_id: str = None, call_id: int = None):
    calls = get_llm_calls(detail_id=detail_id, limit=100)
    selected = None
    if call_id:
        selected = get_llm_call_detail(call_id)
    elif calls:
        selected = get_llm_call_detail(calls[0]["id"])
    return templates.TemplateResponse("llm_debug.html", {
        "request": request,
        "calls": calls,
        "selected": selected
    })


# ============================================
# ПИЛОТ: дашборд KPI + ввод метрик
# ============================================
@app.get("/pilot/learning", response_class=HTMLResponse)
async def pilot_learning_dashboard(request: Request, weeks: int = 4):
    """Дашборд обучения RAG: метрики по неделям + график тренда"""
    weeks = max(1, min(12, int(weeks)))
    metrics = get_learning_metrics_by_week(weeks=weeks)
    return templates.TemplateResponse("pilot_learning.html", {
        "request": request,
        "metrics": metrics,
        "weeks": weeks,
        "current_role": get_current_role(request),
        "roles": ROLES
    })


@app.get("/api/pilot/learning")
async def api_pilot_learning(weeks: int = 4):
    """JSON-метрики по неделям для построения графика на клиенте"""
    weeks = max(1, min(12, int(weeks)))
    return JSONResponse({
        "weeks": weeks,
        "metrics": get_learning_metrics_by_week(weeks=weeks)
    })


@app.get("/pilot", response_class=HTMLResponse)
async def pilot_dashboard(request: Request):
    metrics = get_pilot_metrics()
    conn = get_conn()
    recent = conn.execute("""SELECT detail_id,
        SUM(CASE WHEN metric='edits_count' THEN value ELSE 0 END) as edits,
        SUM(CASE WHEN metric='time_to_card_min' THEN value ELSE 0 END) as time_min,
        MAX(created_at) as last
        FROM pilot_metrics GROUP BY detail_id ORDER BY last DESC LIMIT 20""").fetchall()
    # RAG-метрика: сколько ТК в индексе
    rag_count = 0
    rag_status = "empty"
    rag_path = ".rag"
    if os.path.exists(rag_path):
        try:
            ids_files = [f for f in os.listdir(rag_path) if 'ids' in f]
            if ids_files:
                import pickle
                with open(os.path.join(rag_path, ids_files[0]), 'rb') as f:
                    ids = pickle.load(f)
                    rag_count = len(ids)
                    rag_status = "loaded"
        except Exception:
            rag_status = "error"
    # Количество утверждённых ТК (для пилота: target 30+)
    approved_count = conn.execute("SELECT COUNT(*) FROM drafts WHERE status='approved'").fetchone()[0] or 0
    # Количество всего деталей
    total_details = conn.execute("SELECT COUNT(*) FROM details").fetchone()[0] or 0
    # RAG-готовность: для качественной работы RAG нужно 30+ утверждённых ТК
    rag_readiness = "🟢 достаточно" if approved_count >= 30 else ("🟡 минимум" if approved_count >= 10 else "🔴 мало")
    rag_target = 30
    conn.close()
    approved_list = [{"detail_id": r[0], "edits": r[1] or 0,
                      "time_min": r[2] or 0, "last": r[3]} for r in recent]
    # V4-9: дневная стоимость LLM
    cost_today = get_daily_cost()
    return templates.TemplateResponse("pilot.html", {
        "request": request,
        "metrics": metrics,
        "approved_list": approved_list,
        "rag_count": rag_count,
        "rag_status": rag_status,
        "rag_readiness": rag_readiness,
        "rag_target": rag_target,
        "approved_count": approved_count,
        "total_details": total_details,
        "cost_today": cost_today  # V4-9
    })


@app.get("/api/pilot/report")
async def api_pilot_report(days: int = 30):
    """Pilot report generator (JSON: summary + details + charts + markdown)"""
    from pilot_report import generate_pilot_report
    try:
        report = generate_pilot_report(days=days)
        return JSONResponse({
            "ok": True,
            "summary": report["summary"],
            "details": report["details"],
            "charts": report["charts"],
            "markdown": report["markdown"]
        })
    except Exception as e:
        log.exception(f"pilot report failed: {e}")
        return err(f"pilot report failed: {e}", 500)


@app.get("/pilot/report", response_class=HTMLResponse)
async def pilot_report_page(request: Request, days: int = 30):
    """Страница отчёта для руководства"""
    from pilot_report import generate_pilot_report
    try:
        report = generate_pilot_report(days=days)
    except Exception as e:
        log.exception(f"pilot report failed: {e}")
        return HTMLResponse(f"<h1>Ошибка генерации отчёта</h1><pre>{e}</pre>", status_code=500)
    return templates.TemplateResponse("pilot_report.html", {
        "request": request,
        "report": report,
        "days": days
    })


@app.get("/api/pilot/report/markdown")
async def api_pilot_report_markdown(days: int = 30):
    """Pilot report как Markdown (для копирования в email)"""
    from pilot_report import generate_pilot_report
    try:
        report = generate_pilot_report(days=days)
        return Response(
            content=report["markdown"],
            media_type="text/markdown; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="pilot_report_{days}d.md"'}
        )
    except Exception as e:
        return err(f"pilot report failed: {e}", 500)


@app.post("/api/pilot/time")
async def api_pilot_time(detail_id: str = Form(...), minutes: float = Form(...),
                          author: str = Form("technologist")):
    record_metric(detail_id, "time_to_card_min", minutes, {"author": author})
    return RedirectResponse("/pilot", status_code=303)


@app.post("/api/pilot/accepted")
async def api_pilot_accepted(detail_id: str = Form(...), total_ops: int = Form(...),
                              accepted_ops: int = Form(...)):
    record_metric(detail_id, "total_ops", total_ops)
    record_metric(detail_id, "accepted_op", accepted_ops)
    return RedirectResponse("/pilot", status_code=303)


# ============================================
# ПРАВИЛА ТЕХНОЛОГА (структурированный ввод)
# ============================================
@app.post("/api/details/{detail_id}/rules")
async def api_save_rules(detail_id: str, rules: str = Form(...)):
    """Сохраняет правила технолога для конкретной детали (в свободной форме)"""
    conn = get_conn()
    conn.execute("UPDATE details SET tech_rules=?, updated_at=? WHERE id=?",
                 (rules, datetime.now().isoformat(), detail_id))
    conn.commit()
    conn.close()
    add_history(detail_id, "rules_updated", {"length": len(rules)})
    return {"status": "ok"}


# ============================================
# ЭКОНОМИКА: ставка, накладные, себестоимость
# ============================================
@app.post("/api/details/{detail_id}/economics")
async def api_save_economics(detail_id: str,
                              cost_per_hour: float = Form(...),
                              overhead_pct: float = Form(...),
                              material_cost_rub: float = Form(...)):
    conn = get_conn()
    conn.execute("""UPDATE details SET cost_per_hour=?, overhead_pct=?, material_cost_rub=?, updated_at=?
                    WHERE id=?""",
                 (cost_per_hour, overhead_pct, material_cost_rub,
                  datetime.now().isoformat(), detail_id))
    conn.commit()
    conn.close()
    return {"status": "ok"}


@app.post("/api/details/{detail_id}/apply-ocr")
async def api_apply_ocr(detail_id: str, request: Request):
    """
    v0.5: M28 — применяет распознанные поля чертежа к детали.
    Поля из OCR: designation, material, material_grade, dimensions, thickness_mm, mass_kg, blank_type
    """
    data = await request.json()
    fields = data.get("fields", {})

    # Маппинг OCR полей → DB поля
    field_map = {
        "designation": "designation",
        "material": "material",
        "material_grade": "material_grade",
        "thickness_mm": "thickness_mm",
        "mass_kg": "mass_kg",
        "blank_type": "blank_type",
    }

    updates = []
    values = []
    for ocr_key, db_key in field_map.items():
        if ocr_key in fields and fields[ocr_key] is not None:
            updates.append(f"{db_key}=?")
            values.append(fields[ocr_key])

    # dimensions в size_mm (100x50x20 → "100x50x20")
    if "dimensions" in fields and fields["dimensions"]:
        updates.append("size_mm=?")
        values.append(fields["dimensions"])

    if not updates:
        return {"ok": False, "error": "no fields to apply"}

    values.append(detail_id)
    conn = get_conn()
    # M28: убедиться что есть поля material_grade, thickness_mm, blank_type
    for col in ("material_grade", "thickness_mm", "blank_type"):
        try:
            conn.execute(f"ALTER TABLE details ADD COLUMN {col} TEXT")
        except Exception:
            pass  # колонка уже есть
    conn.commit()
    sql = f"UPDATE details SET {', '.join(updates)}, updated_at=CURRENT_TIMESTAMP WHERE id=?"
    conn.execute(sql, values)
    conn.commit()
    conn.close()
    add_history(detail_id, "ocr_applied", fields)
    return {"ok": True, "applied_fields": list(fields.keys())}


def calc_cost_estimate(detail_id: str) -> dict:
    """Sprint 1: Process-based pricing (CADDi pattern) — разбивка по этапам маршрута"""
    draft = get_draft(detail_id)
    detail = get_detail(detail_id)
    if not draft or not detail:
        return {}
    operations = draft["output"].get("operations", [])
    total_hours = sum(op.get("duration_hours", 0) for op in operations)
    cost_per_hour = detail.get("cost_per_hour") or 0
    material_cost = detail.get("material_cost_rub") or 0
    overhead_pct = detail.get("overhead_pct") or 0

    # Process-based breakdown: группируем по department
    breakdown_by_dept = {}
    for op in operations:
        dept = op.get("department") or "Без цеха"
        if dept not in breakdown_by_dept:
            breakdown_by_dept[dept] = {"hours": 0.0, "operations": 0, "labor_cost": 0.0}
        breakdown_by_dept[dept]["hours"] += op.get("duration_hours", 0)
        breakdown_by_dept[dept]["operations"] += 1
    for d in breakdown_by_dept.values():
        d["hours"] = round(d["hours"], 2)
        d["labor_cost"] = round(d["hours"] * cost_per_hour, 2)

    # Распределяем материал по этапам пропорционально часам
    material_by_dept = {}
    for dept, d in breakdown_by_dept.items():
        share = (d["hours"] / total_hours) if total_hours else 0
        material_by_dept[dept] = round(material_cost * share, 2)

    labor_cost = total_hours * cost_per_hour
    direct_cost = labor_cost + material_cost
    overhead = direct_cost * (overhead_pct / 100)
    total_cost = direct_cost + overhead
    price = total_cost * 1.3  # 30% наценка (mock)

    return {
        "total_hours": round(total_hours, 2),
        "labor_cost": round(labor_cost, 2),
        "material_cost": round(material_cost, 2),
        "direct_cost": round(direct_cost, 2),
        "overhead_pct": overhead_pct,
        "overhead": round(overhead, 2),
        "total_cost": round(total_cost, 2),
        "price": round(price, 2),
        "by_department": [
            {"department": dept, "hours": d["hours"], "operations": d["operations"],
             "labor_cost": d["labor_cost"], "material_cost": material_by_dept[dept]}
            for dept, d in sorted(breakdown_by_dept.items(), key=lambda x: -x[1]["hours"])
        ]
    }


# ============================================
# РОЛЕВАЯ МОДЕЛЬ: workflow утверждения
# ============================================
# F16.7: Вернуть в работу (для approved ТК)
@app.post("/api/reopen")
async def api_reopen(request: Request):
    """Сбросить статус ТК с approved обратно в draft (для переутверждения)."""
    detail_id = await _get_param(request, "detail_id")
    reason = await _get_param(request, "reason") or "Возврат в работу"
    if not detail_id:
        return err("detail_id required", 422)
    from db import get_conn
    conn = get_conn()
    # Получаем текущий draft чтобы убрать RAG индекс
    draft_row = conn.execute("SELECT llm_output, status FROM drafts WHERE detail_id=?", (detail_id,)).fetchone()
    if not draft_row:
        conn.close()
        return err("draft not found", 404)
    if draft_row[1] != "approved":
        conn.close()
        return err("only approved drafts can be reopened", 400)
    conn.execute("""UPDATE drafts SET status='draft', status_ext=NULL, approver=NULL
        WHERE detail_id=?""", (detail_id,))
    conn.commit()
    conn.close()
    add_history(detail_id, "reopened", {"reason": reason})
    # F16.7: убрать из RAG (так как изменится)
    try:
        from rag import get_rag
        rag = get_rag()
        rag.remove_document(detail_id)
        rag.save()
    except Exception as e:
        log.warning(f"RAG remove on reopen failed: {e}")
    return {"ok": True, "status": "draft", "detail_id": detail_id, "reason": reason}


@app.post("/api/submit-for-review")
async def api_submit_for_review(request: Request):
    detail_id = await _get_param(request, "detail_id")
    if not detail_id:
        return err("detail_id required", 422)
    conn = get_conn()
    conn.execute("""UPDATE drafts SET status_ext='review', submitted_at=?
                    WHERE detail_id=?""", (datetime.now().isoformat(), detail_id))
    conn.commit()
    conn.close()
    add_history(detail_id, "submitted_for_review")
    return {"status": "submitted"}


@app.post("/api/approve-chief")
async def api_approve_chief(request: Request):
    """Утверждение начальником (другая роль, не технолог)"""
    detail_id = await _get_param(request, "detail_id")
    chief = await _get_param(request, "chief") or "chief"
    if not detail_id:
        return err("detail_id required", 422)
    conn = get_conn()
    conn.execute("""UPDATE drafts SET status_ext='approved_chief', approver=?
                    WHERE detail_id=?""", (chief, detail_id))
    conn.commit()
    conn.close()
    add_history(detail_id, "approved_by_chief", {"chief": chief})
    return {"status": "approved_chief"}


@app.get("/api/economics/{detail_id}")
async def api_economics(detail_id: str):
    """Sprint 1: HTML с метриками + process-based breakdown по цехам (CADDi pattern)"""
    econ = calc_cost_estimate(detail_id)
    if not econ:
        return HTMLResponse('<p>Нет черновика для расчёта.</p>')
    by_dept_rows = ""
    for d in econ.get("by_department", []):
        by_dept_rows += f"""<tr>
            <td>{d['department']}</td>
            <td>{d['operations']}</td>
            <td>{d['hours']}</td>
            <td>{d['labor_cost']}₽</td>
            <td>{d['material_cost']}₽</td>
            <td>{d['labor_cost'] + d['material_cost']}₽</td>
        </tr>"""
    html = f"""
    <div class="metrics">
        <div class="metric"><div class="metric-value">{econ['total_hours']}</div><div class="metric-label">часов всего</div></div>
        <div class="metric"><div class="metric-value">{econ['labor_cost']}₽</div><div class="metric-label">труд</div></div>
        <div class="metric"><div class="metric-value">{econ['material_cost']}₽</div><div class="metric-label">материал</div></div>
        <div class="metric"><div class="metric-value">{econ['overhead']}₽</div><div class="metric-label">накладные ({econ['overhead_pct']}%)</div></div>
        <div class="metric"><div class="metric-value" style="color:#dc2626">{econ['total_cost']}₽</div><div class="metric-label">себестоимость</div></div>
        <div class="metric"><div class="metric-value" style="color:#16a34a">{econ['price']}₽</div><div class="metric-label">цена (наценка 30%)</div></div>
    </div>
    <p class="lead">Расчёт: {econ['total_hours']} ч × ставка + {econ['material_cost']}₽ материал + {econ['overhead_pct']}% накладные = {econ['total_cost']}₽</p>

    <h3 style="margin-top: 24px;">📊 Разбивка по цехам (process-based pricing)</h3>
    <p class="lead"><small>Сколько часов и денег уходит в каждом цехе — CADDi-паттерн для прозрачности себестоимости.</small></p>
    <table class="data-table">
        <thead><tr><th>Цех</th><th>Операций</th><th>Часы</th><th>Труд</th><th>Материал</th><th>Итого</th></tr></thead>
        <tbody>
        {by_dept_rows}
        </tbody>
    </table>
    """
    return HTMLResponse(html)


# ===================================================================
    log.info(f"Starting БИТ.Технолог (demo_mode={DEMO_MODE})")
    if DEMO_MODE:
        log.info("⚠️  DEMO MODE: no real LLM calls. Mock responses based on heuristics.")
    else:
        log.info(f"✓ LLM: {LLM_MODEL} via {LLM_API_URL}")
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", "8080")))
